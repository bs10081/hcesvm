"""Tests for Boston H1/H2 merge finalization."""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from examples import run_teaching_data_hcesvm_full as runner
from scripts import finalize_boston_h1h2_merge as merge_tool


def _split() -> runner.FullDatasetSplit:
    train_classes = [
        np.asarray([[float(class_index - 1), 0.0], [float(class_index - 1), 1.0]], dtype=float)
        for class_index in range(1, 6)
    ]
    test_classes = [
        np.asarray([[float(class_index - 1), 0.5]], dtype=float)
        for class_index in range(1, 6)
    ]
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [np.full(len(X_class), class_index, dtype=int) for class_index, X_class in enumerate(train_classes, 1)]
    )
    X_test = np.vstack(test_classes)
    y_test = np.asarray([1, 2, 3, 4, 5], dtype=int)
    return runner.FullDatasetSplit(
        name="bostonhousing_ord",
        source_url="https://example.test/bostonhousing_ord.csv",
        feature_names=["feature_1", "feature_2"],
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        split_rule=runner.TEACHING_CSV_SPLIT_RULE,
        random_state=runner.RANDOM_STATE,
    )


def _metadata(split: runner.FullDatasetSplit, *, overrides: dict[str, object] | None = None) -> dict[str, object]:
    metadata: dict[str, object] = {
        "generated_at_utc": "2026-05-18 00:00:00 UTC",
        "branch": "test",
        "commit": "abc123",
        "worktree": "/tmp/hcesvm",
        "log_path": "/tmp/log",
        "excel_path": "/tmp/workbook.xlsx",
        "report_path": "/tmp/report.md",
        "dataset_base_url": runner.DATASET_BASE_URL,
        "datasets": split.name,
        "split_rule": "per-dataset; see <dataset>_split_rule",
        "hcesvm_time_limit_seconds": None,
        "hcesvm_time_limit_label": "none",
        "max_classifiers_per_dataset": 2,
        "threads": 0,
        "soft_mem_limit_gb": 27,
        "mip_gap": 1e-4,
        "mip_focus": 3,
        "nodefile_start_gb": 2,
        "nodefile_dir_requested": "/tmp/nodefiles",
        "nodefile_dir": "/tmp/nodefiles",
        "heartbeat_interval_seconds": 60,
        "feat_lower_bound": 1e-7,
        f"{split.name}_source_url": split.source_url,
        f"{split.name}_source_type": "teaching_csv",
        f"{split.name}_split_rule": split.split_rule,
        f"{split.name}_random_state": split.random_state,
        f"{split.name}_train_counts": json.dumps(split.train_counts),
        f"{split.name}_test_counts": json.dumps(split.test_counts),
        f"{split.name}_train_row_fingerprint": runner.row_fingerprint(split.X_train, split.y_train),
        f"{split.name}_test_row_fingerprint": runner.row_fingerprint(split.X_test, split.y_test),
        f"{split.name}_expected_classifiers": split.n_classes - 1,
    }
    if overrides:
        metadata.update(overrides)
    return metadata


def _parameter_row(split: runner.FullDatasetSplit, hk: int, *, status: str = "optimal") -> dict[str, object]:
    train_counts = split.train_counts
    return {
        "dataset": split.name,
        "classifier": f"H{hk}",
        "description": f"Class {{{', '.join(str(index) for index in range(1, hk + 1))}}} (+1)",
        "positive_sample_count": sum(train_counts[:hk]),
        "negative_sample_count": sum(train_counts[hk:]),
        "weights": json.dumps([-1.0, 0.0]),
        "b": hk - 0.5,
        "objective_value": float(hk),
        "mip_gap": 0.0,
        "solver_status_code": 2 if status == "optimal" else 17,
        "solver_status_label": status,
        "mem_used_gb": 0.1 * hk,
        "max_mem_used_gb": 0.2 * hk,
        "elapsed_seconds": float(hk),
        "cumulative_elapsed_seconds": float(hk),
    }


def _progress_row(row: dict[str, object]) -> dict[str, object]:
    return {
        **row,
        "started_at_utc": "2026-05-18 00:00:00 UTC",
        "finished_at_utc": "2026-05-18 00:00:01 UTC",
    }


def _write_workbook(
    path: Path,
    split: runner.FullDatasetSplit,
    classifiers: tuple[int, ...],
    *,
    metadata_overrides: dict[str, object] | None = None,
    status_overrides: dict[int, str] | None = None,
) -> None:
    status_overrides = status_overrides or {}
    parameter_rows = [
        _parameter_row(split, hk, status=status_overrides.get(hk, "optimal"))
        for hk in classifiers
    ]
    metric_headers = runner.build_metric_headers(split.n_classes)
    parameter_headers = runner.build_parameter_headers()
    progress_headers = runner.build_progress_headers()
    runner.write_excel(
        path,
        metric_headers=metric_headers,
        metric_rows=[
            {
                "dataset": split.name,
                "model": "HCESVM(test3)",
                "status": "stopped_early" if max(classifiers) < 4 else "completed",
                "stop_reason": None,
                "n_classes": split.n_classes,
                "n_features": split.n_features,
                "fit_seconds": None,
                "train_total_accuracy": None,
                "test_total_accuracy": None,
                "source_url": split.source_url,
            }
        ],
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        progress_headers=progress_headers,
        progress_rows=[_progress_row(row) for row in parameter_rows],
        metadata_rows=runner.build_metadata_rows(_metadata(split, overrides=metadata_overrides)),
    )


def test_finalize_merge_recomputes_accuracy_from_merged_weights(tmp_path, monkeypatch) -> None:
    split = _split()
    h1_h2 = tmp_path / "partial.xlsx"
    h3_h4 = tmp_path / "old.xlsx"
    _write_workbook(h1_h2, split, (1, 2))
    _write_workbook(h3_h4, split, (3, 4))
    monkeypatch.setattr(merge_tool, "git_output_safe", lambda *args: "fake")

    outputs = merge_tool.finalize_merge(
        h1_h2_workbook=h1_h2,
        h3_h4_workbook=h3_h4,
        output_root=tmp_path,
        timestamp="20260518_010203",
        split=split,
    )

    assert outputs["workbook"].is_file()
    assert outputs["report"].is_file()
    assert outputs["log"].is_file()

    workbook = load_workbook(outputs["workbook"], read_only=True, data_only=True)
    metric_rows = list(workbook["Metrics"].iter_rows(values_only=True))
    metric = dict(zip(metric_rows[0], metric_rows[1]))
    assert metric["train_total_accuracy"] == 1.0
    assert metric["test_total_accuracy"] == 1.0
    assert metric["train_class_5_accuracy"] == 1.0

    parameter_rows = list(workbook["Parameters"].iter_rows(values_only=True))
    headers = parameter_rows[0]
    h1 = dict(zip(headers, parameter_rows[1]))
    h4 = dict(zip(headers, parameter_rows[4]))
    assert h1["classifier"] == "H1"
    assert h1["b"] == 0.5
    assert h4["classifier"] == "H4"
    assert h4["b"] == 3.5

    metadata = {key: value for key, value in workbook["Run_Metadata"].iter_rows(values_only=True) if key != "key"}
    assert metadata["h1_h2_source_workbook"] == str(h1_h2)
    assert metadata["h3_h4_source_workbook"] == str(h3_h4)
    assert metadata["mip_focus"] == 3


def test_finalize_merge_rejects_split_metadata_mismatch(tmp_path) -> None:
    split = _split()
    h1_h2 = tmp_path / "partial.xlsx"
    h3_h4 = tmp_path / "old.xlsx"
    _write_workbook(h1_h2, split, (1, 2))
    _write_workbook(
        h3_h4,
        split,
        (3, 4),
        metadata_overrides={f"{split.name}_train_counts": json.dumps([99, 2, 2, 2, 2])},
    )

    with pytest.raises(ValueError, match="Split metadata mismatch"):
        merge_tool.finalize_merge(
            h1_h2_workbook=h1_h2,
            h3_h4_workbook=h3_h4,
            output_root=tmp_path,
            timestamp="20260518_010203",
            split=split,
        )


def test_finalize_merge_rejects_non_optimal_h1_h2(tmp_path) -> None:
    split = _split()
    h1_h2 = tmp_path / "partial.xlsx"
    h3_h4 = tmp_path / "old.xlsx"
    _write_workbook(h1_h2, split, (1, 2), status_overrides={2: "mem_limit_with_solution"})
    _write_workbook(h3_h4, split, (3, 4))

    with pytest.raises(ValueError, match="H1/H2 workbook H2 must be optimal"):
        merge_tool.finalize_merge(
            h1_h2_workbook=h1_h2,
            h3_h4_workbook=h3_h4,
            output_root=tmp_path,
            timestamp="20260518_010203",
            split=split,
        )
