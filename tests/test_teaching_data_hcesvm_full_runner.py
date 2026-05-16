"""Tests for the full teaching-data HCESVM-only runner."""

from __future__ import annotations

import numpy as np
from openpyxl import load_workbook
import pytest

from examples import run_teaching_data_hcesvm_full as runner


class FakeClassifier:
    def __init__(self, hk: int) -> None:
        self.weights = np.asarray([0.1 * hk, -0.2 * hk], dtype=float)
        self.intercept = 0.5 * hk
        self.solution = {
            "objective_value": 10.0 * hk,
            "mip_gap": 0.0,
            "solver_status": 2,
            "mem_used_gb": 0.25 * hk,
            "max_mem_used_gb": 0.5 * hk,
        }


class FakeIncrementalHierarchicalCESVM:
    last_init: dict[str, object] | None = None

    def __init__(self, *, cesvm_params, strategy, n_classes) -> None:
        type(self).last_init = {
            "cesvm_params": dict(cesvm_params),
            "strategy": strategy,
            "n_classes": n_classes,
        }
        self.classifiers = {}
        self.h1 = None
        self.h2 = None
        self.expected_classifier_count = n_classes - 1
        self.completed_classifier_count = 0
        self.fit_stop_reason = None
        self.fit_stopped_early = False

    def fit_incremental(self, *train_classes, after_classifier=None) -> None:
        for hk in range(1, self.expected_classifier_count + 1):
            classifier = FakeClassifier(hk)
            self.classifiers[f"h{hk}"] = classifier
            self.completed_classifier_count = hk
            progress = {
                "hk": hk,
                "description": f"Fake H{hk}",
                "started_at_utc": runner.utc_now(),
                "finished_at_utc": runner.utc_now(),
                "elapsed_seconds": float(hk),
                "cumulative_elapsed_seconds": float(sum(range(1, hk + 1))),
                "positive_sample_count": 10 * hk,
                "negative_sample_count": 20 * hk,
            }
            if after_classifier is not None and after_classifier(progress) is False:
                self.fit_stopped_early = True
                self.fit_stop_reason = f"requested stop after h{hk}"
                return

    def is_fully_fitted(self) -> bool:
        return self.completed_classifier_count == self.expected_classifier_count and not self.fit_stopped_early

    def predict(self, X):
        return np.ones(len(X), dtype=int)


def fake_split() -> runner.FullDatasetSplit:
    train_classes = [
        np.asarray([[1.0, 1.1], [1.2, 1.3]], dtype=float),
        np.asarray([[2.0, 2.1], [2.2, 2.3]], dtype=float),
        np.asarray([[3.0, 3.1], [3.2, 3.3]], dtype=float),
        np.asarray([[4.0, 4.1], [4.2, 4.3]], dtype=float),
        np.asarray([[5.0, 5.1], [5.2, 5.3]], dtype=float),
    ]
    test_classes = [
        np.asarray([[1.5, 1.6]], dtype=float),
        np.asarray([[2.5, 2.6]], dtype=float),
        np.asarray([[3.5, 3.6]], dtype=float),
        np.asarray([[4.5, 4.6]], dtype=float),
        np.asarray([[5.5, 5.6]], dtype=float),
    ]
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [np.full(len(X_class), class_index, dtype=int) for class_index, X_class in enumerate(train_classes, 1)]
    )
    X_test = np.vstack(test_classes)
    y_test = np.asarray([1, 2, 3, 4, 5], dtype=int)
    return runner.FullDatasetSplit(
        name="bostonhousing_ord",
        source_url="https://example.test/bostonhousing_ord.csv",
        feature_names=["feature_1", "feature_2"],
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        split_rule=runner.TEACHING_CSV_SPLIT_RULE,
        random_state=runner.RANDOM_STATE,
    )


def test_parse_arguments_defaults_to_boston_with_5400s_time_limit() -> None:
    args = runner.parse_arguments([])

    assert args.datasets == ["bostonhousing_ord"]
    assert args.time_limit == 5400


def test_parse_arguments_accepts_no_time_limit() -> None:
    args = runner.parse_arguments(["--time-limit", "none"])

    assert args.time_limit is None


def test_parse_arguments_accepts_nodefile_controls() -> None:
    args = runner.parse_arguments(
        [
            "--nodefile-start-gb",
            "0.5",
            "--nodefile-dir",
            "~/hcesvm-gurobi-nodefiles/cement",
        ]
    )

    assert args.nodefile_start_gb == 0.5
    assert args.nodefile_dir == "~/hcesvm-gurobi-nodefiles/cement"


def test_resolve_nodefile_dir_only_creates_directory_when_enabled(tmp_path) -> None:
    disabled = runner.resolve_nodefile_dir(
        nodefile_start_gb=None,
        nodefile_dir_option=str(tmp_path / "disabled"),
        dataset_slug="cement_strength",
        timestamp="20260516_000000",
    )
    enabled = runner.resolve_nodefile_dir(
        nodefile_start_gb=0.5,
        nodefile_dir_option=str(tmp_path / "enabled"),
        dataset_slug="cement_strength",
        timestamp="20260516_000000",
    )

    assert disabled is None
    assert not (tmp_path / "disabled").exists()
    assert enabled == str(tmp_path / "enabled")
    assert (tmp_path / "enabled").is_dir()


def test_preflight_reports_missing_dataset_sources(monkeypatch) -> None:
    monkeypatch.setattr(
        runner,
        "PENDING_DATASET_PATHS",
        {"hayes_roth": [runner.Path("/missing/hayes-roth_split.xlsx")]},
    )

    with pytest.raises(FileNotFoundError) as exc_info:
        runner.validate_dataset_sources(["hayes_roth"])

    assert "Dataset source preflight failed" in str(exc_info.value)
    assert "hayes_roth" in str(exc_info.value)
    assert "/missing/hayes-roth_split.xlsx" in str(exc_info.value)


def test_lingo_workbook_source_preserves_train_test_sheets(monkeypatch, tmp_path) -> None:
    class FakeOrdinalDataset:
        workbook_path = tmp_path / "balance.xlsx"
        feature_names = ["f1", "f2"]
        X_train = [[float(index), float(index + 1)] for index in range(15)]
        X_test = [[float(index), float(index + 1)] for index in range(15, 30)]
        y_train = [1] * 5 + [2] * 5 + [3] * 5
        y_test = [1] * 5 + [2] * 5 + [3] * 5

    monkeypatch.setattr(runner, "load_lingo_split_workbook", lambda path: FakeOrdinalDataset())

    split = runner.load_lingo_workbook_preserve_split_dataset("balance", str(tmp_path / "balance.xlsx"))

    assert split.name == "balance"
    assert split.n_classes == 3
    assert split.train_counts == [5, 5, 5]
    assert split.test_counts == [5, 5, 5]
    assert split.n_features == 2
    assert split.split_rule == runner.WORKBOOK_SPLIT_RULE
    assert split.random_state is None
    np.testing.assert_allclose(split.X_train, np.asarray(FakeOrdinalDataset.X_train, dtype=float))
    np.testing.assert_allclose(split.X_test, np.asarray(FakeOrdinalDataset.X_test, dtype=float))


def test_full_runner_writes_accuracy_and_classifier_diagnostics(monkeypatch, tmp_path) -> None:
    split = fake_split()

    monkeypatch.setattr(runner, "ROOT", tmp_path)
    monkeypatch.setattr(runner, "HierarchicalCESVM", FakeIncrementalHierarchicalCESVM)
    monkeypatch.setattr(runner, "git_output", lambda *args: "fake")
    monkeypatch.setattr(
        runner,
        "validate_dataset_sources",
        lambda names: [
            runner.DatasetSourceSpec(
                name="bostonhousing_ord",
                source_type="teaching_csv",
                source="https://example.test/bostonhousing_ord.csv",
            )
        ],
    )
    monkeypatch.setattr(runner, "load_dataset_from_spec", lambda spec: split)
    monkeypatch.setattr(
        runner,
        "evaluate_multiclass",
        lambda y_true, y_pred, *, n_classes: {
            "total_accuracy": 0.2,
            **{f"class_{index}_accuracy": (1.0 if index == 1 else 0.0) for index in range(1, n_classes + 1)},
            **{f"class_{index}_count": int(np.sum(np.asarray(y_true) == index)) for index in range(1, n_classes + 1)},
        },
    )

    exit_code = runner.main(
        [
            "--datasets",
            "bostonhousing_ord",
            "--time-limit",
            "5400",
            "--threads",
            "20",
            "--soft-mem-limit-gb",
            "56",
            "--heartbeat-interval-seconds",
            "none",
        ]
    )

    assert exit_code == 0
    assert FakeIncrementalHierarchicalCESVM.last_init is not None
    assert FakeIncrementalHierarchicalCESVM.last_init["strategy"] == "test3"
    assert FakeIncrementalHierarchicalCESVM.last_init["n_classes"] == 5
    assert FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]["time_limit"] == 5400
    assert FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]["threads"] == 20
    assert FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]["soft_mem_limit_gb"] == 56
    assert "nodefile_start" not in FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]
    assert "nodefile_dir" not in FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]

    workbook_path = next((tmp_path / "docs" / "reports").glob("BOSTONHOUSING_ORD_HCESVM_TEST3_FULL_*.xlsx"))
    report_path = next((tmp_path / "docs" / "reports").glob("BOSTONHOUSING_ORD_HCESVM_TEST3_FULL_*.md"))

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert set(workbook.sheetnames) == {"Metrics", "Parameters", "Progress", "Run_Metadata"}

    metrics_rows = list(workbook["Metrics"].iter_rows(values_only=True))
    metric_headers = metrics_rows[0]
    metric_row = dict(zip(metric_headers, metrics_rows[1]))
    assert metric_row["dataset"] == "bostonhousing_ord"
    assert metric_row["status"] == "completed"
    assert metric_row["train_total_accuracy"] == 0.2
    assert metric_row["test_total_accuracy"] == 0.2
    assert metric_row["train_class_1_accuracy"] == 1.0
    assert metric_row["test_class_5_accuracy"] == 0.0

    parameter_rows = list(workbook["Parameters"].iter_rows(values_only=True))
    parameter_headers = parameter_rows[0]
    h4_row = dict(zip(parameter_headers, parameter_rows[4]))
    assert h4_row["classifier"] == "H4"
    assert h4_row["solver_status_code"] == 2
    assert h4_row["solver_status_label"] == "optimal"
    assert h4_row["mem_used_gb"] == 1.0
    assert h4_row["max_mem_used_gb"] == 2.0
    assert h4_row["weights"] == "[0.4, -0.8]"
    assert h4_row["b"] == 2.0

    metadata_rows = list(workbook["Run_Metadata"].iter_rows(values_only=True))
    metadata = {key: value for key, value in metadata_rows[1:]}
    assert metadata["hcesvm_time_limit_seconds"] == 5400
    assert metadata["threads"] == 20
    assert metadata["soft_mem_limit_gb"] == 56
    assert metadata["nodefile_start_gb"] is None
    assert metadata["nodefile_dir"] is None
    assert metadata["bostonhousing_ord_expected_classifiers"] == 4
    assert metadata["bostonhousing_ord_split_rule"] == runner.TEACHING_CSV_SPLIT_RULE
    assert metadata["bostonhousing_ord_random_state"] == runner.RANDOM_STATE
    assert len(metadata["bostonhousing_ord_train_row_fingerprint"]) == 64
    assert len(metadata["bostonhousing_ord_test_row_fingerprint"]) == 64

    markdown = report_path.read_text(encoding="utf-8")
    assert "## bostonhousing_ord" in markdown
    assert "- Threads: `20`" in markdown
    assert "- SoftMemLimit: `56.0 GB`" in markdown
    assert "- NodeFileStart: `None`" in markdown
    assert f"- Split rule: `{runner.TEACHING_CSV_SPLIT_RULE}`" in markdown
    assert "| Train | 0.2000 | C1=1.0000" in markdown
    assert "| H4 | Fake H4 | 40/80 | optimal | 4.00 | 40.000000 | 0.000000 | 1.000000 | 2.000000" in markdown


def test_full_runner_passes_no_time_limit_to_model(monkeypatch, tmp_path) -> None:
    split = fake_split()

    monkeypatch.setattr(runner, "ROOT", tmp_path)
    monkeypatch.setattr(runner, "HierarchicalCESVM", FakeIncrementalHierarchicalCESVM)
    monkeypatch.setattr(runner, "git_output", lambda *args: "fake")
    monkeypatch.setattr(
        runner,
        "validate_dataset_sources",
        lambda names: [
            runner.DatasetSourceSpec(
                name="bostonhousing_ord",
                source_type="teaching_csv",
                source="https://example.test/bostonhousing_ord.csv",
            )
        ],
    )
    monkeypatch.setattr(runner, "load_dataset_from_spec", lambda spec: split)
    monkeypatch.setattr(
        runner,
        "evaluate_multiclass",
        lambda y_true, y_pred, *, n_classes: {
            "total_accuracy": 0.2,
            **{f"class_{index}_accuracy": (1.0 if index == 1 else 0.0) for index in range(1, n_classes + 1)},
            **{f"class_{index}_count": int(np.sum(np.asarray(y_true) == index)) for index in range(1, n_classes + 1)},
        },
    )

    exit_code = runner.main(
        [
            "--datasets",
            "bostonhousing_ord",
            "--time-limit",
            "none",
            "--threads",
            "2",
            "--soft-mem-limit-gb",
            "18",
            "--heartbeat-interval-seconds",
            "none",
        ]
    )

    assert exit_code == 0
    assert FakeIncrementalHierarchicalCESVM.last_init is not None
    assert FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]["time_limit"] is None

    workbook_path = next((tmp_path / "docs" / "reports").glob("BOSTONHOUSING_ORD_HCESVM_TEST3_FULL_*.xlsx"))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    metadata_rows = list(workbook["Run_Metadata"].iter_rows(values_only=True))
    metadata = {key: value for key, value in metadata_rows[1:]}
    assert metadata["hcesvm_time_limit_seconds"] is None
    assert metadata["hcesvm_time_limit_label"] == "none"


def test_full_runner_passes_nodefile_controls_to_model(monkeypatch, tmp_path) -> None:
    split = fake_split()
    nodefile_dir = tmp_path / "nodefiles"

    monkeypatch.setattr(runner, "ROOT", tmp_path)
    monkeypatch.setattr(runner, "HierarchicalCESVM", FakeIncrementalHierarchicalCESVM)
    monkeypatch.setattr(runner, "git_output", lambda *args: "fake")
    monkeypatch.setattr(
        runner,
        "validate_dataset_sources",
        lambda names: [
            runner.DatasetSourceSpec(
                name="bostonhousing_ord",
                source_type="teaching_csv",
                source="https://example.test/bostonhousing_ord.csv",
            )
        ],
    )
    monkeypatch.setattr(runner, "load_dataset_from_spec", lambda spec: split)
    monkeypatch.setattr(
        runner,
        "evaluate_multiclass",
        lambda y_true, y_pred, *, n_classes: {
            "total_accuracy": 0.2,
            **{f"class_{index}_accuracy": (1.0 if index == 1 else 0.0) for index in range(1, n_classes + 1)},
            **{f"class_{index}_count": int(np.sum(np.asarray(y_true) == index)) for index in range(1, n_classes + 1)},
        },
    )

    exit_code = runner.main(
        [
            "--datasets",
            "bostonhousing_ord",
            "--time-limit",
            "none",
            "--threads",
            "0",
            "--soft-mem-limit-gb",
            "56",
            "--nodefile-start-gb",
            "0.5",
            "--nodefile-dir",
            str(nodefile_dir),
            "--heartbeat-interval-seconds",
            "none",
        ]
    )

    assert exit_code == 0
    assert FakeIncrementalHierarchicalCESVM.last_init is not None
    cesvm_params = FakeIncrementalHierarchicalCESVM.last_init["cesvm_params"]
    assert cesvm_params["time_limit"] is None
    assert cesvm_params["threads"] == 0
    assert cesvm_params["nodefile_start"] == 0.5
    assert cesvm_params["nodefile_dir"] == str(nodefile_dir)
    assert nodefile_dir.is_dir()

    workbook_path = next((tmp_path / "docs" / "reports").glob("BOSTONHOUSING_ORD_HCESVM_TEST3_FULL_*.xlsx"))
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    metadata_rows = list(workbook["Run_Metadata"].iter_rows(values_only=True))
    metadata = {key: value for key, value in metadata_rows[1:]}
    assert metadata["hcesvm_time_limit_label"] == "none"
    assert metadata["nodefile_start_gb"] == 0.5
    assert metadata["nodefile_dir"] == str(nodefile_dir)
