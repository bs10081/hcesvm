"""Tests for the exact skill 1000-sample SVOR/NPSVOR baseline runner."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from hcesvm import skill_1000_baselines_runner as runner


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def make_manifest(tmp_path: Path) -> Path:
    train_csv = tmp_path / "skill_method3_4class_1000_train.csv"
    test_csv = tmp_path / "skill_method3_4class_1000_test.csv"
    write_text(train_csv, "feature_1,feature_2,response\n1.0,2.0,1\n")
    write_text(test_csv, "feature_1,feature_2,response\n1.5,2.5,1\n")

    manifest = {
        "dataset": "skill_method3_4class_1000",
        "source_dataset": "skill",
        "source_url": "https://example.test/skill.csv",
        "description": "Exact archived split for skill method3 4-class 1000.",
        "class_map": {"4": 1, "5": 2, "6": 3, "7": 4},
        "train_original_counts": [649, 642, 497, 28],
        "test_original_counts": [162, 161, 124, 7],
        "train_sampled_counts": [286, 283, 219, 12],
        "test_sampled_counts": [71, 71, 55, 3],
        "train_target_size": 800,
        "test_target_size": 200,
        "feature_names": ["feature_1", "feature_2"],
        "train_csv": str(train_csv),
        "test_csv": str(test_csv),
    }
    manifest_path = tmp_path / "skill_method3_4class_1000_manifest_20260509_055457.json"
    write_text(manifest_path, json.dumps(manifest, ensure_ascii=False, indent=2))
    return manifest_path


def test_load_manifest_split_reads_expected_counts_and_csvs(tmp_path: Path) -> None:
    manifest_path = make_manifest(tmp_path)

    split = runner.load_manifest_split(manifest_path)

    assert split.dataset == "skill_method3_4class_1000"
    assert split.source_dataset == "skill"
    assert split.train_csv == (tmp_path / "skill_method3_4class_1000_train.csv").resolve()
    assert split.test_csv == (tmp_path / "skill_method3_4class_1000_test.csv").resolve()
    assert split.train_sampled_counts == [286, 283, 219, 12]
    assert split.test_sampled_counts == [71, 71, 55, 3]
    assert split.relabel_summary == "4->1, 5->2, 6->3, 7->4"


def test_run_manifest_experiment_runs_svor_then_npsvor_and_writes_artifacts(
    tmp_path: Path,
    monkeypatch,
) -> None:
    manifest_path = make_manifest(tmp_path)
    artifact_paths = runner.ArtifactPaths(
        log_path=tmp_path / "results" / "run.log",
        workbook_path=tmp_path / "docs" / "run.xlsx",
        markdown_path=tmp_path / "docs" / "run.md",
    )
    call_order: list[str] = []

    split = runner.load_manifest_split(manifest_path)
    synthetic_bundle = runner.DatasetBundle(
        split=split,
        dataset=runner.OrdinalDatasetSplit(
            workbook_path=split.train_csv,
            feature_names=["feature_1", "feature_2"],
            X_train=[[1.0, 2.0], [2.0, 3.0], [3.0, 4.0], [4.0, 5.0]],
            y_train=[1, 2, 3, 4],
            X_test=[[1.5, 2.5], [2.5, 3.5], [3.5, 4.5], [4.5, 5.5]],
            y_test=[1, 2, 3, 4],
            class_token_to_label={"1": 1, "2": 2, "3": 3, "4": 4},
        ),
        X_train=runner.np.asarray([[1.0, 2.0], [2.0, 3.0], [3.0, 4.0], [4.0, 5.0]], dtype=float),
        y_train=runner.np.asarray([1, 2, 3, 4], dtype=int),
        X_test=runner.np.asarray([[1.5, 2.5], [2.5, 3.5], [3.5, 4.5], [4.5, 5.5]], dtype=float),
        y_test=runner.np.asarray([1, 2, 3, 4], dtype=int),
    )

    hcesvm_metrics = {
        "dataset": "skill_method3_4class_1000",
        "model": "HCESVM(test3)",
        "train_total_accuracy": 0.35375,
        "test_total_accuracy": 0.355,
        "train_class_4_accuracy": 0.0,
        "test_class_4_accuracy": 0.0,
        "workbook_path": str(tmp_path / "docs" / "hcesvm.xlsx"),
    }

    def fake_result(model_name: str, total_accuracy: float) -> runner.ModelRunResult:
        train_metrics = {
            "total_accuracy": total_accuracy,
            "class_1_accuracy": 0.0,
            "class_2_accuracy": 1.0,
            "class_3_accuracy": 0.0,
            "class_4_accuracy": 0.0,
            "class_1_count": 286,
            "class_2_count": 283,
            "class_3_count": 219,
            "class_4_count": 12,
        }
        test_metrics = {
            "total_accuracy": total_accuracy,
            "class_1_accuracy": 0.0,
            "class_2_accuracy": 1.0,
            "class_3_accuracy": 0.0,
            "class_4_accuracy": 0.0,
            "class_1_count": 71,
            "class_2_count": 71,
            "class_3_count": 55,
            "class_4_count": 3,
        }
        component = "GLOBAL" if model_name == "SVOR" else "RANK_1"
        return runner.ModelRunResult(
            dataset="skill_method3_4class_1000",
            model=model_name,
            status="completed",
            fit_seconds=12.5,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=[
                {
                    "dataset": "skill_method3_4class_1000",
                    "source_dataset": "skill",
                    "model": model_name,
                    "component": component,
                    "description": "stub",
                    "status": "completed",
                    "fit_seconds": 12.5,
                    "solver_status_code": 2,
                    "solver_status_label": "optimal",
                    "weights": "[0.1, -0.2]",
                    "b": "[0.3, 0.4]" if model_name == "SVOR" else 0.3,
                    "objective_value": 1.23,
                    "prediction_rule": "min_distance",
                }
            ],
        )

    def fake_run_svor(bundle, tee, params):
        call_order.append("SVOR")
        assert bundle.split.train_sampled_counts == [286, 283, 219, 12]
        assert params["solver_params"]["TimeLimit"] == 1800
        return fake_result("SVOR", 0.36)

    def fake_run_npsvor(bundle, tee, params):
        call_order.append("NPSVOR")
        assert bundle.split.test_sampled_counts == [71, 71, 55, 3]
        assert params["solver_params"]["TimeLimit"] == 1800
        return fake_result("NPSVOR", 0.37)

    monkeypatch.setattr(runner, "build_artifact_paths", lambda started_at: artifact_paths)
    monkeypatch.setattr(runner, "git_output", lambda *args: "stub-value")
    monkeypatch.setattr(runner, "load_bundle", lambda split: synthetic_bundle)
    monkeypatch.setattr(runner, "load_matching_hcesvm_metrics", lambda split: hcesvm_metrics)
    monkeypatch.setattr(runner, "run_svor", fake_run_svor)
    monkeypatch.setattr(runner, "run_npsvor", fake_run_npsvor)

    outcome = runner.run_manifest_experiment(
        manifest_json=manifest_path,
        svor_time_limit=1800,
        npsvor_time_limit=1800,
        started_at=datetime(2026, 5, 9, 8, 30, 0, tzinfo=timezone.utc),
    )

    assert call_order == ["SVOR", "NPSVOR"]
    assert outcome["artifacts"] == artifact_paths
    assert artifact_paths.log_path.exists()
    assert artifact_paths.workbook_path.exists()
    assert artifact_paths.markdown_path.exists()

    workbook = load_workbook(artifact_paths.workbook_path, read_only=True, data_only=True)
    metrics_rows = list(workbook["Metrics"].iter_rows(values_only=True))
    metrics_headers = metrics_rows[0]
    model_index = metrics_headers.index("model")
    status_index = metrics_headers.index("status")
    metric_models = [row[model_index] for row in metrics_rows[1:]]
    metric_statuses = [row[status_index] for row in metrics_rows[1:]]

    assert metric_models == ["SVOR", "NPSVOR"]
    assert all(status == "completed" for status in metric_statuses)
    assert "HCESVM(test3)" not in metric_models

    metadata_rows = list(workbook["Run_Metadata"].iter_rows(values_only=True))
    metadata = {key: value for key, value in metadata_rows[1:]}
    assert metadata["manifest_json"] == str(manifest_path.resolve())
    assert metadata["train_sampled_counts"] == "[286, 283, 219, 12]"
    assert metadata["test_sampled_counts"] == "[71, 71, 55, 3]"

    markdown = artifact_paths.markdown_path.read_text(encoding="utf-8")
    assert "Train counts: `[649, 642, 497, 28] -> [286, 283, 219, 12]`" in markdown
    assert "Test counts: `[162, 161, 124, 7] -> [71, 71, 55, 3]`" in markdown
    assert "Minority-only explanation still looks plausible on this exact split" in markdown

