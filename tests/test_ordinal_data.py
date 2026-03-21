"""Tests for ordinal dataset loaders."""

from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from hcesvm.utils.ordinal_data import load_lingo_split_workbook, load_tabular_dataset_split


ROOT = Path(__file__).resolve().parents[1]
SVOR_BALANCE = ROOT / "data" / "svor" / "SVOR_balance_split.xlsx"
NPSVOR_BALANCE = ROOT / "data" / "npsvor" / "NPSVOR_balance_split.xlsx"


class WorkbookLoaderTests(unittest.TestCase):
    def test_svor_balance_workbook_loads(self) -> None:
        dataset = load_lingo_split_workbook(SVOR_BALANCE)

        self.assertEqual(dataset.feature_names, ["left-weight", "left-distance", "right-weight", "right-distance"])
        self.assertEqual(len(dataset.X_train), 500)
        self.assertEqual(len(dataset.X_test), 125)
        self.assertEqual(sorted(set(dataset.y_train)), [1, 2, 3])
        self.assertAlmostEqual(dataset.reported_train_accuracy or 0.0, 0.932, places=3)
        self.assertAlmostEqual(dataset.reported_test_accuracy or 0.0, 0.856, places=3)

    def test_npsvor_balance_workbook_loads(self) -> None:
        dataset = load_lingo_split_workbook(NPSVOR_BALANCE)

        self.assertEqual(dataset.feature_names, ["left-weight", "left-distance", "right-weight", "right-distance"])
        self.assertEqual(len(dataset.X_train), 500)
        self.assertEqual(len(dataset.X_test), 125)
        self.assertEqual(sorted(set(dataset.y_train)), [1, 2, 3])
        self.assertAlmostEqual(dataset.reported_train_accuracy or 0.0, 0.932, places=3)
        self.assertAlmostEqual(dataset.reported_test_accuracy or 0.0, 0.856, places=3)

    def test_generic_csv_train_and_test_load(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            train_path = Path(temp_dir) / "train.csv"
            test_path = Path(temp_dir) / "test.csv"
            rows_train = [
                {"f1": 1.0, "f2": 2.0, "note": "a", "label": 10},
                {"f1": 2.0, "f2": 3.0, "note": "b", "label": 20},
                {"f1": 3.0, "f2": 4.0, "note": "c", "label": 30},
            ]
            rows_test = [
                {"f1": 1.5, "f2": 2.5, "note": "x", "label": 10},
                {"f1": 2.5, "f2": 3.5, "note": "y", "label": 20},
            ]

            for path, rows in [(train_path, rows_train), (test_path, rows_test)]:
                with path.open("w", encoding="utf-8", newline="") as handle:
                    writer = csv.DictWriter(handle, fieldnames=["f1", "f2", "note", "label"])
                    writer.writeheader()
                    writer.writerows(rows)

            dataset = load_tabular_dataset_split(train_path, test_path=test_path, target_column="label")
            self.assertEqual(dataset.feature_names, ["f1", "f2"])
            self.assertEqual(dataset.y_train, [10, 20, 30])
            self.assertEqual(dataset.y_test, [10, 20])

    def test_generic_excel_train_test_sheets_load(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "ordinal.xlsx"
            workbook = Workbook()
            train_sheet = workbook.active
            train_sheet.title = "Train"
            test_sheet = workbook.create_sheet("Test")

            for sheet, rows in [
                (
                    train_sheet,
                    [
                        ["feat_a", "feat_b", "text", "target"],
                        [1, 2, "row1", 1],
                        [2, 3, "row2", 3],
                        [3, 4, "row3", 5],
                    ],
                ),
                (
                    test_sheet,
                    [
                        ["feat_a", "feat_b", "text", "target"],
                        [1.5, 2.5, "row4", 1],
                        [2.5, 3.5, "row5", 3],
                    ],
                ),
            ]:
                for row in rows:
                    sheet.append(row)

            workbook.save(workbook_path)

            dataset = load_tabular_dataset_split(workbook_path, target_column="target")
            self.assertEqual(dataset.feature_names, ["feat_a", "feat_b"])
            self.assertEqual(dataset.y_train, [1, 3, 5])
            self.assertEqual(dataset.y_test, [1, 3])

    def test_lingo_loader_handles_shifted_headers_and_numeric_class_column(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "shifted.xlsx"
            workbook = Workbook()
            train_sheet = workbook.active
            train_sheet.title = "Train"
            test_sheet = workbook.create_sheet("Test")

            train_rows = [
                ["Total Accuracy", 0.75],
                ["metadata"],
                [],
                [],
                [],
                ["feat_a", "feat_b", "feat_c", "Class", "Predict", "Actual"],
                [0.0, 0.1, 0.2, 1, 1, 1],
                [1.0, 1.1, 1.2, 2, 2, 2],
                [2.0, 2.1, 2.2, 3, 3, 3],
            ]
            test_rows = [
                ["Total", 1.0],
                [],
                ["stats"],
                [],
                ["feat_a", "feat_b", "feat_c", "k1", "k2", "k3", "Predict", "Class"],
                [0.5, 0.6, 0.7, 1, 0, 0, 1, 1],
                [1.5, 1.6, 1.7, 0, 1, 0, 2, 2],
                [2.5, 2.6, 2.7, 0, 0, 1, 3, 3],
            ]

            for row in train_rows:
                train_sheet.append(row)
            for row in test_rows:
                test_sheet.append(row)

            workbook.save(workbook_path)

            dataset = load_lingo_split_workbook(workbook_path)

            self.assertEqual(dataset.feature_names, ["feat_a", "feat_b", "feat_c"])
            self.assertEqual(dataset.y_train, [1, 2, 3])
            self.assertEqual(dataset.y_test, [1, 2, 3])
            self.assertAlmostEqual(dataset.reported_train_accuracy or 0.0, 0.75, places=6)
            self.assertAlmostEqual(dataset.reported_test_accuracy or 0.0, 1.0, places=6)


if __name__ == "__main__":
    unittest.main()
