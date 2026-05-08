#!/usr/bin/env python3
"""Run teaching-data HCESVM(test3) with a deadline-aware stop policy."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from hcesvm.models.hierarchical import HierarchicalCESVM
from hcesvm.utils.data_loader import load_csv_ordinal_data
from hcesvm.utils.deadline_control import evaluate_deadline_policy, parse_deadline
from hcesvm.utils.evaluator import evaluate_multiclass
from hcesvm.utils.teaching_data_runtime import format_hcesvm_time_limit_message


DATASET_BASE_URL = "https://raw.githubusercontent.com/gagolews/teaching-data/master/ordinal_regression"
DEFAULT_DATASETS = [
    "cement_strength",
    "bostonhousing_ord",
    "californiahousing",
    "skill",
    "stock_ord",
]
DEFAULT_DATASET = "skill"
DEFAULT_DEADLINE_AT = "2026-04-25T12:00:00"
DEFAULT_DEADLINE_TZ = "Asia/Taipei"
TEST_SIZE = 0.2
RANDOM_STATE = 42
MODEL_NAME = "HCESVM(test3)"
DEFAULT_HCESVM_THREADS = 2
DEFAULT_HCESVM_SOFT_MEM_LIMIT_GB = 18.0


class TeeStream:
    """Mirror output to stdout/stderr and the run log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass(slots=True)
class DatasetBundle:
    name: str
    source_url: str
    n_classes: int
    n_features: int
    train_classes: list[np.ndarray]
    test_classes: list[np.ndarray]
    X_train: np.ndarray
    y_train: np.ndarray
    X_test: np.ndarray
    y_test: np.ndarray

    @property
    def train_counts(self) -> list[int]:
        return [len(X_class) for X_class in self.train_classes]

    @property
    def test_counts(self) -> list[int]:
        return [len(X_class) for X_class in self.test_classes]


def parse_optional_int(value: str) -> int | None:
    """Parse optional integer CLI values."""
    if value.lower() in {"none", "null", "unlimited"}:
        return None
    return int(value)


def parse_optional_float(value: str) -> float | None:
    """Parse optional floating-point CLI values."""
    if value.lower() in {"none", "null", "off", "disabled"}:
        return None
    return float(value)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run HCESVM(test3) on a teaching-data dataset with a deadline-aware stop policy."
    )
    parser.add_argument(
        "--dataset",
        default=DEFAULT_DATASET,
        choices=DEFAULT_DATASETS,
        help="Teaching-data dataset to run. Defaults to skill.",
    )
    parser.add_argument(
        "--deadline-at",
        default=DEFAULT_DEADLINE_AT,
        help="Local deadline in ISO format, without timezone. Default: 2026-04-25T12:00:00",
    )
    parser.add_argument(
        "--deadline-tz",
        default=DEFAULT_DEADLINE_TZ,
        help="Timezone for --deadline-at. Default: Asia/Taipei.",
    )
    parser.add_argument(
        "--overrun-buffer-minutes",
        type=int,
        default=30,
        help="Allowed overrun after the deadline when the estimated finish stays within this buffer.",
    )
    parser.add_argument(
        "--time-limit",
        type=parse_optional_int,
        default=None,
        help="Per-classifier Gurobi time limit in seconds.",
    )
    parser.add_argument(
        "--heartbeat-interval-seconds",
        type=parse_optional_float,
        default=60.0,
        help=(
            "Emit solver heartbeats this often while a classifier is still solving. "
            "Use 'none' to disable. Default: 60."
        ),
    )
    parser.add_argument(
        "--mip-gap",
        type=float,
        default=1e-4,
        help="Gurobi MIP gap tolerance.",
    )
    parser.add_argument(
        "--threads",
        type=int,
        default=DEFAULT_HCESVM_THREADS,
        help="Gurobi thread count. Default: 2.",
    )
    parser.add_argument(
        "--soft-mem-limit-gb",
        type=float,
        default=DEFAULT_HCESVM_SOFT_MEM_LIMIT_GB,
        help="Gurobi SoftMemLimit in GB. Default: 18.",
    )
    parser.add_argument(
        "--C-hyper",
        type=float,
        default=1.0,
        help="Binary CE-SVM slack penalty coefficient.",
    )
    parser.add_argument(
        "--M",
        type=float,
        default=1000.0,
        help="Binary CE-SVM big-M constant.",
    )
    return parser.parse_args()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def format_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.floating, float)):
        return float(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    return value


def format_vector(values: Any) -> str:
    if values is None:
        return ""
    array = np.asarray(values, dtype=float).tolist()
    return json.dumps([round(float(value), 10) for value in array], ensure_ascii=False)


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    print(f"\n{line}", file=tee)
    print(f"[{human_timestamp(utc_now())}] {title}", file=tee)
    print(line, file=tee)


def git_output(*args: str) -> str:
    return subprocess.check_output(args, cwd=ROOT, text=True).strip()


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def load_dataset(name: str) -> DatasetBundle:
    source_url = f"{DATASET_BASE_URL}/{name}.csv"
    train_classes, test_classes, X_test, y_test, n_features = load_csv_ordinal_data(
        source_url,
        target_col="response",
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
    )

    n_classes = len(train_classes)
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [np.full(len(X_class), class_index, dtype=int) for class_index, X_class in enumerate(train_classes, start=1)]
    )

    return DatasetBundle(
        name=name,
        source_url=source_url,
        n_classes=n_classes,
        n_features=n_features,
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=X_train,
        y_train=y_train.astype(int),
        X_test=np.asarray(X_test, dtype=float),
        y_test=np.asarray(y_test, dtype=int),
    )


def build_metric_headers(bundle: DatasetBundle) -> list[str]:
    headers = [
        "dataset",
        "model",
        "status",
        "stop_reason",
        "n_classes",
        "n_features",
        "fit_seconds",
        "train_total_accuracy",
        "test_total_accuracy",
        "source_url",
    ]
    for class_index in range(1, bundle.n_classes + 1):
        headers.extend(
            [
                f"train_class_{class_index}_count",
                f"train_class_{class_index}_accuracy",
                f"test_class_{class_index}_count",
                f"test_class_{class_index}_accuracy",
            ]
        )
    return headers


def build_metric_row(
    bundle: DatasetBundle,
    *,
    status: str,
    fit_seconds: float,
    stop_reason: str | None,
    train_metrics: dict[str, Any] | None,
    test_metrics: dict[str, Any] | None,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "dataset": bundle.name,
        "model": MODEL_NAME,
        "status": status,
        "stop_reason": stop_reason,
        "n_classes": bundle.n_classes,
        "n_features": bundle.n_features,
        "fit_seconds": fit_seconds,
        "train_total_accuracy": None if train_metrics is None else format_scalar(train_metrics["total_accuracy"]),
        "test_total_accuracy": None if test_metrics is None else format_scalar(test_metrics["total_accuracy"]),
        "source_url": bundle.source_url,
    }

    for class_index in range(1, bundle.n_classes + 1):
        row[f"train_class_{class_index}_count"] = bundle.train_counts[class_index - 1]
        row[f"test_class_{class_index}_count"] = bundle.test_counts[class_index - 1]
        row[f"train_class_{class_index}_accuracy"] = None
        row[f"test_class_{class_index}_accuracy"] = None

        if train_metrics is not None:
            row[f"train_class_{class_index}_accuracy"] = format_scalar(
                train_metrics.get(f"class_{class_index}_accuracy")
            )
        if test_metrics is not None:
            row[f"test_class_{class_index}_accuracy"] = format_scalar(
                test_metrics.get(f"class_{class_index}_accuracy")
            )

    return row


def build_parameter_headers() -> list[str]:
    return [
        "dataset",
        "model",
        "component",
        "description",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "positive_class_accuracy_lb",
        "negative_class_accuracy_lb",
        "mip_gap",
    ]


def build_progress_headers() -> list[str]:
    return [
        "dataset",
        "model",
        "component",
        "hk",
        "description",
        "started_at_utc",
        "finished_at_utc",
        "elapsed_seconds",
        "cumulative_elapsed_seconds",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "positive_class_accuracy_lb",
        "negative_class_accuracy_lb",
        "mip_gap",
        "deadline_reached",
        "estimated_remaining_seconds",
        "estimated_finish_utc",
        "allowed_finish_utc",
        "decision",
    ]


def build_metadata_rows(metadata: dict[str, Any]) -> list[tuple[str, Any]]:
    return [(key, value) for key, value in metadata.items()]


def write_excel(
    output_path: Path,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    metrics_sheet = workbook.active
    metrics_sheet.title = "Metrics"
    metrics_sheet.append(metric_headers)
    for row in metric_rows:
        metrics_sheet.append([row.get(header) for header in metric_headers])

    parameters_sheet = workbook.create_sheet("Parameters")
    parameters_sheet.append(parameter_headers)
    for row in parameter_rows:
        parameters_sheet.append([row.get(header) for header in parameter_headers])

    progress_sheet = workbook.create_sheet("Progress")
    progress_sheet.append(progress_headers)
    for row in progress_rows:
        progress_sheet.append([row.get(header) for header in progress_headers])

    metadata_sheet = workbook.create_sheet("Run_Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in (metrics_sheet, parameters_sheet, progress_sheet, metadata_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def checkpoint_excel(
    output_path: Path,
    tee: TeeStream,
    label: str,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    write_excel(
        output_path,
        metric_headers=metric_headers,
        metric_rows=metric_rows,
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        progress_headers=progress_headers,
        progress_rows=progress_rows,
        metadata_rows=metadata_rows,
    )
    print(f"Checkpoint Excel saved after {label}: {output_path}", file=tee)


def log_metrics(tee: TeeStream, label: str, metrics: dict[str, Any], n_classes: int) -> None:
    print(f"{label} total accuracy: {metrics['total_accuracy']:.4f}", file=tee)
    for class_index in range(1, n_classes + 1):
        accuracy = metrics[f"class_{class_index}_accuracy"]
        count = metrics[f"class_{class_index}_count"]
        print(f"  Class {class_index}: {accuracy:.4f} ({count} samples)", file=tee)


def main() -> int:
    args = parse_arguments()
    started_at = utc_now()
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")
    archive_dir = ROOT / "results" / "archive" / f"{date_prefix}_test3_{args.dataset}_deadline"
    archive_dir.mkdir(parents=True, exist_ok=True)

    log_path = archive_dir / f"test3_{args.dataset}_deadline_{timestamp}.log"
    xlsx_path = ROOT / "docs" / "reports" / f"{args.dataset.upper()}_HCESVM_TEST3_DEADLINE_{timestamp}.xlsx"

    deadline_local, deadline_utc = parse_deadline(args.deadline_at, args.deadline_tz)
    resolved_time_limit = args.time_limit
    resolved_time_limit_message = format_hcesvm_time_limit_message(resolved_time_limit)
    hcesvm_params = {
        "C_hyper": args.C_hyper,
        "M": args.M,
        "time_limit": resolved_time_limit,
        "mip_gap": args.mip_gap,
        "threads": args.threads,
        "soft_mem_limit_gb": args.soft_mem_limit_gb,
        "heartbeat_interval_seconds": args.heartbeat_interval_seconds,
        "retain_raw_solution_arrays": False,
        "release_solver_resources_after_fit": True,
        "verbose": False,
    }

    metric_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    progress_rows: list[dict[str, Any]] = []
    completed_durations: list[float] = []
    final_status = "running"
    stop_reason: str | None = None

    with log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, f"Teaching-data deadline-aware HCESVM run ({args.dataset})")
        print(f"Branch: {git_output('git', 'rev-parse', '--abbrev-ref', 'HEAD')}", file=tee)
        print(f"Commit: {git_output('git', 'rev-parse', 'HEAD')}", file=tee)
        print("Split rule: Stratified 80/20 train/test split", file=tee)
        print(f"Random state: {RANDOM_STATE}", file=tee)
        print(f"Dataset: {args.dataset}", file=tee)
        print(f"HCESVM params: {json.dumps(hcesvm_params, ensure_ascii=False)}", file=tee)
        print(resolved_time_limit_message, file=tee)
        print(f"Deadline local: {deadline_local.isoformat()}", file=tee)
        print(f"Deadline UTC: {deadline_utc.isoformat()}", file=tee)
        print(f"Allowed overrun buffer: {args.overrun_buffer_minutes} minutes", file=tee)

        bundle: DatasetBundle
        with redirect_stdout(tee), redirect_stderr(tee):
            bundle = load_dataset(args.dataset)

        metric_headers = build_metric_headers(bundle)
        parameter_headers = build_parameter_headers()
        progress_headers = build_progress_headers()

        metadata = {
            "generated_at_utc": human_timestamp(started_at),
            "branch": git_output("git", "rev-parse", "--abbrev-ref", "HEAD"),
            "commit": git_output("git", "rev-parse", "HEAD"),
            "worktree": str(ROOT),
            "log_path": str(log_path),
            "excel_path": str(xlsx_path),
            "dataset": bundle.name,
            "dataset_base_url": DATASET_BASE_URL,
            "source_url": bundle.source_url,
            "split_rule": f"Stratified train_test_split with test_size={TEST_SIZE}",
            "random_state": RANDOM_STATE,
            "train_counts": json.dumps(bundle.train_counts, ensure_ascii=False),
            "test_counts": json.dumps(bundle.test_counts, ensure_ascii=False),
            "n_classes": bundle.n_classes,
            "n_features": bundle.n_features,
            "model": MODEL_NAME,
            "hcesvm_params": json.dumps(hcesvm_params, ensure_ascii=False),
            "hcesvm_time_limit_message": resolved_time_limit_message,
            "deadline_local": deadline_local.isoformat(),
            "deadline_utc": deadline_utc.isoformat(),
            "deadline_timezone": args.deadline_tz,
            "overrun_buffer_minutes": args.overrun_buffer_minutes,
            "stop_policy": "between_hk",
            "final_status": final_status,
            "completed_classifier_count": 0,
            "expected_classifier_count": bundle.n_classes - 1,
            "stop_reason": stop_reason,
        }

        checkpoint_excel(
            xlsx_path,
            tee,
            "dataset loading",
            metric_headers=metric_headers,
            metric_rows=metric_rows,
            parameter_headers=parameter_headers,
            parameter_rows=parameter_rows,
            progress_headers=progress_headers,
            progress_rows=progress_rows,
            metadata_rows=build_metadata_rows(metadata),
        )

        banner(tee, f"{bundle.name} | {MODEL_NAME}", major=False)
        model = HierarchicalCESVM(
            cesvm_params=hcesvm_params,
            strategy="test3",
            n_classes=bundle.n_classes,
        )

        def after_classifier(progress: dict[str, Any]) -> bool:
            nonlocal stop_reason

            completed_durations.append(float(progress["elapsed_seconds"]))
            remaining_steps = int(progress["n_classifiers"]) - int(progress["hk"])
            decision = evaluate_deadline_policy(
                now_utc=progress["finished_at_utc"],
                deadline_utc=deadline_utc,
                overrun_buffer_minutes=args.overrun_buffer_minutes,
                completed_durations=completed_durations,
                remaining_steps=remaining_steps,
            )

            parameter_rows.append(
                {
                    "dataset": bundle.name,
                    "model": MODEL_NAME,
                    "component": str(progress["component"]).upper(),
                    "description": progress["description"],
                    "positive_sample_count": progress["positive_sample_count"],
                    "negative_sample_count": progress["negative_sample_count"],
                    "weights": format_vector(progress["weights"]),
                    "b": format_scalar(progress["b"]),
                    "objective_value": format_scalar(progress["objective_value"]),
                    "positive_class_accuracy_lb": format_scalar(progress["positive_class_accuracy_lb"]),
                    "negative_class_accuracy_lb": format_scalar(progress["negative_class_accuracy_lb"]),
                    "mip_gap": format_scalar(progress["mip_gap"]),
                }
            )

            progress_rows.append(
                {
                    "dataset": bundle.name,
                    "model": MODEL_NAME,
                    "component": str(progress["component"]).upper(),
                    "hk": int(progress["hk"]),
                    "description": progress["description"],
                    "started_at_utc": human_timestamp(progress["started_at_utc"]),
                    "finished_at_utc": human_timestamp(progress["finished_at_utc"]),
                    "elapsed_seconds": format_scalar(progress["elapsed_seconds"]),
                    "cumulative_elapsed_seconds": format_scalar(progress["cumulative_elapsed_seconds"]),
                    "positive_sample_count": progress["positive_sample_count"],
                    "negative_sample_count": progress["negative_sample_count"],
                    "weights": format_vector(progress["weights"]),
                    "b": format_scalar(progress["b"]),
                    "objective_value": format_scalar(progress["objective_value"]),
                    "positive_class_accuracy_lb": format_scalar(progress["positive_class_accuracy_lb"]),
                    "negative_class_accuracy_lb": format_scalar(progress["negative_class_accuracy_lb"]),
                    "mip_gap": format_scalar(progress["mip_gap"]),
                    "deadline_reached": decision.deadline_reached,
                    "estimated_remaining_seconds": format_scalar(decision.estimated_remaining_seconds),
                    "estimated_finish_utc": human_timestamp(decision.estimated_finish_utc),
                    "allowed_finish_utc": human_timestamp(decision.allowed_finish_utc),
                    "decision": decision.decision,
                }
            )

            metadata["completed_classifier_count"] = len(parameter_rows)

            print(
                f"Completed {str(progress['component']).upper()}: "
                f"elapsed={progress['elapsed_seconds']:.2f}s, "
                f"cumulative={progress['cumulative_elapsed_seconds']:.2f}s",
                file=tee,
            )
            print(
                f"  Deadline reached: {decision.deadline_reached}; "
                f"estimated remaining: {decision.estimated_remaining_seconds}",
                file=tee,
            )
            print(
                f"  Estimated finish UTC: {human_timestamp(decision.estimated_finish_utc)}; "
                f"allowed finish UTC: {human_timestamp(decision.allowed_finish_utc)}",
                file=tee,
            )
            print(f"  Decision: {decision.decision}", file=tee)

            if not decision.continue_training and remaining_steps > 0:
                stop_reason = (
                    f"deadline reached; estimated finish "
                    f"{human_timestamp(decision.estimated_finish_utc)} exceeds allowed finish "
                    f"{human_timestamp(decision.allowed_finish_utc)}"
                )
                metadata["stop_reason"] = stop_reason

            checkpoint_excel(
                xlsx_path,
                tee,
                f"{bundle.name} / {str(progress['component']).upper()}",
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                progress_headers=progress_headers,
                progress_rows=progress_rows,
                metadata_rows=build_metadata_rows(metadata),
            )

            return decision.continue_training

        fit_started_at = utc_now()

        try:
            with redirect_stdout(tee), redirect_stderr(tee):
                model.fit_incremental(*bundle.train_classes, after_classifier=after_classifier)
        except Exception:
            final_status = "failed"
            stop_reason = traceback.format_exc()
            metadata["final_status"] = final_status
            metadata["stop_reason"] = stop_reason
            metric_rows[:] = [
                build_metric_row(
                    bundle,
                    status=final_status,
                    fit_seconds=(utc_now() - fit_started_at).total_seconds(),
                    stop_reason=stop_reason,
                    train_metrics=None,
                    test_metrics=None,
                )
            ]
            print(stop_reason, file=tee)
            checkpoint_excel(
                xlsx_path,
                tee,
                "failure",
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                progress_headers=progress_headers,
                progress_rows=progress_rows,
                metadata_rows=build_metadata_rows(metadata),
            )
            return 1

        fit_seconds = (utc_now() - fit_started_at).total_seconds()
        metadata["completed_classifier_count"] = model.completed_classifier_count
        metadata["expected_classifier_count"] = model.expected_classifier_count

        if model.is_fully_fitted():
            final_status = "completed"
            train_predictions = np.asarray(model.predict(bundle.X_train), dtype=int)
            test_predictions = np.asarray(model.predict(bundle.X_test), dtype=int)
            train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
            test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)

            metric_rows[:] = [
                build_metric_row(
                    bundle,
                    status=final_status,
                    fit_seconds=fit_seconds,
                    stop_reason=None,
                    train_metrics=train_metrics,
                    test_metrics=test_metrics,
                )
            ]

            log_metrics(tee, "Training", train_metrics, bundle.n_classes)
            log_metrics(tee, "Testing", test_metrics, bundle.n_classes)
        else:
            final_status = "stopped_by_deadline"
            if stop_reason is None:
                stop_reason = metadata.get("stop_reason") or model.fit_stop_reason

            metric_rows[:] = [
                build_metric_row(
                    bundle,
                    status=final_status,
                    fit_seconds=fit_seconds,
                    stop_reason=stop_reason,
                    train_metrics=None,
                    test_metrics=None,
                )
            ]
            print(f"Run stopped before all classifiers completed.", file=tee)
            print(f"Stop reason: {stop_reason}", file=tee)

        metadata["final_status"] = final_status
        metadata["stop_reason"] = stop_reason

        banner(tee, "Run complete")
        print(f"Final status: {final_status}", file=tee)
        print(f"Completed classifiers: {model.completed_classifier_count}/{model.expected_classifier_count}", file=tee)
        print(f"Log saved to: {log_path}", file=tee)
        print(f"Excel saved to: {xlsx_path}", file=tee)

        checkpoint_excel(
            xlsx_path,
            tee,
            "finalization",
            metric_headers=metric_headers,
            metric_rows=metric_rows,
            parameter_headers=parameter_headers,
            parameter_rows=parameter_rows,
            progress_headers=progress_headers,
            progress_rows=progress_rows,
            metadata_rows=build_metadata_rows(metadata),
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
