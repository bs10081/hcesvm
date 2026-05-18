#!/usr/bin/env python3
"""Run HCESVM(test3) on full teaching-data ordinal datasets."""

from __future__ import annotations

import argparse
import hashlib
import json
import subprocess
import sys
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from hcesvm.models.hierarchical import HierarchicalCESVM
from hcesvm.utils.data_loader import load_csv_ordinal_data
from hcesvm.utils.evaluator import evaluate_multiclass
from hcesvm.utils.ordinal_data import load_lingo_split_workbook
from hcesvm.utils.teaching_data_1000 import (
    DATASET_BASE_URL,
    build_classifier_diagnostics_row,
    build_classifier_progress_row,
    class_counts,
    format_scalar,
    source_dataset_url,
)
from hcesvm.utils.teaching_data_runtime import format_hcesvm_time_limit_message


MODEL_NAME = "HCESVM(test3)"
DEFAULT_DATASETS = [
    "bostonhousing_ord",
    "hayes_roth",
    "new_thyroid",
    "car_evaluation",
    "balance",
    "thyroid",
    "cement_strength",
    "stock_ord",
]
TEACHING_DATASETS = {"bostonhousing_ord", "cement_strength", "stock_ord"}
PENDING_DATASET_PATHS = {
    "hayes_roth": [
        Path("~/Developer/NSVORA/Archive/hayes-roth_split.xlsx"),
        Path("~/Developer/NSVORA/Archive/hayes_roth_split.xlsx"),
    ],
    "new_thyroid": [
        Path("~/Developer/NSVORA/Archive/new-thyroid_split.xlsx"),
        Path("~/Developer/NSVORA/Archive/new_thyroid_split.xlsx"),
    ],
    "car_evaluation": [
        Path("~/Developer/NSVORA/datasets/primary/car_evaluation/car_evaluation_split.xlsx"),
    ],
    "thyroid": [
        Path("~/Developer/NSVORA/Archive/thyroid_split.xlsx"),
    ],
}
BALANCE_WORKBOOK = ROOT / "data" / "svor" / "SVOR_balance_split.xlsx"
TEST_SIZE = 0.2
RANDOM_STATE = 42
TEACHING_CSV_SPLIT_RULE = f"Stratified train_test_split with test_size={TEST_SIZE}"
WORKBOOK_SPLIT_RULE = "Preserve workbook Train/Test sheets"


class TeeStream:
    """Mirror output to stdout/stderr and the run log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass(slots=True)
class FullDatasetSplit:
    name: str
    source_url: str
    feature_names: list[str]
    train_classes: list[np.ndarray]
    test_classes: list[np.ndarray]
    X_train: np.ndarray
    y_train: np.ndarray
    X_test: np.ndarray
    y_test: np.ndarray
    split_rule: str = "unknown"
    random_state: int | None = None

    @property
    def n_classes(self) -> int:
        return len(self.train_classes)

    @property
    def n_features(self) -> int:
        return int(self.X_train.shape[1])

    @property
    def train_counts(self) -> list[int]:
        return class_counts(self.y_train, n_classes=self.n_classes)

    @property
    def test_counts(self) -> list[int]:
        return class_counts(self.y_test, n_classes=self.n_classes)


@dataclass(frozen=True, slots=True)
class DatasetSourceSpec:
    name: str
    source_type: str
    source: str
    missing_candidates: tuple[str, ...] = ()


@dataclass(slots=True)
class DatasetRunOutcome:
    split: FullDatasetSplit
    status: str
    stop_reason: str | None
    fit_seconds: float | None
    train_metrics: dict[str, Any] | None
    test_metrics: dict[str, Any] | None
    diagnostics_rows: list[dict[str, Any]]
    per_classifier_time_limit: int | None
    time_limit_message: str


def parse_optional_float(value: str) -> float | None:
    """Parse optional floating-point CLI values."""
    if value.lower() in {"none", "null", "off", "disabled"}:
        return None
    return float(value)


def parse_optional_int(value: str) -> int | None:
    """Parse optional integer CLI values."""
    if value.lower() in {"none", "null", "off", "disabled"}:
        return None
    return int(value)


def parse_optional_path(value: str) -> str | None:
    """Parse optional path-like CLI values."""
    if value.lower() in {"none", "null", "off", "disabled"}:
        return None
    return value


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run HCESVM(test3) on full teaching-data ordinal datasets."
    )
    parser.add_argument(
        "--datasets",
        nargs="+",
        choices=DEFAULT_DATASETS,
        default=["bostonhousing_ord"],
        help="Datasets to run. Default: bostonhousing_ord.",
    )
    parser.add_argument(
        "--time-limit",
        type=parse_optional_int,
        default=5400,
        help="HCESVM per-classifier time limit in seconds. Use 'none' for no Gurobi TimeLimit. Default: 5400.",
    )
    parser.add_argument(
        "--max-classifiers-per-dataset",
        type=int,
        default=None,
        help="Stop after this many classifiers per dataset. Useful for smoke runs.",
    )
    parser.add_argument(
        "--heartbeat-interval-seconds",
        type=parse_optional_float,
        default=60.0,
        help="Emit solver heartbeats this often. Use 'none' to disable. Default: 60.",
    )
    parser.add_argument("--threads", type=int, default=20, help="Gurobi thread count. Default: 20.")
    parser.add_argument(
        "--soft-mem-limit-gb",
        type=float,
        default=56.0,
        help="Gurobi SoftMemLimit in GB. Default: 56.",
    )
    parser.add_argument(
        "--nodefile-start-gb",
        type=parse_optional_float,
        default=None,
        help="Gurobi NodeFileStart in GB. Use 'none' to keep Gurobi default. Default: none.",
    )
    parser.add_argument(
        "--nodefile-dir",
        type=parse_optional_path,
        default="auto",
        help=(
            "Gurobi NodeFileDir path, 'auto' for "
            "$HOME/hcesvm-gurobi-nodefiles/<dataset>_<timestamp>, or 'none'. "
            "Used when --nodefile-start-gb is set. Default: auto."
        ),
    )
    parser.add_argument("--mip-gap", type=float, default=1e-4, help="Gurobi MIP gap tolerance.")
    parser.add_argument(
        "--mip-focus",
        type=parse_optional_int,
        default=None,
        help="Gurobi MIPFocus value 0..3. Use 'none' to keep Gurobi default. Default: none.",
    )
    parser.add_argument("--C-hyper", type=float, default=1.0, help="Binary CE-SVM slack penalty.")
    parser.add_argument("--M", type=float, default=1000.0, help="Binary CE-SVM big-M constant.")
    parser.add_argument(
        "--feat-lower-bound",
        type=float,
        default=1e-7,
        help="Binary CE-SVM lower bound for active features. Default: 1e-7.",
    )
    return parser.parse_args(argv)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def git_output(*args: str) -> str:
    return subprocess.check_output(args, cwd=ROOT, text=True).strip()


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    print(f"\n{line}", file=tee)
    print(f"[{human_timestamp(utc_now())}] {title}", file=tee)
    print(line, file=tee)


def format_time_limit_value(time_limit: int | None) -> str:
    return "none" if time_limit is None else f"{time_limit}s"


def resolve_nodefile_dir(
    *,
    nodefile_start_gb: float | None,
    nodefile_dir_option: str | None,
    dataset_slug: str,
    timestamp: str,
) -> str | None:
    """Resolve and create the nodefile directory only for nodefile-enabled runs."""
    if nodefile_start_gb is None or nodefile_dir_option is None:
        return None

    if nodefile_dir_option.lower() == "auto":
        nodefile_dir = Path.home() / "hcesvm-gurobi-nodefiles" / f"{dataset_slug}_{timestamp}"
    else:
        nodefile_dir = Path(nodefile_dir_option).expanduser()

    nodefile_dir.mkdir(parents=True, exist_ok=True)
    return str(nodefile_dir)


def resolve_dataset_source(name: str) -> DatasetSourceSpec:
    if name in TEACHING_DATASETS:
        return DatasetSourceSpec(
            name=name,
            source_type="teaching_csv",
            source=source_dataset_url(name),
        )
    if name == "balance":
        return DatasetSourceSpec(
            name=name,
            source_type="lingo_workbook_preserve_split",
            source=str(BALANCE_WORKBOOK),
        )

    candidates = PENDING_DATASET_PATHS.get(name)
    if candidates:
        expanded = tuple(str(path.expanduser()) for path in candidates)
        existing = [path.expanduser() for path in candidates if path.expanduser().exists()]
        if existing:
            return DatasetSourceSpec(
                name=name,
                source_type="lingo_workbook_preserve_split",
                source=str(existing[0]),
                missing_candidates=expanded,
            )
        return DatasetSourceSpec(
            name=name,
            source_type="missing",
            source="",
            missing_candidates=expanded,
        )

    raise KeyError(f"Unknown dataset: {name}")


def validate_dataset_sources(dataset_names: list[str]) -> list[DatasetSourceSpec]:
    specs = [resolve_dataset_source(name) for name in dataset_names]
    missing = [spec for spec in specs if spec.source_type == "missing"]
    if missing:
        lines = ["Dataset source preflight failed. Missing source files:"]
        for spec in missing:
            lines.append(f"- {spec.name}:")
            for candidate in spec.missing_candidates:
                lines.append(f"  - {candidate}")
        raise FileNotFoundError("\n".join(lines))
    return specs


def arrays_to_classes(X: np.ndarray, y: np.ndarray, *, name: str) -> list[np.ndarray]:
    classes = sorted(np.unique(y).astype(int).tolist())
    expected_classes = list(range(1, len(classes) + 1))
    if classes != expected_classes:
        raise ValueError(f"{name} labels must be consecutive integers from 1, got {classes!r}.")

    return [
        np.asarray(X[np.asarray(y, dtype=int) == class_index], dtype=float)
        for class_index in expected_classes
    ]


def row_fingerprint(X: np.ndarray, y: np.ndarray) -> str:
    rows = [
        {
            "y": int(label),
            "x": [float(value) for value in row],
        }
        for row, label in zip(np.asarray(X, dtype=float), np.asarray(y, dtype=int))
    ]
    payload = json.dumps(rows, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def load_teaching_csv_dataset(name: str, source_url: str) -> FullDatasetSplit:
    """Load the canonical full teaching-data split for one dataset."""
    train_classes, test_classes, X_test, y_test, n_features = load_csv_ordinal_data(
        source_url,
        target_col="response",
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
    )
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [
            np.full(len(X_class), class_index, dtype=int)
            for class_index, X_class in enumerate(train_classes, start=1)
        ]
    )
    return FullDatasetSplit(
        name=name,
        source_url=source_url,
        feature_names=[f"feature_{index}" for index in range(1, n_features + 1)],
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=np.asarray(X_train, dtype=float),
        y_train=y_train.astype(int),
        X_test=np.asarray(X_test, dtype=float),
        y_test=np.asarray(y_test, dtype=int),
        split_rule=TEACHING_CSV_SPLIT_RULE,
        random_state=RANDOM_STATE,
    )


def load_lingo_workbook_preserve_split_dataset(name: str, source_path: str) -> FullDatasetSplit:
    """Load a LINGO workbook while preserving its Train/Test sheets."""
    dataset = load_lingo_split_workbook(source_path)
    X_train = np.asarray(dataset.X_train, dtype=float)
    y_train = np.asarray(dataset.y_train, dtype=int)
    X_test = np.asarray(dataset.X_test, dtype=float)
    y_test = np.asarray(dataset.y_test, dtype=int)
    train_classes = arrays_to_classes(X_train, y_train, name=name)
    test_classes = arrays_to_classes(X_test, y_test, name=name)
    if len(train_classes) != len(test_classes):
        raise ValueError(
            f"{name} workbook Train/Test sheets have different class counts: "
            f"{len(train_classes)} vs {len(test_classes)}."
        )
    return FullDatasetSplit(
        name=name,
        source_url=str(dataset.workbook_path),
        feature_names=list(dataset.feature_names),
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        split_rule=WORKBOOK_SPLIT_RULE,
        random_state=None,
    )


def load_dataset_from_spec(spec: DatasetSourceSpec) -> FullDatasetSplit:
    if spec.source_type == "teaching_csv":
        return load_teaching_csv_dataset(spec.name, spec.source)
    if spec.source_type == "lingo_workbook_preserve_split":
        return load_lingo_workbook_preserve_split_dataset(spec.name, spec.source)
    raise ValueError(f"Dataset {spec.name!r} is not loadable: {spec.source_type}")


def load_dataset(name: str) -> FullDatasetSplit:
    """Load one configured dataset after resolving its source."""
    return load_dataset_from_spec(resolve_dataset_source(name))


def build_metric_headers(max_classes: int) -> list[str]:
    headers = [
        "dataset",
        "model",
        "status",
        "stop_reason",
        "n_classes",
        "n_features",
        "fit_seconds",
        "train_total_accuracy",
        "test_total_accuracy",
        "source_url",
    ]
    for class_index in range(1, max_classes + 1):
        headers.extend(
            [
                f"train_class_{class_index}_count",
                f"train_class_{class_index}_accuracy",
                f"test_class_{class_index}_count",
                f"test_class_{class_index}_accuracy",
            ]
        )
    return headers


def build_metric_row(outcome: DatasetRunOutcome, *, max_classes: int) -> dict[str, Any]:
    split = outcome.split
    row: dict[str, Any] = {
        "dataset": split.name,
        "model": MODEL_NAME,
        "status": outcome.status,
        "stop_reason": outcome.stop_reason,
        "n_classes": split.n_classes,
        "n_features": split.n_features,
        "fit_seconds": outcome.fit_seconds,
        "train_total_accuracy": None,
        "test_total_accuracy": None,
        "source_url": split.source_url,
    }
    if outcome.train_metrics is not None:
        row["train_total_accuracy"] = format_scalar(outcome.train_metrics["total_accuracy"])
    if outcome.test_metrics is not None:
        row["test_total_accuracy"] = format_scalar(outcome.test_metrics["total_accuracy"])

    for class_index in range(1, max_classes + 1):
        row[f"train_class_{class_index}_count"] = None
        row[f"train_class_{class_index}_accuracy"] = None
        row[f"test_class_{class_index}_count"] = None
        row[f"test_class_{class_index}_accuracy"] = None
        if class_index <= split.n_classes:
            row[f"train_class_{class_index}_count"] = split.train_counts[class_index - 1]
            row[f"test_class_{class_index}_count"] = split.test_counts[class_index - 1]
        if outcome.train_metrics is not None and class_index <= split.n_classes:
            row[f"train_class_{class_index}_accuracy"] = format_scalar(
                outcome.train_metrics.get(f"class_{class_index}_accuracy")
            )
        if outcome.test_metrics is not None and class_index <= split.n_classes:
            row[f"test_class_{class_index}_accuracy"] = format_scalar(
                outcome.test_metrics.get(f"class_{class_index}_accuracy")
            )
    return row


def build_parameter_headers() -> list[str]:
    return [
        "dataset",
        "classifier",
        "description",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "mip_gap",
        "solver_status_code",
        "solver_status_label",
        "mem_used_gb",
        "max_mem_used_gb",
        "elapsed_seconds",
        "cumulative_elapsed_seconds",
    ]


def build_progress_headers() -> list[str]:
    return [
        "dataset",
        "classifier",
        "description",
        "started_at_utc",
        "finished_at_utc",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "mip_gap",
        "solver_status_code",
        "solver_status_label",
        "mem_used_gb",
        "max_mem_used_gb",
        "elapsed_seconds",
        "cumulative_elapsed_seconds",
    ]


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def build_metadata_rows(metadata: dict[str, Any]) -> list[tuple[str, Any]]:
    return [(key, value) for key, value in metadata.items()]


def write_excel(
    output_path: Path,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    metrics_sheet = workbook.active
    metrics_sheet.title = "Metrics"
    metrics_sheet.append(metric_headers)
    for row in metric_rows:
        metrics_sheet.append([row.get(header) for header in metric_headers])

    parameters_sheet = workbook.create_sheet("Parameters")
    parameters_sheet.append(parameter_headers)
    for row in parameter_rows:
        parameters_sheet.append([row.get(header) for header in parameter_headers])

    progress_sheet = workbook.create_sheet("Progress")
    progress_sheet.append(progress_headers)
    for row in progress_rows:
        progress_sheet.append([row.get(header) for header in progress_headers])

    metadata_sheet = workbook.create_sheet("Run_Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in (metrics_sheet, parameters_sheet, progress_sheet, metadata_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def checkpoint_excel(
    output_path: Path,
    tee: TeeStream,
    label: str,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    write_excel(
        output_path,
        metric_headers=metric_headers,
        metric_rows=metric_rows,
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        progress_headers=progress_headers,
        progress_rows=progress_rows,
        metadata_rows=metadata_rows,
    )
    print(f"Checkpoint Excel saved after {label}: {output_path}", file=tee)


def log_metrics(tee: TeeStream, label: str, metrics: dict[str, Any], n_classes: int) -> None:
    print(f"{label} total accuracy: {metrics['total_accuracy']:.4f}", file=tee)
    for class_index in range(1, n_classes + 1):
        accuracy = metrics[f"class_{class_index}_accuracy"]
        count = metrics[f"class_{class_index}_count"]
        print(f"  Class {class_index}: {accuracy:.4f} ({count} samples)", file=tee)


def format_report_table_row(values: list[str]) -> str:
    return "| " + " | ".join(values) + " |"


def render_accuracy_table(outcome: DatasetRunOutcome) -> str:
    split = outcome.split
    lines = [
        format_report_table_row(["Split", "Total Accuracy", "Per-Class Accuracy"]),
        format_report_table_row(["---", "---:", "---"]),
    ]
    for split_name, metrics, counts in (
        ("Train", outcome.train_metrics, split.train_counts),
        ("Test", outcome.test_metrics, split.test_counts),
    ):
        if metrics is None:
            lines.append(format_report_table_row([split_name, "", outcome.stop_reason or ""]))
            continue
        per_class = ", ".join(
            f"C{class_index}={metrics[f'class_{class_index}_accuracy']:.4f} (n={counts[class_index - 1]})"
            for class_index in range(1, split.n_classes + 1)
        )
        lines.append(format_report_table_row([split_name, f"{metrics['total_accuracy']:.4f}", per_class]))
    return "\n".join(lines)


def render_diagnostics_table(outcome: DatasetRunOutcome) -> str:
    lines = [
        format_report_table_row(
            [
                "Classifier",
                "Description",
                "Samples (+/-)",
                "Status",
                "Elapsed (s)",
                "Objective",
                "mip_gap",
                "RAM Used (GB)",
                "Peak RAM (GB)",
                "b",
                "weights",
            ]
        ),
        format_report_table_row(["---", "---", "---", "---", "---:", "---:", "---:", "---:", "---:", "---:", "---"]),
    ]
    for row in outcome.diagnostics_rows:
        lines.append(
            format_report_table_row(
                [
                    str(row["classifier"]),
                    str(row["description"]),
                    f"{row['positive_sample_count']}/{row['negative_sample_count']}",
                    str(row["solver_status_label"]),
                    "" if row["elapsed_seconds"] is None else f"{float(row['elapsed_seconds']):.2f}",
                    "" if row["objective_value"] is None else f"{float(row['objective_value']):.6f}",
                    "" if row["mip_gap"] is None else f"{float(row['mip_gap']):.6f}",
                    "" if row["mem_used_gb"] is None else f"{float(row['mem_used_gb']):.6f}",
                    "" if row["max_mem_used_gb"] is None else f"{float(row['max_mem_used_gb']):.6f}",
                    "" if row["b"] is None else f"{float(row['b']):.6f}",
                    str(row["weights"]),
                ]
            )
        )
    return "\n".join(lines)


def render_markdown_report(
    *,
    outcomes: list[DatasetRunOutcome],
    generated_at: str,
    workbook_path: Path,
    log_path: Path,
    run_config: dict[str, Any],
) -> str:
    lines = [
        "# HCESVM(test3) Full Teaching-Data Validation",
        "",
        f"Generated at: `{generated_at}`",
        f"Workbook: `{workbook_path}`",
        f"Log: `{log_path}`",
        "",
        "## Run Configuration",
        "",
        f"- Threads: `{run_config['threads']}`",
        f"- SoftMemLimit: `{run_config['soft_mem_limit_gb']} GB`",
        f"- Per-classifier time limit: `{run_config['time_limit_label']}`",
        f"- MIPGap: `{run_config['mip_gap']}`",
        f"- MIPFocus: `{run_config['mip_focus']}`",
        f"- NodeFileStart: `{run_config['nodefile_start_gb']}`",
        f"- NodeFileDir: `{run_config['nodefile_dir']}`",
        "",
    ]
    for outcome in outcomes:
        split = outcome.split
        lines.extend(
            [
                f"## {split.name}",
                "",
                f"- Source URL: `{split.source_url}`",
                f"- Split rule: `{split.split_rule}`",
                f"- Train counts: `{split.train_counts}`",
                f"- Test counts: `{split.test_counts}`",
                f"- Per-classifier time limit: `{format_time_limit_value(outcome.per_classifier_time_limit)}`",
                f"- Expected classifiers: `{split.n_classes - 1}`",
                f"- Final status: `{outcome.status}`",
            ]
        )
        if split.random_state is not None:
            lines.append(f"- Random state: `{split.random_state}`")
        if outcome.stop_reason:
            lines.append(f"- Stop reason: `{outcome.stop_reason}`")
        lines.extend(["", render_accuracy_table(outcome), "", render_diagnostics_table(outcome), ""])
    return "\n".join(lines)


def get_completed_classifier(model: HierarchicalCESVM, hk: int) -> Any:
    key = f"h{hk}"
    if key in model.classifiers:
        return model.classifiers[key]
    if hk == 1 and model.h1 is not None:
        return model.h1
    if hk == 2 and model.h2 is not None:
        return model.h2
    raise KeyError(f"Classifier {key} is not available after callback.")


def is_expected_early_stop(outcome: DatasetRunOutcome) -> bool:
    return outcome.status == "stopped_early" and bool(outcome.stop_reason) and str(
        outcome.stop_reason
    ).startswith("requested stop after")


def main(argv: list[str] | None = None) -> int:
    args = parse_arguments(argv)
    started_at = utc_now()
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")

    archive_label = "test3_no_time_limit_full" if args.time_limit is None else "test3_teaching_data_full"
    archive_dir = ROOT / "results" / "archive" / f"{date_prefix}_{archive_label}"
    archive_dir.mkdir(parents=True, exist_ok=True)
    dataset_slug = "all" if len(args.datasets) > 1 else args.datasets[0]
    resolved_nodefile_dir = resolve_nodefile_dir(
        nodefile_start_gb=args.nodefile_start_gb,
        nodefile_dir_option=args.nodefile_dir,
        dataset_slug=dataset_slug,
        timestamp=timestamp,
    )
    log_path = archive_dir / f"test3_{dataset_slug}_{timestamp}.log"
    xlsx_path = ROOT / "docs" / "reports" / f"{dataset_slug.upper()}_HCESVM_TEST3_FULL_{timestamp}.xlsx"
    report_path = ROOT / "docs" / "reports" / f"{dataset_slug.upper()}_HCESVM_TEST3_FULL_{timestamp}.md"

    metric_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    progress_rows: list[dict[str, Any]] = []
    outcomes: list[DatasetRunOutcome] = []
    loaded_splits: list[FullDatasetSplit] = []

    metadata = {
        "generated_at_utc": human_timestamp(started_at),
        "branch": git_output("git", "rev-parse", "--abbrev-ref", "HEAD"),
        "commit": git_output("git", "rev-parse", "HEAD"),
        "worktree": str(ROOT),
        "log_path": str(log_path),
        "excel_path": str(xlsx_path),
        "report_path": str(report_path),
        "dataset_base_url": DATASET_BASE_URL,
        "datasets": ", ".join(args.datasets),
        "split_rule": "per-dataset; see <dataset>_split_rule",
        "hcesvm_time_limit_seconds": args.time_limit,
        "hcesvm_time_limit_label": "none" if args.time_limit is None else str(args.time_limit),
        "max_classifiers_per_dataset": args.max_classifiers_per_dataset,
        "threads": args.threads,
        "soft_mem_limit_gb": args.soft_mem_limit_gb,
        "mip_gap": args.mip_gap,
        "mip_focus": args.mip_focus,
        "nodefile_start_gb": args.nodefile_start_gb,
        "nodefile_dir_requested": args.nodefile_dir,
        "nodefile_dir": resolved_nodefile_dir,
        "heartbeat_interval_seconds": args.heartbeat_interval_seconds,
        "feat_lower_bound": args.feat_lower_bound,
    }

    with log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, "HCESVM(test3) full teaching-data validation")
        print(f"Branch: {metadata['branch']}", file=tee)
        print(f"Commit: {metadata['commit']}", file=tee)
        print(f"Datasets: {metadata['datasets']}", file=tee)
        print(format_hcesvm_time_limit_message(args.time_limit), file=tee)
        print(f"Threads: {args.threads}", file=tee)
        print(f"SoftMemLimit: {args.soft_mem_limit_gb} GB", file=tee)
        print(f"MIPGap: {args.mip_gap}", file=tee)
        print(f"MIPFocus: {args.mip_focus}", file=tee)
        print(f"NodeFileStart: {args.nodefile_start_gb}", file=tee)
        print(f"NodeFileDir: {resolved_nodefile_dir}", file=tee)
        print(f"Heartbeat interval: {args.heartbeat_interval_seconds}", file=tee)
        print(f"Feature lower bound: {args.feat_lower_bound}", file=tee)
        print(f"Split rule: {metadata['split_rule']}", file=tee)

        try:
            source_specs = validate_dataset_sources(args.datasets)
        except FileNotFoundError as exc:
            print(str(exc), file=tee)
            return 1

        for spec in source_specs:
            print(f"Source [{spec.name}]: {spec.source_type} | {spec.source}", file=tee)

        for spec in source_specs:
            with redirect_stdout(tee), redirect_stderr(tee):
                split = load_dataset_from_spec(spec)
            loaded_splits.append(split)

            metadata[f"{split.name}_source_url"] = split.source_url
            metadata[f"{split.name}_source_type"] = spec.source_type
            metadata[f"{split.name}_split_rule"] = split.split_rule
            metadata[f"{split.name}_random_state"] = split.random_state
            metadata[f"{split.name}_train_counts"] = json.dumps(split.train_counts, ensure_ascii=False)
            metadata[f"{split.name}_test_counts"] = json.dumps(split.test_counts, ensure_ascii=False)
            metadata[f"{split.name}_train_row_fingerprint"] = row_fingerprint(split.X_train, split.y_train)
            metadata[f"{split.name}_test_row_fingerprint"] = row_fingerprint(split.X_test, split.y_test)
            metadata[f"{split.name}_expected_classifiers"] = split.n_classes - 1

        max_classes = max(split.n_classes for split in loaded_splits)
        metric_headers = build_metric_headers(max_classes)
        parameter_headers = build_parameter_headers()
        progress_headers = build_progress_headers()

        checkpoint_excel(
            xlsx_path,
            tee,
            "dataset loading",
            metric_headers=metric_headers,
            metric_rows=metric_rows,
            parameter_headers=parameter_headers,
            parameter_rows=parameter_rows,
            progress_headers=progress_headers,
            progress_rows=progress_rows,
            metadata_rows=build_metadata_rows(metadata),
        )

        for split in loaded_splits:

            banner(tee, f"{split.name} | {MODEL_NAME}", major=False)
            print(f"Source URL: {split.source_url}", file=tee)
            print(f"Split rule: {split.split_rule}", file=tee)
            print(f"Features: {split.n_features}", file=tee)
            print(f"Train counts: {split.train_counts}", file=tee)
            print(f"Test counts: {split.test_counts}", file=tee)

            per_classifier_time_limit = args.time_limit
            time_limit_message = format_hcesvm_time_limit_message(per_classifier_time_limit)
            print(time_limit_message, file=tee)
            print(f"Feature lower bound: {args.feat_lower_bound}", file=tee)
            hcesvm_params = {
                "C_hyper": args.C_hyper,
                "M": args.M,
                "time_limit": per_classifier_time_limit,
                "mip_gap": args.mip_gap,
                "threads": args.threads,
                "soft_mem_limit_gb": args.soft_mem_limit_gb,
                "mip_focus": args.mip_focus,
                "heartbeat_interval_seconds": args.heartbeat_interval_seconds,
                "feat_lower_bound": args.feat_lower_bound,
                "retain_raw_solution_arrays": False,
                "release_solver_resources_after_fit": True,
                "verbose": False,
            }
            if args.nodefile_start_gb is not None:
                hcesvm_params["nodefile_start"] = args.nodefile_start_gb
            if resolved_nodefile_dir is not None:
                hcesvm_params["nodefile_dir"] = resolved_nodefile_dir
            if args.mip_focus is not None:
                hcesvm_params["mip_focus"] = args.mip_focus
            model = HierarchicalCESVM(
                cesvm_params=hcesvm_params,
                strategy="test3",
                n_classes=split.n_classes,
            )

            dataset_diagnostics: list[dict[str, Any]] = []
            dataset_stop_reason: str | None = None
            fit_started_at = utc_now()

            def after_classifier(progress: dict[str, Any]) -> bool:
                nonlocal dataset_stop_reason
                hk = int(progress["hk"])
                classifier = get_completed_classifier(model, hk)
                diagnostics_row = build_classifier_diagnostics_row(
                    dataset_name=split.name,
                    progress=progress,
                    classifier=classifier,
                )
                progress_row = build_classifier_progress_row(
                    diagnostics_row=diagnostics_row,
                    progress={
                        "started_at_utc": human_timestamp(progress["started_at_utc"]),
                        "finished_at_utc": human_timestamp(progress["finished_at_utc"]),
                    },
                )
                parameter_rows.append(diagnostics_row)
                progress_rows.append(progress_row)
                dataset_diagnostics.append(diagnostics_row)
                print(
                    f"{diagnostics_row['classifier']} status: {diagnostics_row['solver_status_label']} "
                    f"(code={diagnostics_row['solver_status_code']}), "
                    f"elapsed={diagnostics_row['elapsed_seconds']:.2f}s, "
                    f"mem={diagnostics_row['mem_used_gb']} GB, "
                    f"peak_mem={diagnostics_row['max_mem_used_gb']} GB",
                    file=tee,
                )
                checkpoint_excel(
                    xlsx_path,
                    tee,
                    f"{split.name} / {diagnostics_row['classifier']}",
                    metric_headers=metric_headers,
                    metric_rows=metric_rows,
                    parameter_headers=parameter_headers,
                    parameter_rows=parameter_rows,
                    progress_headers=progress_headers,
                    progress_rows=progress_rows,
                    metadata_rows=build_metadata_rows(metadata),
                )
                if (
                    args.max_classifiers_per_dataset is not None
                    and len(dataset_diagnostics) >= args.max_classifiers_per_dataset
                    and model.completed_classifier_count < model.expected_classifier_count
                ):
                    dataset_stop_reason = (
                        f"requested stop after {args.max_classifiers_per_dataset} classifiers"
                    )
                    return False
                return True

            try:
                with redirect_stdout(tee), redirect_stderr(tee):
                    model.fit_incremental(*split.train_classes, after_classifier=after_classifier)
            except Exception:
                dataset_stop_reason = traceback.format_exc()
                outcome = DatasetRunOutcome(
                    split=split,
                    status="failed",
                    stop_reason=dataset_stop_reason,
                    fit_seconds=(utc_now() - fit_started_at).total_seconds(),
                    train_metrics=None,
                    test_metrics=None,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )
                outcomes.append(outcome)
                metric_rows.append(build_metric_row(outcome, max_classes=max_classes))
                metadata[f"{split.name}_final_status"] = outcome.status
                metadata[f"{split.name}_stop_reason"] = outcome.stop_reason
                print(dataset_stop_reason, file=tee)
                checkpoint_excel(
                    xlsx_path,
                    tee,
                    f"{split.name} failure",
                    metric_headers=metric_headers,
                    metric_rows=metric_rows,
                    parameter_headers=parameter_headers,
                    parameter_rows=parameter_rows,
                    progress_headers=progress_headers,
                    progress_rows=progress_rows,
                    metadata_rows=build_metadata_rows(metadata),
                )
                continue

            fit_seconds = (utc_now() - fit_started_at).total_seconds()
            if model.is_fully_fitted():
                train_predictions = np.asarray(model.predict(split.X_train), dtype=int)
                test_predictions = np.asarray(model.predict(split.X_test), dtype=int)
                train_metrics = evaluate_multiclass(split.y_train, train_predictions, n_classes=split.n_classes)
                test_metrics = evaluate_multiclass(split.y_test, test_predictions, n_classes=split.n_classes)
                log_metrics(tee, "Training", train_metrics, split.n_classes)
                log_metrics(tee, "Testing", test_metrics, split.n_classes)
                outcome = DatasetRunOutcome(
                    split=split,
                    status="completed",
                    stop_reason=None,
                    fit_seconds=fit_seconds,
                    train_metrics=train_metrics,
                    test_metrics=test_metrics,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )
            else:
                stop_reason = dataset_stop_reason or model.fit_stop_reason or "stopped_before_completion"
                print(f"Run stopped before all classifiers completed: {stop_reason}", file=tee)
                outcome = DatasetRunOutcome(
                    split=split,
                    status="stopped_early",
                    stop_reason=stop_reason,
                    fit_seconds=fit_seconds,
                    train_metrics=None,
                    test_metrics=None,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )

            outcomes.append(outcome)
            metric_rows.append(build_metric_row(outcome, max_classes=max_classes))
            metadata[f"{split.name}_final_status"] = outcome.status
            metadata[f"{split.name}_stop_reason"] = outcome.stop_reason
            checkpoint_excel(
                xlsx_path,
                tee,
                f"{split.name} finalization",
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                progress_headers=progress_headers,
                progress_rows=progress_rows,
                metadata_rows=build_metadata_rows(metadata),
            )

        banner(tee, "Run complete")
        print(f"Log saved to: {log_path}", file=tee)
        print(f"Excel saved to: {xlsx_path}", file=tee)
        print(f"Markdown report target: {report_path}", file=tee)

    report_content = render_markdown_report(
        outcomes=outcomes,
        generated_at=human_timestamp(utc_now()) or "",
        workbook_path=xlsx_path,
        log_path=log_path,
        run_config={
            "threads": args.threads,
            "soft_mem_limit_gb": args.soft_mem_limit_gb,
            "time_limit_label": "none" if args.time_limit is None else str(args.time_limit),
            "mip_gap": args.mip_gap,
            "mip_focus": args.mip_focus,
            "nodefile_start_gb": args.nodefile_start_gb,
            "nodefile_dir": resolved_nodefile_dir,
        },
    )
    report_path.write_text(report_content, encoding="utf-8")
    if any(outcome.status == "failed" for outcome in outcomes):
        return 1
    if all(outcome.status == "completed" or is_expected_early_stop(outcome) for outcome in outcomes):
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
