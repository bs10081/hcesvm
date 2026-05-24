#!/usr/bin/env python3
"""Run HCESVM(test3) on the derived 1000-sample teaching-data datasets."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from hcesvm.models.hierarchical import HierarchicalCESVM
from hcesvm.utils.data_loader import load_csv_ordinal_data
from hcesvm.utils.evaluator import evaluate_multiclass
from hcesvm.utils.teaching_data_1000 import (
    CALIFORNIAHOUSING_1000_RECIPE,
    DERIVED_DATASET_RECIPES,
    SKILL_METHOD3_4CLASS_1000_RECIPE,
    DerivedDatasetRecipe,
    DerivedDatasetSplit,
    build_classifier_diagnostics_row,
    build_classifier_progress_row,
    derive_dataset_split,
    format_scalar,
    solver_status_label,
    source_dataset_url,
    write_derived_split_artifacts,
)
from hcesvm.utils.teaching_data_runtime import format_hcesvm_time_limit_message


MODEL_NAME = "HCESVM(test3)"


class TeeStream:
    """Mirror output to stdout/stderr and the run log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass(slots=True)
class SourceDatasetBundle:
    """Canonical training/test split for one teaching-data source dataset."""

    name: str
    source_url: str
    n_features: int
    X_train: np.ndarray
    y_train: np.ndarray
    X_test: np.ndarray
    y_test: np.ndarray


@dataclass(slots=True)
class DatasetRunOutcome:
    """Final result bundle for one derived dataset run."""

    split: DerivedDatasetSplit
    status: str
    stop_reason: str | None
    fit_seconds: float | None
    train_metrics: dict[str, Any] | None
    test_metrics: dict[str, Any] | None
    diagnostics_rows: list[dict[str, Any]]
    per_classifier_time_limit: int
    time_limit_message: str


def parse_optional_float(value: str) -> float | None:
    """Parse optional floating-point CLI values."""
    if value.lower() in {"none", "null", "off", "disabled"}:
        return None
    return float(value)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run HCESVM(test3) on the derived 1000-sample teaching-data datasets."
    )
    parser.add_argument(
        "--datasets",
        nargs="+",
        choices=sorted(DERIVED_DATASET_RECIPES),
        default=[
            SKILL_METHOD3_4CLASS_1000_RECIPE.name,
            CALIFORNIAHOUSING_1000_RECIPE.name,
        ],
        help="Derived datasets to run. Defaults to both 1000-sample validation sets.",
    )
    parser.add_argument(
        "--time-limit",
        type=int,
        default=1800,
        help="HCESVM per-classifier time limit in seconds. Default: 1800.",
    )
    parser.add_argument(
        "--max-classifiers-per-dataset",
        type=int,
        default=None,
        help="Stop after this many classifiers per dataset. Useful for smoke runs.",
    )
    parser.add_argument(
        "--heartbeat-interval-seconds",
        type=parse_optional_float,
        default=60.0,
        help="Emit solver heartbeats this often. Use 'none' to disable. Default: 60.",
    )
    parser.add_argument("--threads", type=int, default=20, help="Gurobi thread count. Default: 20.")
    parser.add_argument(
        "--soft-mem-limit-gb",
        type=float,
        default=56.0,
        help="Gurobi SoftMemLimit in GB. Default: 56.",
    )
    parser.add_argument("--mip-gap", type=float, default=1e-4, help="Gurobi MIP gap tolerance.")
    parser.add_argument("--C-hyper", type=float, default=1.0, help="Binary CE-SVM slack penalty.")
    parser.add_argument("--M", type=float, default=1000.0, help="Binary CE-SVM big-M constant.")
    return parser.parse_args()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def git_output(*args: str) -> str:
    return subprocess.check_output(args, cwd=ROOT, text=True).strip()


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    print(f"\n{line}", file=tee)
    print(f"[{human_timestamp(utc_now())}] {title}", file=tee)
    print(line, file=tee)


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def load_source_dataset(name: str) -> SourceDatasetBundle:
    """Load the canonical stratified split for one teaching-data source dataset."""
    source_url = source_dataset_url(name)
    train_classes, _, X_test, y_test, n_features = load_csv_ordinal_data(
        source_url,
        target_col="response",
        test_size=0.2,
        random_state=42,
    )
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [
            np.full(len(X_class), class_index, dtype=int)
            for class_index, X_class in enumerate(train_classes, start=1)
        ]
    )
    return SourceDatasetBundle(
        name=name,
        source_url=source_url,
        n_features=n_features,
        X_train=X_train,
        y_train=y_train.astype(int),
        X_test=np.asarray(X_test, dtype=float),
        y_test=np.asarray(y_test, dtype=int),
    )


def prepare_split(
    *,
    recipe: DerivedDatasetRecipe,
    archive_dir: Path,
    timestamp: str,
) -> tuple[DerivedDatasetSplit, dict[str, Path]]:
    """Load the source dataset, derive the smaller split, and persist its artifacts."""
    source_bundle = load_source_dataset(recipe.source_dataset)
    split = derive_dataset_split(
        X_train=source_bundle.X_train,
        y_train=source_bundle.y_train,
        X_test=source_bundle.X_test,
        y_test=source_bundle.y_test,
        source_url=source_bundle.source_url,
        recipe=recipe,
        feature_names=[f"feature_{index}" for index in range(1, source_bundle.n_features + 1)],
    )
    artifact_dir = archive_dir / split.name
    artifacts = write_derived_split_artifacts(split, output_dir=artifact_dir, timestamp=timestamp)
    return split, artifacts


def recipe_n_classes(recipe: DerivedDatasetRecipe) -> int:
    """Return the derived class count for workbook header construction."""
    if recipe.expected_train_counts is not None:
        return len(recipe.expected_train_counts)
    if recipe.class_map is not None:
        return len(recipe.class_map)
    raise ValueError(f"Cannot infer n_classes for recipe {recipe.name}")


def build_metric_headers(max_classes: int) -> list[str]:
    headers = [
        "dataset",
        "source_dataset",
        "model",
        "status",
        "stop_reason",
        "n_classes",
        "n_features",
        "fit_seconds",
        "train_total_accuracy",
        "test_total_accuracy",
        "source_url",
    ]
    for class_index in range(1, max_classes + 1):
        headers.extend(
            [
                f"train_class_{class_index}_count",
                f"train_class_{class_index}_accuracy",
                f"test_class_{class_index}_count",
                f"test_class_{class_index}_accuracy",
            ]
        )
    return headers


def build_metric_row(
    outcome: DatasetRunOutcome,
    *,
    max_classes: int,
) -> dict[str, Any]:
    split = outcome.split
    row: dict[str, Any] = {
        "dataset": split.name,
        "source_dataset": split.source_dataset,
        "model": MODEL_NAME,
        "status": outcome.status,
        "stop_reason": outcome.stop_reason,
        "n_classes": split.n_classes,
        "n_features": split.n_features,
        "fit_seconds": outcome.fit_seconds,
        "train_total_accuracy": None,
        "test_total_accuracy": None,
        "source_url": split.source_url,
    }
    if outcome.train_metrics is not None:
        row["train_total_accuracy"] = format_scalar(outcome.train_metrics["total_accuracy"])
    if outcome.test_metrics is not None:
        row["test_total_accuracy"] = format_scalar(outcome.test_metrics["total_accuracy"])

    for class_index in range(1, max_classes + 1):
        row[f"train_class_{class_index}_count"] = None
        row[f"train_class_{class_index}_accuracy"] = None
        row[f"test_class_{class_index}_count"] = None
        row[f"test_class_{class_index}_accuracy"] = None
        if class_index <= split.n_classes:
            row[f"train_class_{class_index}_count"] = split.train_sampled_counts[class_index - 1]
            row[f"test_class_{class_index}_count"] = split.test_sampled_counts[class_index - 1]
        if outcome.train_metrics is not None and class_index <= split.n_classes:
            row[f"train_class_{class_index}_accuracy"] = format_scalar(
                outcome.train_metrics.get(f"class_{class_index}_accuracy")
            )
        if outcome.test_metrics is not None and class_index <= split.n_classes:
            row[f"test_class_{class_index}_accuracy"] = format_scalar(
                outcome.test_metrics.get(f"class_{class_index}_accuracy")
            )
    return row


def build_parameter_headers() -> list[str]:
    return [
        "dataset",
        "classifier",
        "description",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "mip_gap",
        "solver_status_code",
        "solver_status_label",
        "elapsed_seconds",
        "cumulative_elapsed_seconds",
    ]


def build_progress_headers() -> list[str]:
    return [
        "dataset",
        "classifier",
        "description",
        "started_at_utc",
        "finished_at_utc",
        "positive_sample_count",
        "negative_sample_count",
        "weights",
        "b",
        "objective_value",
        "mip_gap",
        "solver_status_code",
        "solver_status_label",
        "elapsed_seconds",
        "cumulative_elapsed_seconds",
    ]


def build_metadata_rows(metadata: dict[str, Any]) -> list[tuple[str, Any]]:
    return [(key, value) for key, value in metadata.items()]


def write_excel(
    output_path: Path,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    metrics_sheet = workbook.active
    metrics_sheet.title = "Metrics"
    metrics_sheet.append(metric_headers)
    for row in metric_rows:
        metrics_sheet.append([row.get(header) for header in metric_headers])

    parameters_sheet = workbook.create_sheet("Parameters")
    parameters_sheet.append(parameter_headers)
    for row in parameter_rows:
        parameters_sheet.append([row.get(header) for header in parameter_headers])

    progress_sheet = workbook.create_sheet("Progress")
    progress_sheet.append(progress_headers)
    for row in progress_rows:
        progress_sheet.append([row.get(header) for header in progress_headers])

    metadata_sheet = workbook.create_sheet("Run_Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in (metrics_sheet, parameters_sheet, progress_sheet, metadata_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def checkpoint_excel(
    output_path: Path,
    tee: TeeStream,
    label: str,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    progress_headers: list[str],
    progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    write_excel(
        output_path,
        metric_headers=metric_headers,
        metric_rows=metric_rows,
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        progress_headers=progress_headers,
        progress_rows=progress_rows,
        metadata_rows=metadata_rows,
    )
    print(f"Checkpoint Excel saved after {label}: {output_path}", file=tee)


def log_metrics(tee: TeeStream, label: str, metrics: dict[str, Any], n_classes: int) -> None:
    print(f"{label} total accuracy: {metrics['total_accuracy']:.4f}", file=tee)
    for class_index in range(1, n_classes + 1):
        accuracy = metrics[f"class_{class_index}_accuracy"]
        count = metrics[f"class_{class_index}_count"]
        print(f"  Class {class_index}: {accuracy:.4f} ({count} samples)", file=tee)


def format_report_table_row(values: list[str]) -> str:
    return "| " + " | ".join(values) + " |"


def render_accuracy_table(outcome: DatasetRunOutcome) -> str:
    split = outcome.split
    lines = [
        format_report_table_row(["Split", "Total Accuracy", "Per-Class Accuracy"]),
        format_report_table_row(["---", "---:", "---"]),
    ]
    for split_name, metrics, counts in (
        ("Train", outcome.train_metrics, split.train_sampled_counts),
        ("Test", outcome.test_metrics, split.test_sampled_counts),
    ):
        if metrics is None:
            lines.append(format_report_table_row([split_name, "", outcome.stop_reason or ""]))
            continue
        per_class = ", ".join(
            f"C{class_index}={metrics[f'class_{class_index}_accuracy']:.4f} (n={counts[class_index - 1]})"
            for class_index in range(1, split.n_classes + 1)
        )
        lines.append(
            format_report_table_row(
                [split_name, f"{metrics['total_accuracy']:.4f}", per_class]
            )
        )
    return "\n".join(lines)


def render_diagnostics_table(outcome: DatasetRunOutcome) -> str:
    lines = [
        format_report_table_row(
            [
                "Classifier",
                "Description",
                "Samples (+/-)",
                "Status",
                "Elapsed (s)",
                "Objective",
                "mip_gap",
                "b",
                "weights",
            ]
        ),
        format_report_table_row(
            ["---", "---", "---", "---", "---:", "---:", "---:", "---:", "---"]
        ),
    ]
    for row in outcome.diagnostics_rows:
        lines.append(
            format_report_table_row(
                [
                    str(row["classifier"]),
                    str(row["description"]),
                    f"{row['positive_sample_count']}/{row['negative_sample_count']}",
                    str(row["solver_status_label"]),
                    "" if row["elapsed_seconds"] is None else f"{float(row['elapsed_seconds']):.2f}",
                    "" if row["objective_value"] is None else f"{float(row['objective_value']):.6f}",
                    "" if row["mip_gap"] is None else f"{float(row['mip_gap']):.6f}",
                    "" if row["b"] is None else f"{float(row['b']):.6f}",
                    str(row["weights"]),
                ]
            )
        )
    return "\n".join(lines)


def render_markdown_report(
    *,
    outcomes: list[DatasetRunOutcome],
    generated_at: str,
    workbook_path: Path,
    log_path: Path,
) -> str:
    lines = [
        "# HCESVM(test3) Derived 1000-Sample Validation",
        "",
        f"Generated at: `{generated_at}`",
        f"Workbook: `{workbook_path}`",
        f"Log: `{log_path}`",
        "",
    ]
    for outcome in outcomes:
        split = outcome.split
        lines.extend(
            [
                f"## {split.name}",
                "",
                f"- Source dataset: `{split.source_dataset}`",
                f"- Split rule: reuse the existing stratified split, then downsample to `{split.train_target_size}/{split.test_target_size}` independently",
                f"- Train counts: `{split.train_original_counts} -> {split.train_sampled_counts}`",
                f"- Test counts: `{split.test_original_counts} -> {split.test_sampled_counts}`",
                f"- Per-classifier time limit: `{outcome.per_classifier_time_limit}s`",
                f"- Final status: `{outcome.status}`",
            ]
        )
        if outcome.stop_reason:
            lines.append(f"- Stop reason: `{outcome.stop_reason}`")
        lines.extend(["", render_accuracy_table(outcome), "", render_diagnostics_table(outcome), ""])
    return "\n".join(lines)


def is_expected_early_stop(outcome: DatasetRunOutcome) -> bool:
    """Return whether an early stop was explicitly requested by the caller."""
    return outcome.status == "stopped_early" and bool(outcome.stop_reason) and str(
        outcome.stop_reason
    ).startswith("requested stop after")


def main() -> int:
    args = parse_arguments()
    started_at = utc_now()
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")
    selected_recipes = [DERIVED_DATASET_RECIPES[name] for name in args.datasets]
    max_classes = max(recipe_n_classes(recipe) for recipe in selected_recipes)

    archive_dir = ROOT / "results" / "archive" / f"{date_prefix}_test3_teaching_data_1000"
    archive_dir.mkdir(parents=True, exist_ok=True)
    log_path = archive_dir / f"test3_teaching_data_1000_{timestamp}.log"
    xlsx_path = ROOT / "docs" / "reports" / f"HCESVM_TEST3_1000_VALIDATION_{timestamp}.xlsx"
    report_path = ROOT / "docs" / "reports" / f"HCESVM_TEST3_1000_VALIDATION_{timestamp}.md"

    metric_headers = build_metric_headers(max_classes)
    parameter_headers = build_parameter_headers()
    progress_headers = build_progress_headers()
    metric_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    progress_rows: list[dict[str, Any]] = []
    outcomes: list[DatasetRunOutcome] = []

    metadata = {
        "generated_at_utc": human_timestamp(started_at),
        "branch": git_output("git", "rev-parse", "--abbrev-ref", "HEAD"),
        "commit": git_output("git", "rev-parse", "HEAD"),
        "worktree": str(ROOT),
        "log_path": str(log_path),
        "excel_path": str(xlsx_path),
        "report_path": str(report_path),
        "datasets": ", ".join(recipe.name for recipe in selected_recipes),
        "hcesvm_time_limit_seconds": args.time_limit,
        "max_classifiers_per_dataset": args.max_classifiers_per_dataset,
        "threads": args.threads,
        "soft_mem_limit_gb": args.soft_mem_limit_gb,
        "heartbeat_interval_seconds": args.heartbeat_interval_seconds,
    }

    with log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, "HCESVM(test3) derived 1000-sample validation")
        print(f"Branch: {metadata['branch']}", file=tee)
        print(f"Commit: {metadata['commit']}", file=tee)
        print(f"Datasets: {metadata['datasets']}", file=tee)
        print(format_hcesvm_time_limit_message(args.time_limit), file=tee)
        print(f"Threads: {args.threads}", file=tee)
        print(f"SoftMemLimit: {args.soft_mem_limit_gb} GB", file=tee)
        print(f"Heartbeat interval: {args.heartbeat_interval_seconds}", file=tee)

        for recipe in selected_recipes:
            split, artifacts = prepare_split(recipe=recipe, archive_dir=archive_dir, timestamp=timestamp)
            metadata[f"{split.name}_source_dataset"] = split.source_dataset
            metadata[f"{split.name}_source_url"] = split.source_url
            metadata[f"{split.name}_train_counts"] = json.dumps(split.train_sampled_counts, ensure_ascii=False)
            metadata[f"{split.name}_test_counts"] = json.dumps(split.test_sampled_counts, ensure_ascii=False)
            metadata[f"{split.name}_manifest_json"] = str(artifacts["manifest_json"])

            checkpoint_excel(
                xlsx_path,
                tee,
                f"{split.name} dataset loading",
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                progress_headers=progress_headers,
                progress_rows=progress_rows,
                metadata_rows=build_metadata_rows(metadata),
            )

            banner(tee, f"{split.name} | {MODEL_NAME}", major=False)
            print(f"Source dataset: {split.source_dataset}", file=tee)
            print(f"Train counts: {split.train_original_counts} -> {split.train_sampled_counts}", file=tee)
            print(f"Test counts: {split.test_original_counts} -> {split.test_sampled_counts}", file=tee)

            per_classifier_time_limit = int(args.time_limit)
            time_limit_message = format_hcesvm_time_limit_message(per_classifier_time_limit)
            print(time_limit_message, file=tee)
            hcesvm_params = {
                "C_hyper": args.C_hyper,
                "M": args.M,
                "time_limit": per_classifier_time_limit,
                "mip_gap": args.mip_gap,
                "threads": args.threads,
                "soft_mem_limit_gb": args.soft_mem_limit_gb,
                "heartbeat_interval_seconds": args.heartbeat_interval_seconds,
                "retain_raw_solution_arrays": False,
                "release_solver_resources_after_fit": True,
                "verbose": False,
            }
            model = HierarchicalCESVM(
                cesvm_params=hcesvm_params,
                strategy="test3",
                n_classes=split.n_classes,
            )

            dataset_diagnostics: list[dict[str, Any]] = []
            dataset_stop_reason: str | None = None
            fit_started_at = utc_now()

            def after_classifier(progress: dict[str, Any]) -> bool:
                nonlocal dataset_stop_reason
                classifier = model.classifiers[f"h{int(progress['hk'])}"]
                diagnostics_row = build_classifier_diagnostics_row(
                    dataset_name=split.name,
                    progress=progress,
                    classifier=classifier,
                )
                progress_row = build_classifier_progress_row(
                    diagnostics_row=diagnostics_row,
                    progress={
                        "started_at_utc": human_timestamp(progress["started_at_utc"]),
                        "finished_at_utc": human_timestamp(progress["finished_at_utc"]),
                    },
                )
                parameter_rows.append(diagnostics_row)
                progress_rows.append(progress_row)
                dataset_diagnostics.append(diagnostics_row)
                print(
                    f"{diagnostics_row['classifier']} status: {diagnostics_row['solver_status_label']} "
                    f"(code={diagnostics_row['solver_status_code']}), "
                    f"elapsed={diagnostics_row['elapsed_seconds']:.2f}s",
                    file=tee,
                )
                checkpoint_excel(
                    xlsx_path,
                    tee,
                    f"{split.name} / {diagnostics_row['classifier']}",
                    metric_headers=metric_headers,
                    metric_rows=metric_rows,
                    parameter_headers=parameter_headers,
                    parameter_rows=parameter_rows,
                    progress_headers=progress_headers,
                    progress_rows=progress_rows,
                    metadata_rows=build_metadata_rows(metadata),
                )
                if (
                    args.max_classifiers_per_dataset is not None
                    and len(dataset_diagnostics) >= args.max_classifiers_per_dataset
                    and model.completed_classifier_count < model.expected_classifier_count
                ):
                    dataset_stop_reason = (
                        f"requested stop after {args.max_classifiers_per_dataset} classifiers"
                    )
                    return False
                return True

            try:
                with redirect_stdout(tee), redirect_stderr(tee):
                    model.fit_incremental(*split.train_classes, after_classifier=after_classifier)
            except Exception:
                dataset_stop_reason = traceback.format_exc()
                outcome = DatasetRunOutcome(
                    split=split,
                    status="failed",
                    stop_reason=dataset_stop_reason,
                    fit_seconds=(utc_now() - fit_started_at).total_seconds(),
                    train_metrics=None,
                    test_metrics=None,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )
                outcomes.append(outcome)
                metric_rows.append(build_metric_row(outcome, max_classes=max_classes))
                metadata[f"{split.name}_final_status"] = outcome.status
                metadata[f"{split.name}_stop_reason"] = outcome.stop_reason
                print(dataset_stop_reason, file=tee)
                checkpoint_excel(
                    xlsx_path,
                    tee,
                    f"{split.name} failure",
                    metric_headers=metric_headers,
                    metric_rows=metric_rows,
                    parameter_headers=parameter_headers,
                    parameter_rows=parameter_rows,
                    progress_headers=progress_headers,
                    progress_rows=progress_rows,
                    metadata_rows=build_metadata_rows(metadata),
                )
                continue

            fit_seconds = (utc_now() - fit_started_at).total_seconds()
            if model.is_fully_fitted():
                train_predictions = np.asarray(model.predict(split.X_train), dtype=int)
                test_predictions = np.asarray(model.predict(split.X_test), dtype=int)
                train_metrics = evaluate_multiclass(split.y_train, train_predictions, n_classes=split.n_classes)
                test_metrics = evaluate_multiclass(split.y_test, test_predictions, n_classes=split.n_classes)
                log_metrics(tee, "Training", train_metrics, split.n_classes)
                log_metrics(tee, "Testing", test_metrics, split.n_classes)
                outcome = DatasetRunOutcome(
                    split=split,
                    status="completed",
                    stop_reason=None,
                    fit_seconds=fit_seconds,
                    train_metrics=train_metrics,
                    test_metrics=test_metrics,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )
            else:
                stop_reason = dataset_stop_reason or model.fit_stop_reason or "stopped_before_completion"
                print(f"Run stopped before all classifiers completed: {stop_reason}", file=tee)
                outcome = DatasetRunOutcome(
                    split=split,
                    status="stopped_early",
                    stop_reason=stop_reason,
                    fit_seconds=fit_seconds,
                    train_metrics=None,
                    test_metrics=None,
                    diagnostics_rows=list(dataset_diagnostics),
                    per_classifier_time_limit=per_classifier_time_limit,
                    time_limit_message=time_limit_message,
                )

            outcomes.append(outcome)
            metric_rows.append(build_metric_row(outcome, max_classes=max_classes))
            metadata[f"{split.name}_final_status"] = outcome.status
            metadata[f"{split.name}_stop_reason"] = outcome.stop_reason
            checkpoint_excel(
                xlsx_path,
                tee,
                f"{split.name} finalization",
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                progress_headers=progress_headers,
                progress_rows=progress_rows,
                metadata_rows=build_metadata_rows(metadata),
            )

        banner(tee, "Run complete")
        print(f"Log saved to: {log_path}", file=tee)
        print(f"Excel saved to: {xlsx_path}", file=tee)

    report_content = render_markdown_report(
        outcomes=outcomes,
        generated_at=human_timestamp(utc_now()) or "",
        workbook_path=xlsx_path,
        log_path=log_path,
    )
    report_path.write_text(report_content, encoding="utf-8")
    if any(outcome.status == "failed" for outcome in outcomes):
        return 1
    if all(outcome.status == "completed" or is_expected_early_stop(outcome) for outcome in outcomes):
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
