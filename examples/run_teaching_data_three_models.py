#!/usr/bin/env python3
"""Run HCESVM(test3), SVOR, and NPSVOR on teaching-data ordinal datasets."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from hcesvm.models.hierarchical import HierarchicalCESVM
from hcesvm.models.npsvor import NPSVORQP
from hcesvm.models.svor import SVORImplicitQP
from hcesvm.utils.data_loader import load_csv_ordinal_data
from hcesvm.utils.evaluator import evaluate_multiclass
from hcesvm.utils.teaching_data_runtime import format_hcesvm_time_limit_message


DATASET_BASE_URL = "https://raw.githubusercontent.com/gagolews/teaching-data/master/ordinal_regression"
DEFAULT_DATASETS = [
    "cement_strength",
    "bostonhousing_ord",
    "californiahousing",
    "skill",
    "stock_ord",
]
TEST_SIZE = 0.2
RANDOM_STATE = 42
MAX_CLASSES = 7

DEFAULT_HCESVM_PARAMS = {
    "C_hyper": 1.0,
    "M": 1000.0,
    "time_limit": 1800,
    "mip_gap": 1e-4,
    "threads": 2,
    "soft_mem_limit_gb": 18.0,
    "retain_raw_solution_arrays": False,
    "release_solver_resources_after_fit": True,
    "verbose": False,
}
DEFAULT_SVOR_PARAMS = {
    "C": 1.0,
    "add_order_constraints": True,
    "solver_params": {
        "TimeLimit": 1800,
        "MIPGap": 1e-4,
        "OutputFlag": 0,
    },
}
DEFAULT_NPSVOR_PARAMS = {
    "C1": 1.0,
    "C2": 1.0,
    "epsilon": 0.2,
    "prediction_rule": "min_distance",
    "solver_params": {
        "TimeLimit": 1800,
        "MIPGap": 1e-4,
        "OutputFlag": 0,
    },
}


class TeeStream:
    """Mirror output to stdout/stderr and the run log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass(slots=True)
class DatasetBundle:
    name: str
    source_url: str
    n_classes: int
    n_features: int
    train_classes: list[np.ndarray]
    test_classes: list[np.ndarray]
    X_train: np.ndarray
    y_train: np.ndarray
    X_test: np.ndarray
    y_test: np.ndarray

    @property
    def train_counts(self) -> list[int]:
        return [len(X_class) for X_class in self.train_classes]

    @property
    def test_counts(self) -> list[int]:
        return [len(X_class) for X_class in self.test_classes]


@dataclass(slots=True)
class ModelRunResult:
    dataset: str
    model: str
    status: str
    fit_seconds: float | None
    train_metrics: dict[str, Any] | None
    test_metrics: dict[str, Any] | None
    parameter_rows: list[dict[str, Any]]
    error: str | None = None


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run HCESVM(test3), SVOR, and NPSVOR on teaching-data ordinal datasets."
    )
    parser.add_argument(
        "--datasets",
        nargs="+",
        default=DEFAULT_DATASETS,
        choices=DEFAULT_DATASETS,
        help="Datasets to run. Defaults to all teaching-data ordinal datasets.",
    )
    parser.add_argument(
        "--hcesvm-time-limit",
        type=int,
        default=DEFAULT_HCESVM_PARAMS["time_limit"],
        help="HCESVM per-classifier time limit in seconds.",
    )
    parser.add_argument(
        "--svor-time-limit",
        type=int,
        default=DEFAULT_SVOR_PARAMS["solver_params"]["TimeLimit"],
        help="SVOR time limit in seconds.",
    )
    parser.add_argument(
        "--npsvor-time-limit",
        type=int,
        default=DEFAULT_NPSVOR_PARAMS["solver_params"]["TimeLimit"],
        help="NPSVOR per-rank time limit in seconds.",
    )
    return parser.parse_args()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def format_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.floating, float)):
        return float(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    return value


def format_vector(values: Any) -> str:
    if values is None:
        return ""
    array = np.asarray(values, dtype=float).tolist()
    return json.dumps([round(float(value), 10) for value in array], ensure_ascii=False)


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    timestamp = human_timestamp(utc_now())
    print(f"\n{line}", file=tee)
    print(f"[{timestamp}] {title}", file=tee)
    print(line, file=tee)


def git_output(*args: str) -> str:
    return subprocess.check_output(args, cwd=ROOT, text=True).strip()


def load_dataset(name: str) -> DatasetBundle:
    source_url = f"{DATASET_BASE_URL}/{name}.csv"
    train_classes, test_classes, X_test, y_test, n_features = load_csv_ordinal_data(
        source_url,
        target_col="response",
        test_size=TEST_SIZE,
        random_state=RANDOM_STATE,
    )

    n_classes = len(train_classes)
    X_train = np.vstack(train_classes)
    y_train = np.concatenate(
        [np.full(len(X_class), class_index, dtype=int) for class_index, X_class in enumerate(train_classes, start=1)]
    )

    return DatasetBundle(
        name=name,
        source_url=source_url,
        n_classes=n_classes,
        n_features=n_features,
        train_classes=train_classes,
        test_classes=test_classes,
        X_train=X_train,
        y_train=y_train.astype(int),
        X_test=np.asarray(X_test, dtype=float),
        y_test=np.asarray(y_test, dtype=int),
    )


def collect_metric_row(bundle: DatasetBundle, result: ModelRunResult) -> dict[str, Any]:
    row: dict[str, Any] = {
        "dataset": bundle.name,
        "model": result.model,
        "status": result.status,
        "n_classes": bundle.n_classes,
        "n_features": bundle.n_features,
        "fit_seconds": result.fit_seconds,
        "train_total_accuracy": None,
        "test_total_accuracy": None,
        "source_url": bundle.source_url,
    }

    if result.train_metrics is not None:
        row["train_total_accuracy"] = format_scalar(result.train_metrics["total_accuracy"])
    if result.test_metrics is not None:
        row["test_total_accuracy"] = format_scalar(result.test_metrics["total_accuracy"])

    for class_index in range(1, MAX_CLASSES + 1):
        row[f"train_class_{class_index}_count"] = (
            bundle.train_counts[class_index - 1] if class_index <= bundle.n_classes else None
        )
        row[f"test_class_{class_index}_count"] = (
            bundle.test_counts[class_index - 1] if class_index <= bundle.n_classes else None
        )
        row[f"train_class_{class_index}_accuracy"] = None
        row[f"test_class_{class_index}_accuracy"] = None

        if result.train_metrics is not None and class_index <= bundle.n_classes:
            row[f"train_class_{class_index}_accuracy"] = format_scalar(
                result.train_metrics.get(f"class_{class_index}_accuracy")
            )
        if result.test_metrics is not None and class_index <= bundle.n_classes:
            row[f"test_class_{class_index}_accuracy"] = format_scalar(
                result.test_metrics.get(f"class_{class_index}_accuracy")
            )

    if result.error is not None:
        row["error"] = result.error

    return row


def log_metrics(tee: TeeStream, label: str, metrics: dict[str, Any], n_classes: int) -> None:
    print(f"{label} total accuracy: {metrics['total_accuracy']:.4f}", file=tee)
    for class_index in range(1, n_classes + 1):
        accuracy = metrics[f"class_{class_index}_accuracy"]
        count = metrics[f"class_{class_index}_count"]
        print(f"  Class {class_index}: {accuracy:.4f} ({count} samples)", file=tee)


def run_hcesvm(bundle: DatasetBundle, tee: TeeStream, hcesvm_params: dict[str, Any]) -> ModelRunResult:
    model_name = "HCESVM(test3)"
    banner(tee, f"{bundle.name} | {model_name}", major=False)
    start = utc_now()

    try:
        actual_params = hcesvm_params.copy()
        time_limit_message = format_hcesvm_time_limit_message(actual_params.get("time_limit"))
        print(time_limit_message, file=tee)

        model = HierarchicalCESVM(
            cesvm_params=actual_params,
            strategy="test3",
            n_classes=bundle.n_classes,
        )

        with redirect_stdout(tee), redirect_stderr(tee):
            model.fit(*bundle.train_classes)

        train_predictions = np.asarray(model.predict(bundle.X_train), dtype=int)
        test_predictions = np.asarray(model.predict(bundle.X_test), dtype=int)
        train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
        test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)
        fit_seconds = (utc_now() - start).total_seconds()

        log_metrics(tee, "Training", train_metrics, bundle.n_classes)
        log_metrics(tee, "Testing", test_metrics, bundle.n_classes)

        summary = model.get_model_summary()
        parameter_rows: list[dict[str, Any]] = []
        for component_name, component_summary in summary["classifiers"].items():
            classifier = model.classifiers[component_name]
            parameter_rows.append(
                {
                    "dataset": bundle.name,
                    "model": model_name,
                    "component": component_name.upper(),
                    "description": component_summary["description"],
                    "weights": format_vector(classifier.weights),
                    "b": format_scalar(classifier.intercept),
                    "objective_value": format_scalar(component_summary["objective_value"]),
                    "positive_class_accuracy_lb": format_scalar(component_summary["positive_class_accuracy_lb"]),
                    "negative_class_accuracy_lb": format_scalar(component_summary["negative_class_accuracy_lb"]),
                    "mip_gap": format_scalar(component_summary["mip_gap"]),
                }
            )
            print(
                f"{component_name.upper()} weights: {format_vector(classifier.weights)}",
                file=tee,
            )
            print(
                f"{component_name.upper()} b: {classifier.intercept:.10f}",
                file=tee,
            )

        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="success",
            fit_seconds=fit_seconds,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=parameter_rows,
        )
    except Exception:
        error = traceback.format_exc()
        print(error, file=tee)
        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="failed",
            fit_seconds=None,
            train_metrics=None,
            test_metrics=None,
            parameter_rows=[],
            error=error,
        )


def run_svor(bundle: DatasetBundle, tee: TeeStream, svor_params: dict[str, Any]) -> ModelRunResult:
    model_name = "SVOR"
    banner(tee, f"{bundle.name} | {model_name}", major=False)
    start = utc_now()

    try:
        print(
            f"Total SVOR time budget: {svor_params['solver_params']['TimeLimit']}s "
            f"(single optimization problem)",
            file=tee,
        )
        model = SVORImplicitQP(**svor_params)
        with redirect_stdout(tee), redirect_stderr(tee):
            model.fit(bundle.X_train.tolist(), bundle.y_train.tolist())

        train_predictions = np.asarray(model.predict(bundle.X_train.tolist()), dtype=int)
        test_predictions = np.asarray(model.predict(bundle.X_test.tolist()), dtype=int)
        train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
        test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)
        fit_seconds = (utc_now() - start).total_seconds()

        log_metrics(tee, "Training", train_metrics, bundle.n_classes)
        log_metrics(tee, "Testing", test_metrics, bundle.n_classes)
        print(f"Weights (w): {format_vector(model.weights_)}", file=tee)
        print(f"Thresholds (b): {format_vector(model.thresholds_)}", file=tee)

        parameter_rows = [
            {
                "dataset": bundle.name,
                "model": model_name,
                "component": "GLOBAL",
                "description": "Shared hyperplane with ordered thresholds",
                "weights": format_vector(model.weights_),
                "b": format_vector(model.thresholds_),
                "objective_value": format_scalar(model.objective_value_),
                "positive_class_accuracy_lb": None,
                "negative_class_accuracy_lb": None,
                "mip_gap": None,
            }
        ]

        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="success",
            fit_seconds=fit_seconds,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=parameter_rows,
        )
    except Exception:
        error = traceback.format_exc()
        print(error, file=tee)
        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="failed",
            fit_seconds=None,
            train_metrics=None,
            test_metrics=None,
            parameter_rows=[],
            error=error,
        )


def run_npsvor(bundle: DatasetBundle, tee: TeeStream, npsvor_params: dict[str, Any]) -> ModelRunResult:
    model_name = "NPSVOR"
    banner(tee, f"{bundle.name} | {model_name}", major=False)
    start = utc_now()

    try:
        total_time_limit = int(npsvor_params["solver_params"]["TimeLimit"])
        subproblem_count = max(bundle.n_classes, 1)
        per_rank_time_limit = max(1, total_time_limit // subproblem_count)
        actual_params = dict(npsvor_params)
        actual_params["solver_params"] = dict(npsvor_params["solver_params"])
        actual_params["solver_params"]["TimeLimit"] = per_rank_time_limit
        print(
            f"Total NPSVOR time budget: {total_time_limit}s; "
            f"per rank time limit: {per_rank_time_limit}s across {subproblem_count} ranks",
            file=tee,
        )

        model = NPSVORQP(**actual_params)
        with redirect_stdout(tee), redirect_stderr(tee):
            model.fit(bundle.X_train.tolist(), bundle.y_train.tolist())

        train_predictions = np.asarray(model.predict(bundle.X_train.tolist()), dtype=int)
        test_predictions = np.asarray(model.predict(bundle.X_test.tolist()), dtype=int)
        train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
        test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)
        fit_seconds = (utc_now() - start).total_seconds()

        log_metrics(tee, "Training", train_metrics, bundle.n_classes)
        log_metrics(tee, "Testing", test_metrics, bundle.n_classes)

        parameter_rows: list[dict[str, Any]] = []
        for rank_index, (weights, bias) in enumerate(zip(model.hyperplanes_, model.biases_), start=1):
            print(f"Rank {rank_index} weights: {format_vector(weights)}", file=tee)
            print(f"Rank {rank_index} b: {bias:.10f}", file=tee)
            parameter_rows.append(
                {
                    "dataset": bundle.name,
                    "model": model_name,
                    "component": f"RANK_{rank_index}",
                    "description": f"Non-parallel hyperplane for class {rank_index}",
                    "weights": format_vector(weights),
                    "b": format_scalar(bias),
                    "objective_value": format_scalar(model.objective_values_[rank_index - 1]),
                    "positive_class_accuracy_lb": None,
                    "negative_class_accuracy_lb": None,
                    "mip_gap": None,
                }
            )

        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="success",
            fit_seconds=fit_seconds,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=parameter_rows,
        )
    except Exception:
        error = traceback.format_exc()
        print(error, file=tee)
        return ModelRunResult(
            dataset=bundle.name,
            model=model_name,
            status="failed",
            fit_seconds=None,
            train_metrics=None,
            test_metrics=None,
            parameter_rows=[],
            error=error,
        )


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_excel(
    output_path: Path,
    metric_rows: list[dict[str, Any]],
    parameter_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Metrics"
    summary_headers = list(metric_rows[0].keys()) if metric_rows else ["dataset", "model", "status"]
    summary_sheet.append(summary_headers)
    for row in metric_rows:
        summary_sheet.append([row.get(header) for header in summary_headers])

    params_sheet = workbook.create_sheet("Parameters")
    params_headers = list(parameter_rows[0].keys()) if parameter_rows else [
        "dataset",
        "model",
        "component",
        "description",
        "weights",
        "b",
    ]
    params_sheet.append(params_headers)
    for row in parameter_rows:
        params_sheet.append([row.get(header) for header in params_headers])

    metadata_sheet = workbook.create_sheet("Run_Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in (summary_sheet, params_sheet, metadata_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def checkpoint_excel(
    output_path: Path,
    metric_rows: list[dict[str, Any]],
    parameter_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
    tee: TeeStream,
    label: str,
) -> None:
    write_excel(output_path, metric_rows, parameter_rows, metadata_rows)
    print(f"Checkpoint Excel saved after {label}: {output_path}", file=tee)


def main() -> int:
    args = parse_arguments()
    started_at = utc_now()
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")
    archive_dir = ROOT / "results" / "archive" / f"{date_prefix}_three_model_ordinal_regression"
    archive_dir.mkdir(parents=True, exist_ok=True)

    log_path = archive_dir / f"three_model_ordinal_regression_all_{timestamp}.log"
    xlsx_path = ROOT / "docs" / "reports" / f"TEACHING_DATA_THREE_MODEL_COMPARISON_{timestamp}.xlsx"

    datasets = list(args.datasets)
    hcesvm_params = dict(DEFAULT_HCESVM_PARAMS)
    hcesvm_params["time_limit"] = args.hcesvm_time_limit
    svor_params = dict(DEFAULT_SVOR_PARAMS)
    svor_params["solver_params"] = dict(DEFAULT_SVOR_PARAMS["solver_params"])
    svor_params["solver_params"]["TimeLimit"] = args.svor_time_limit
    npsvor_params = dict(DEFAULT_NPSVOR_PARAMS)
    npsvor_params["solver_params"] = dict(DEFAULT_NPSVOR_PARAMS["solver_params"])
    npsvor_params["solver_params"]["TimeLimit"] = args.npsvor_time_limit

    metric_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []

    with log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, "Teaching-data ordinal regression comparison")
        print(f"Branch: {git_output('git', 'rev-parse', '--abbrev-ref', 'HEAD')}", file=tee)
        print(f"Commit: {git_output('git', 'rev-parse', 'HEAD')}", file=tee)
        print("Split rule: Stratified 80/20 train/test split", file=tee)
        print(f"Random state: {RANDOM_STATE}", file=tee)
        print(f"Datasets: {', '.join(datasets)}", file=tee)
        print(f"HCESVM params: {json.dumps(hcesvm_params, ensure_ascii=False)}", file=tee)
        print(f"SVOR params: {json.dumps(svor_params, ensure_ascii=False)}", file=tee)
        print(f"NPSVOR params: {json.dumps(npsvor_params, ensure_ascii=False)}", file=tee)

        metadata_rows = [
            ("generated_at_utc", human_timestamp(started_at)),
            ("branch", git_output("git", "rev-parse", "--abbrev-ref", "HEAD")),
            ("commit", git_output("git", "rev-parse", "HEAD")),
            ("worktree", str(ROOT)),
            ("log_path", str(log_path)),
            ("excel_path", str(xlsx_path)),
            ("dataset_base_url", DATASET_BASE_URL),
            ("datasets", ", ".join(datasets)),
            ("split_rule", "Stratified train_test_split with test_size=0.2"),
            ("random_state", RANDOM_STATE),
            ("hcesvm_params", json.dumps(hcesvm_params, ensure_ascii=False)),
            ("svor_params", json.dumps(svor_params, ensure_ascii=False)),
            ("npsvor_params", json.dumps(npsvor_params, ensure_ascii=False)),
        ]

        checkpoint_excel(xlsx_path, metric_rows, parameter_rows, metadata_rows, tee, "initialization")

        for dataset_name in datasets:
            bundle = load_dataset(dataset_name)
            banner(tee, f"Dataset {bundle.name}")
            print(f"Source: {bundle.source_url}", file=tee)
            print(f"Classes: {bundle.n_classes}", file=tee)
            print(f"Features: {bundle.n_features}", file=tee)
            print(f"Train counts: {bundle.train_counts}", file=tee)
            print(f"Test counts: {bundle.test_counts}", file=tee)

            results = (
                run_hcesvm(bundle, tee, hcesvm_params),
                run_svor(bundle, tee, svor_params),
                run_npsvor(bundle, tee, npsvor_params),
            )

            for result in results:
                metric_rows.append(collect_metric_row(bundle, result))
                parameter_rows.extend(result.parameter_rows)
                checkpoint_excel(
                    xlsx_path,
                    metric_rows,
                    parameter_rows,
                    metadata_rows,
                    tee,
                    f"{bundle.name} / {result.model}",
                )

        banner(tee, "Run complete")
        print(f"Log saved to: {log_path}", file=tee)
        print(f"Excel report target: {xlsx_path}", file=tee)
    write_excel(xlsx_path, metric_rows, parameter_rows, metadata_rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
