#!/usr/bin/env python3
"""Build an Excel comparison between the 1000-sample runs and previous full runs."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "docs" / "reports"
DEFAULT_CURRENT_PATTERN = "HCESVM_TEST3_1000_VALIDATION_*.xlsx"
DEFAULT_BASELINE_PATTERN = "TEACHING_DATA_THREE_MODEL_COMPARISON_*.xlsx"
MODEL_NAME = "HCESVM(test3)"

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
GRAY_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")


@dataclass(frozen=True, slots=True)
class DatasetComparisonConfig:
    """Mapping rules for one current dataset vs its baseline dataset."""

    current_dataset: str
    baseline_dataset: str
    display_name: str
    current_to_baseline_class: dict[int, int]
    current_to_baseline_classifier: dict[str, str]
    total_compare_classes: tuple[int, ...]
    total_note: str


COMPARISON_CONFIGS = {
    "skill_method3_4class_1000": DatasetComparisonConfig(
        current_dataset="skill_method3_4class_1000",
        baseline_dataset="skill",
        display_name="skill_method3_4class_1000 vs skill(full)",
        current_to_baseline_class={1: 4, 2: 5, 3: 6, 4: 7},
        current_to_baseline_classifier={"H1": "H4", "H2": "H5", "H3": "H6"},
        total_compare_classes=(4, 5, 6, 7),
        total_note=(
            "Current run keeps original skill classes 4/5/6/7 only; "
            "baseline comparable total is recomputed from the previous full-skill run "
            "restricted to classes 4-7."
        ),
    ),
    "californiahousing_1000": DatasetComparisonConfig(
        current_dataset="californiahousing_1000",
        baseline_dataset="californiahousing",
        display_name="californiahousing_1000 vs californiahousing(full)",
        current_to_baseline_class={1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6},
        current_to_baseline_classifier={
            "H1": "H1",
            "H2": "H2",
            "H3": "H3",
            "H4": "H4",
            "H5": "H5",
        },
        total_compare_classes=(1, 2, 3, 4, 5, 6),
        total_note="Class set is unchanged, so total accuracy is compared directly.",
    ),
}


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def latest_matching(pattern: str) -> Path:
    matches = sorted(REPORTS_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No workbook matches {pattern!r} in {REPORTS_DIR}")
    return matches[-1]


def sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def metric_record(workbook_path: Path, *, dataset: str) -> dict[str, Any]:
    for row in sheet_records(workbook_path, "Metrics"):
        if row.get("dataset") == dataset and row.get("model") == MODEL_NAME:
            return row
    raise KeyError(f"{dataset} / {MODEL_NAME} not found in {workbook_path}")


def parameter_records(workbook_path: Path, *, dataset: str) -> dict[str, dict[str, Any]]:
    rows = {}
    sheet_name = "Parameters"
    component_key = "classifier" if "1000_VALIDATION" in workbook_path.name else "component"
    for row in sheet_records(workbook_path, sheet_name):
        if row.get("dataset") != dataset:
            continue
        if row.get("model") not in (None, MODEL_NAME):
            continue
        component = row.get(component_key)
        if component is None:
            continue
        rows[str(component)] = row
    return rows


def compare_flag(previous_value: float | None, current_value: float | None) -> str:
    if previous_value is None or current_value is None:
        return "not_available"
    delta = float(current_value) - float(previous_value)
    if abs(delta) <= 1e-12:
        return "same"
    return "better" if delta > 0 else "worse"


def fill_for_flag(flag: str) -> PatternFill:
    if flag == "better":
        return GREEN_FILL
    if flag == "worse":
        return RED_FILL
    if flag == "same":
        return YELLOW_FILL
    return GRAY_FILL


def weighted_total_for_classes(
    metrics_row: dict[str, Any],
    *,
    split: str,
    class_indices: tuple[int, ...],
) -> float:
    weighted_sum = 0.0
    total_count = 0
    for class_index in class_indices:
        count = int(metrics_row[f"{split}_class_{class_index}_count"])
        accuracy = float(metrics_row[f"{split}_class_{class_index}_accuracy"])
        weighted_sum += count * accuracy
        total_count += count
    return weighted_sum / total_count


def build_summary_rows(
    *,
    current_workbook: Path,
    baseline_workbook: Path,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for config in COMPARISON_CONFIGS.values():
        current_metrics = metric_record(current_workbook, dataset=config.current_dataset)
        baseline_metrics = metric_record(baseline_workbook, dataset=config.baseline_dataset)
        for split in ("train", "test"):
            previous_full_total = float(baseline_metrics[f"{split}_total_accuracy"])
            previous_comparable_total = weighted_total_for_classes(
                baseline_metrics,
                split=split,
                class_indices=config.total_compare_classes,
            )
            current_total = float(current_metrics[f"{split}_total_accuracy"])
            flag = compare_flag(previous_comparable_total, current_total)
            rows.append(
                {
                    "dataset": config.display_name,
                    "split": split,
                    "previous_total_accuracy_full_dataset": previous_full_total,
                    "previous_total_accuracy_comparable_scope": previous_comparable_total,
                    "current_total_accuracy_1000": current_total,
                    "delta_vs_comparable_scope": current_total - previous_comparable_total,
                    "improvement_flag": flag,
                    "comparison_note": config.total_note,
                }
            )
    return rows


def build_class_rows(
    *,
    current_workbook: Path,
    baseline_workbook: Path,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for config in COMPARISON_CONFIGS.values():
        current_metrics = metric_record(current_workbook, dataset=config.current_dataset)
        baseline_metrics = metric_record(baseline_workbook, dataset=config.baseline_dataset)
        for split in ("train", "test"):
            for current_class, baseline_class in config.current_to_baseline_class.items():
                previous_accuracy = float(
                    baseline_metrics[f"{split}_class_{baseline_class}_accuracy"]
                )
                current_accuracy = float(current_metrics[f"{split}_class_{current_class}_accuracy"])
                flag = compare_flag(previous_accuracy, current_accuracy)
                rows.append(
                    {
                        "dataset": config.display_name,
                        "split": split,
                        "current_class": (
                            f"C{current_class} (orig {baseline_class})"
                            if config.current_dataset == "skill_method3_4class_1000"
                            else f"C{current_class}"
                        ),
                        "baseline_class": f"Class {baseline_class}",
                        "previous_count_full_dataset": int(
                            baseline_metrics[f"{split}_class_{baseline_class}_count"]
                        ),
                        "current_count_1000": int(
                            current_metrics[f"{split}_class_{current_class}_count"]
                        ),
                        "previous_accuracy": previous_accuracy,
                        "current_accuracy": current_accuracy,
                        "delta": current_accuracy - previous_accuracy,
                        "improvement_flag": flag,
                        "comparison_note": (
                            "Current class is the relabeled version of the previous full-dataset class."
                            if config.current_dataset == "skill_method3_4class_1000"
                            else "Direct class-to-class comparison."
                        ),
                    }
                )
    return rows


def build_classifier_rows(
    *,
    current_workbook: Path,
    baseline_workbook: Path,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for config in COMPARISON_CONFIGS.values():
        current_params = parameter_records(current_workbook, dataset=config.current_dataset)
        baseline_params = parameter_records(baseline_workbook, dataset=config.baseline_dataset)
        for current_classifier, baseline_classifier in config.current_to_baseline_classifier.items():
            current_row = current_params[current_classifier]
            baseline_row = baseline_params[baseline_classifier]
            rows.append(
                {
                    "dataset": config.display_name,
                    "current_classifier": current_classifier,
                    "baseline_classifier": baseline_classifier,
                    "current_description": current_row.get("description"),
                    "baseline_description": baseline_row.get("description"),
                    "current_positive_sample_count": current_row.get("positive_sample_count"),
                    "current_negative_sample_count": current_row.get("negative_sample_count"),
                    "baseline_weights": baseline_row.get("weights"),
                    "baseline_b": baseline_row.get("b"),
                    "baseline_objective_value": baseline_row.get("objective_value"),
                    "baseline_mip_gap": baseline_row.get("mip_gap"),
                    "baseline_status": "not_recorded_in_baseline_workbook",
                    "current_weights": current_row.get("weights"),
                    "current_b": current_row.get("b"),
                    "current_objective_value": current_row.get("objective_value"),
                    "current_mip_gap": current_row.get("mip_gap"),
                    "current_status": current_row.get("solver_status_label"),
                    "current_elapsed_seconds": current_row.get("elapsed_seconds"),
                    "comparison_note": (
                        "Derived-skill classifier is aligned to the previous full-skill classifier on classes 4-7."
                        if config.current_dataset == "skill_method3_4class_1000"
                        else "Direct classifier index comparison."
                    ),
                }
            )
    return rows


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_sheet(workbook: Workbook, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    headers = list(rows[0].keys()) if rows else []
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"

    if "improvement_flag" in headers:
        flag_index = headers.index("improvement_flag") + 1
        for row_index in range(2, worksheet.max_row + 1):
            flag_cell = worksheet.cell(row=row_index, column=flag_index)
            flag_cell.fill = fill_for_flag(str(flag_cell.value))

    autosize_worksheet(worksheet)


def build_notes_rows(current_workbook: Path, baseline_workbook: Path) -> list[tuple[str, str]]:
    return [
        ("generated_at_utc", human_timestamp(utc_now())),
        ("current_workbook", str(current_workbook)),
        ("baseline_workbook", str(baseline_workbook)),
        (
            "skill_total_comparison_rule",
            "Use previous full-skill classes 4-7 only to compute a comparable total accuracy.",
        ),
        (
            "skill_class_mapping",
            "Current C1/C2/C3/C4 correspond to previous full-skill classes 4/5/6/7.",
        ),
        (
            "skill_classifier_mapping",
            "Current H1/H2/H3 correspond to previous full-skill H4/H5/H6.",
        ),
        (
            "californiahousing_comparison_rule",
            "Class set is unchanged; compare totals, classes, and H1-H5 directly.",
        ),
        (
            "improvement_flag_rule",
            "better if current > previous comparable accuracy, worse if current < previous, same if equal.",
        ),
    ]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build an Excel comparison between 1000-sample HCESVM runs and previous full-dataset runs."
    )
    parser.add_argument(
        "--current-workbook",
        type=Path,
        default=None,
        help=f"Current 1000-run workbook. Default: latest {DEFAULT_CURRENT_PATTERN}.",
    )
    parser.add_argument(
        "--baseline-workbook",
        type=Path,
        default=None,
        help=f"Baseline full-run workbook. Default: latest {DEFAULT_BASELINE_PATTERN}.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output workbook path. Default: docs/reports/HCESVM_TEST3_1000_VS_FULL_COMPARISON_<timestamp>.xlsx",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    started_at = utc_now()
    current_workbook = args.current_workbook or latest_matching(DEFAULT_CURRENT_PATTERN)
    baseline_workbook = args.baseline_workbook or latest_matching(DEFAULT_BASELINE_PATTERN)
    output_path = args.output or (
        REPORTS_DIR / f"HCESVM_TEST3_1000_VS_FULL_COMPARISON_{short_timestamp(started_at)}.xlsx"
    )

    summary_rows = build_summary_rows(
        current_workbook=current_workbook,
        baseline_workbook=baseline_workbook,
    )
    class_rows = build_class_rows(
        current_workbook=current_workbook,
        baseline_workbook=baseline_workbook,
    )
    classifier_rows = build_classifier_rows(
        current_workbook=current_workbook,
        baseline_workbook=baseline_workbook,
    )

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    write_sheet(workbook, "Summary", summary_rows)
    write_sheet(workbook, "Class_Accuracy", class_rows)
    write_sheet(workbook, "Classifier_Values", classifier_rows)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["key", "value"])
    for key, value in build_notes_rows(current_workbook, baseline_workbook):
        notes_sheet.append([key, value])
    for cell in notes_sheet[1]:
        cell.font = Font(bold=True)
    notes_sheet.freeze_panes = "A2"
    autosize_worksheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"Current workbook: {current_workbook}")
    print(f"Baseline workbook: {baseline_workbook}")
    print(f"Comparison workbook: {output_path}")
    print(f"Generated at: {human_timestamp(started_at)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
