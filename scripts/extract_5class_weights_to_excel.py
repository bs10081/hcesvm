#!/usr/bin/env python
"""
從參考 Excel 檔案提取 5 分類 datasets 的 weights 和截距項
創建每個 dataset 一個 sheet，每個 sheet 包含三個表格（每個方法一個）
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_dataset_data(df: pd.DataFrame, dataset_name: str) -> Dict:
    """從 long format 的 DataFrame 中提取指定 dataset 的資料"""
    dataset_df = df[df['dataset'] == dataset_name].copy()

    result = {
        'HCESVM': [],
        'NPSVOR': [],
        'SVOR': []
    }

    # 處理 HCESVM
    hcesvm_df = dataset_df[dataset_df['model'] == 'HCESVM(test3)'].copy()
    for _, row in hcesvm_df.iterrows():
        weights = eval(row['weights']) if isinstance(row['weights'], str) else row['weights']
        result['HCESVM'].append({
            'component': row['component'],
            'description': row['description'],
            'weights': weights,
            'b': row['b']
        })

    # 處理 NPSVOR
    npsvor_df = dataset_df[dataset_df['model'] == 'NPSVOR'].copy()
    for _, row in npsvor_df.iterrows():
        weights = eval(row['weights']) if isinstance(row['weights'], str) else row['weights']
        result['NPSVOR'].append({
            'component': row['component'],
            'description': row['description'],
            'weights': weights,
            'b': row['b']
        })

    # 處理 SVOR
    svor_df = dataset_df[dataset_df['model'] == 'SVOR'].copy()
    for _, row in svor_df.iterrows():
        weights = eval(row['weights']) if isinstance(row['weights'], str) else row['weights']
        b_values = eval(row['b']) if isinstance(row['b'], str) else row['b']
        result['SVOR'].append({
            'component': row['component'],
            'description': row['description'],
            'weights': weights,
            'b': b_values
        })

    return result

def create_method_table(method_data: List[Dict], method_name: str, n_features: int) -> pd.DataFrame:
    """為單一方法建立表格

    格式：每個 hyperplane 一欄，rows 為 x1, x2, ..., xn, b
    """
    if not method_data:
        return pd.DataFrame()

    # 確定欄位數量（hyperplanes 數量）
    n_hyperplanes = len(method_data)

    # 建立欄位名稱
    columns = [item['component'] for item in method_data]

    # 建立列索引
    index = [f'x{i+1}' for i in range(n_features)]

    # 特殊處理 SVOR（b 是多個值）
    if method_name == 'SVOR':
        # SVOR 只有一組 weights，多個 thresholds
        b_values = method_data[0]['b']
        n_thresholds = len(b_values) if isinstance(b_values, list) else 1
        index.extend([f'b{i+1}' for i in range(n_thresholds)])
    else:
        index.append('b')

    # 建立 DataFrame
    df = pd.DataFrame(index=index, columns=columns)

    # 填入資料
    for col_idx, item in enumerate(method_data):
        weights = item['weights']
        component = item['component']

        # 填入 weights
        for i, w in enumerate(weights):
            if i < n_features:
                df.loc[f'x{i+1}', component] = round(w, 4)

        # 填入 b
        if method_name == 'SVOR':
            b_values = item['b']
            if isinstance(b_values, list):
                for i, b in enumerate(b_values):
                    df.loc[f'b{i+1}', component] = round(b, 4)
        else:
            df.loc['b', component] = round(item['b'], 4)

    return df

def write_sheet_with_three_tables(wb, sheet_name: str, hcesvm_table: pd.DataFrame,
                                   npsvor_table: pd.DataFrame, svor_table: pd.DataFrame):
    """在一個 sheet 中寫入三個表格"""
    ws = wb.create_sheet(sheet_name)

    current_row = 1

    # Times New Roman 字體
    normal_font = Font(name='Times New Roman', size=11, bold=False)
    header_font = Font(name='Times New Roman', size=11, bold=True)

    # 邊框樣式
    thin_border = Side(style='thin', color='000000')
    top_bottom_border = Border(top=thin_border, bottom=thin_border)
    bottom_border = Border(bottom=thin_border)

    # 寫入三個表格
    tables = [
        ('HCESVM(test3)', hcesvm_table),
        ('NPSVOR', npsvor_table),
        ('SVOR', svor_table)
    ]

    for method_name, table in tables:
        if table.empty:
            continue

        # 方法標題
        ws.cell(current_row, 1, method_name)
        ws.cell(current_row, 1).font = header_font
        current_row += 1

        # 寫入表格
        # Header (欄位名稱)
        ws.cell(current_row, 1, '')  # 左上角空白
        for col_idx, col_name in enumerate(table.columns, start=2):
            cell = ws.cell(current_row, col_idx, col_name)
            cell.font = normal_font
            cell.border = top_bottom_border

        # Index 欄位也要邊框
        ws.cell(current_row, 1).border = top_bottom_border
        current_row += 1

        # 資料列
        for row_idx, (index_name, row_data) in enumerate(table.iterrows()):
            # Index 名稱
            ws.cell(current_row, 1, index_name)
            ws.cell(current_row, 1).font = normal_font

            # 資料
            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(current_row, col_idx, value if pd.notna(value) else '')
                cell.font = normal_font

                # 最後一列加下邊框
                if row_idx == len(table) - 1:
                    cell.border = bottom_border

            # Index 欄位的最後一列也要下邊框
            if row_idx == len(table) - 1:
                ws.cell(current_row, 1).border = bottom_border

            current_row += 1

        # 表格之間空兩列
        current_row += 2

def apply_formatting(file_path: Path):
    """調整欄寬"""
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 設定欄寬
        ws.column_dimensions['A'].width = 15  # Index 欄
        for col_idx in range(2, ws.max_column + 1):
            col_letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[col_letter].width = 18

    wb.save(file_path)

def main():
    # 讀取參考 Excel 檔案
    reference_file = Path('/tmp/reference.xlsx')
    df_weights_b = pd.read_excel(reference_file, sheet_name='weights_b')

    # 要處理的 datasets
    datasets = ['bostonhousing_ord', 'cement_strength', 'stock_ord']

    print(f"處理 {len(datasets)} 個 5-class datasets...")
    print(f"時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 80)

    # 建立輸出檔案
    output_path = Path('/home/justin/lab/hcesvm/docs/reports/weights_comparison_5class.xlsx')
    wb = Workbook()
    wb.remove(wb.active)  # 移除預設的 sheet

    for dataset in datasets:
        print(f"\n處理 Dataset: {dataset}")

        # 提取該 dataset 的資料
        dataset_data = extract_dataset_data(df_weights_b, dataset)

        # 確定特徵數量
        if dataset_data['HCESVM']:
            n_features = len(dataset_data['HCESVM'][0]['weights'])
        elif dataset_data['NPSVOR']:
            n_features = len(dataset_data['NPSVOR'][0]['weights'])
        else:
            n_features = 8  # 預設值

        print(f"  特徵數量: {n_features}")

        # 建立三個表格
        hcesvm_table = create_method_table(dataset_data['HCESVM'], 'HCESVM', n_features)
        npsvor_table = create_method_table(dataset_data['NPSVOR'], 'NPSVOR', n_features)
        svor_table = create_method_table(dataset_data['SVOR'], 'SVOR', n_features)

        print(f"  ✓ HCESVM 表格: {hcesvm_table.shape}")
        print(f"  ✓ NPSVOR 表格: {npsvor_table.shape}")
        print(f"  ✓ SVOR 表格: {svor_table.shape}")

        # 寫入 sheet
        write_sheet_with_three_tables(wb, dataset, hcesvm_table, npsvor_table, svor_table)
        print(f"  ✓ 已寫入 sheet: {dataset}")

    # 儲存檔案
    wb.save(output_path)
    print(f"\n✓ 已儲存 Excel 檔案")

    # 套用格式
    apply_formatting(output_path)
    print(f"✓ 已套用格式設定")

    print("\n" + "=" * 80)
    print(f"✓ 完成！輸出檔案: {output_path}")
    print(f"時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

if __name__ == '__main__':
    main()
