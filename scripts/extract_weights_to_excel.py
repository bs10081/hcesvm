#!/usr/bin/env python
"""
提取 SVOR、NPSVOR、HCESVM 的 weights 和截距項到 Excel
每個 dataset 一個 sheet
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

def extract_svor_weights(log_file: Path, dataset: str) -> Optional[Dict]:
    """提取 SVOR 的 weights 和 thresholds"""
    with open(log_file, 'r') as f:
        content = f.read()

    # 找到對應 dataset 的 SVOR 結果
    pattern = rf"Dataset: {re.escape(dataset)}.*?Testing SVOR on {re.escape(dataset)}.*?SVOR Results:(.*?)(?:={80}|Testing NPSVOR)"
    match = re.search(pattern, content, re.DOTALL)

    if not match:
        return None

    svor_section = match.group(1)

    # 提取 Weights
    weights_match = re.search(r"Weights: \[(.*?)\]", svor_section)
    if not weights_match:
        return None
    weights_str = weights_match.group(1)
    weights = [float(x.strip()) for x in weights_str.split(',')]

    # 提取 Thresholds (對應 H1, H2)
    thresholds_match = re.search(r"Thresholds: \[(.*?)\]", svor_section)
    if not thresholds_match:
        return None
    thresholds_str = thresholds_match.group(1)
    thresholds = [float(x.strip()) for x in thresholds_str.split(',')]

    # SVOR 使用共享權重，不同的 threshold
    # H1 和 H2 都使用相同的 weights，但 threshold 不同
    return {
        'H1_weights': weights,
        'H1_b': -thresholds[0],  # threshold 對應負的 b
        'H2_weights': weights,
        'H2_b': -thresholds[1],
        'H3_weights': None,  # SVOR 只有兩個 hyperplanes
        'H3_b': None
    }

def extract_npsvor_weights(log_file: Path, dataset: str) -> Optional[Dict]:
    """提取 NPSVOR 的 weights 和 b"""
    with open(log_file, 'r') as f:
        content = f.read()

    # 找到對應 dataset 的 NPSVOR 結果
    pattern = rf"Dataset: {re.escape(dataset)}.*?Testing NPSVOR on {re.escape(dataset)}.*?NPSVOR Results:(.*?)(?:={80}|$)"
    match = re.search(pattern, content, re.DOTALL)

    if not match:
        return None

    npsvor_section = match.group(1)

    result = {}

    # 提取三個 hyperplanes
    for i in range(1, 4):
        # Class 1: w=[...], b=...
        class_pattern = rf"Class {i}: w=\[(.*?)\], b=([-\d.]+)"
        class_match = re.search(class_pattern, npsvor_section)

        if class_match:
            weights_str = class_match.group(1)
            weights = [float(x.strip()) for x in weights_str.split(',')]
            b = float(class_match.group(2))

            result[f'H{i}_weights'] = weights
            result[f'H{i}_b'] = b
        else:
            result[f'H{i}_weights'] = None
            result[f'H{i}_b'] = None

    return result

def extract_hcesvm_weights(log_file: Path, dataset: str = None) -> Optional[Dict]:
    """提取 HCESVM 的 H1 和 H2 weights 和 b

    Args:
        log_file: 日誌檔案路徑
        dataset: 如果日誌包含多個 datasets，指定要提取的 dataset 名稱
    """
    with open(log_file, 'r') as f:
        content = f.read()

    # 如果指定了 dataset，先找到對應的 section
    if dataset:
        # 嘗試找到 "Testing Dataset: {dataset}" 的 section
        dataset_pattern = rf"Testing Dataset: {re.escape(dataset)}.*?(?=Testing Dataset:|$)"
        dataset_match = re.search(dataset_pattern, content, re.DOTALL)
        if dataset_match:
            content = dataset_match.group(0)
        else:
            # 如果沒找到，嘗試單一 dataset 的格式 "Dataset: {dataset}"
            dataset_pattern2 = rf"Dataset: {re.escape(dataset)}.*?(?=Dataset:|$)"
            dataset_match2 = re.search(dataset_pattern2, content, re.DOTALL)
            if dataset_match2:
                content = dataset_match2.group(0)

    result = {}

    # 嘗試格式 1: "H1 Complete Weights and Bias"
    h1_pattern = r"H1 Complete Weights and Bias.*?Intercept \(b\): ([-\d.]+).*?Weight vector \(w\) - \d+ features:(.*?)(?:={80}|H2 Complete)"
    h1_match = re.search(h1_pattern, content, re.DOTALL)

    if h1_match:
        h1_b = float(h1_match.group(1))
        h1_weights_section = h1_match.group(2)

        # 提取每個 w[i] = value
        h1_weights = []
        weight_pattern = r"w\[\d+\] = ([-\d.]+)"
        for w_match in re.finditer(weight_pattern, h1_weights_section):
            h1_weights.append(float(w_match.group(1)))

        result['H1_weights'] = h1_weights
        result['H1_b'] = h1_b
    else:
        # 嘗試格式 2: "H1 Solution:"
        h1_pattern2 = r"H1 Solution:.*?Weights \(w\): \[(.*?)\].*?Intercept \(b\): ([-\d.]+)"
        h1_match2 = re.search(h1_pattern2, content, re.DOTALL)

        if h1_match2:
            weights_str = h1_match2.group(1)
            h1_weights = [float(x.strip()) for x in weights_str.split()]
            h1_b = float(h1_match2.group(2))

            result['H1_weights'] = h1_weights
            result['H1_b'] = h1_b
        else:
            result['H1_weights'] = None
            result['H1_b'] = None

    # 提取 H2 - 格式 1
    h2_pattern = r"H2 Complete Weights and Bias.*?Intercept \(b\): ([-\d.]+).*?Weight vector \(w\) - \d+ features:(.*?)(?:={80}|$)"
    h2_match = re.search(h2_pattern, content, re.DOTALL)

    if h2_match:
        h2_b = float(h2_match.group(1))
        h2_weights_section = h2_match.group(2)

        # 提取每個 w[i] = value
        h2_weights = []
        weight_pattern = r"w\[\d+\] = ([-\d.]+)"
        for w_match in re.finditer(weight_pattern, h2_weights_section):
            h2_weights.append(float(w_match.group(1)))

        result['H2_weights'] = h2_weights
        result['H2_b'] = h2_b
    else:
        # 嘗試格式 2: "H2 Solution:"
        h2_pattern2 = r"H2 Solution:.*?Weights \(w\): \[(.*?)\].*?Intercept \(b\): ([-\d.]+)"
        h2_match2 = re.search(h2_pattern2, content, re.DOTALL)

        if h2_match2:
            weights_str = h2_match2.group(1)
            h2_weights = [float(x.strip()) for x in weights_str.split()]
            h2_b = float(h2_match2.group(2))

            result['H2_weights'] = h2_weights
            result['H2_b'] = h2_b
        else:
            result['H2_weights'] = None
            result['H2_b'] = None

    return result

def create_weights_table(svor_data: Dict, npsvor_data: Dict, hcesvm_data: Dict,
                        n_features: int) -> pd.DataFrame:
    """建立 weights 表格"""

    # 建立欄位名稱（移除 SVORIM_H3）
    columns = [
        'HCESVM_H1', 'HCESVM_H2',
        'SVORIM_H1', 'SVORIM_H2',
        'NPSVOR_H1', 'NPSVOR_H2', 'NPSVOR_H3'
    ]

    # 建立列索引：x1, x2, ..., xn, b
    index = [f'x{i+1}' for i in range(n_features)] + ['b']

    # 建立空的 DataFrame
    df = pd.DataFrame(index=index, columns=columns)

    # 填入 HCESVM 資料
    if hcesvm_data and hcesvm_data.get('H1_weights'):
        for i, w in enumerate(hcesvm_data['H1_weights']):
            df.loc[f'x{i+1}', 'HCESVM_H1'] = round(w, 4)
        df.loc['b', 'HCESVM_H1'] = round(hcesvm_data.get('H1_b'), 4)

    if hcesvm_data and hcesvm_data.get('H2_weights'):
        for i, w in enumerate(hcesvm_data['H2_weights']):
            df.loc[f'x{i+1}', 'HCESVM_H2'] = round(w, 4)
        df.loc['b', 'HCESVM_H2'] = round(hcesvm_data.get('H2_b'), 4)

    # 填入 SVOR 資料（只保留 H1 和 H2）
    if svor_data:
        for h in [1, 2]:
            weights = svor_data.get(f'H{h}_weights')
            b = svor_data.get(f'H{h}_b')

            if weights:
                for i, w in enumerate(weights):
                    df.loc[f'x{i+1}', f'SVORIM_H{h}'] = round(w, 4)
                df.loc['b', f'SVORIM_H{h}'] = round(b, 4)

    # 填入 NPSVOR 資料
    if npsvor_data:
        for h in [1, 2, 3]:
            weights = npsvor_data.get(f'H{h}_weights')
            b = npsvor_data.get(f'H{h}_b')

            if weights:
                for i, w in enumerate(weights):
                    df.loc[f'x{i+1}', f'NPSVOR_H{h}'] = round(w, 4)
                df.loc['b', f'NPSVOR_H{h}'] = round(b, 4)

    return df

def apply_formatting(file_path: Path):
    """套用 Excel 格式設定

    規則：
    - 全部表格無邊框
    - 最上面的 row（header）有上下邊框
    - 最下面的 row（b）有下邊框
    - Title 不粗體
    - 字體：Times New Roman
    """
    wb = load_workbook(file_path)

    # Times New Roman 字體（非粗體）
    normal_font = Font(name='Times New Roman', size=11, bold=False)

    # 邊框樣式
    thin_border = Side(style='thin', color='000000')
    top_bottom_border = Border(top=thin_border, bottom=thin_border)
    bottom_border = Border(bottom=thin_border)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 取得最大行數和列數
        max_row = ws.max_row
        max_col = ws.max_column

        # 處理所有儲存格
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row,
                                                     min_col=1, max_col=max_col),
                                       start=1):
            for col_idx, cell in enumerate(row, start=1):
                # 設定字體
                cell.font = normal_font

                # 設定邊框
                if row_idx == 1:
                    # 第一列（header）：上下邊框
                    cell.border = top_bottom_border
                elif row_idx == max_row:
                    # 最後一列（b）：下邊框
                    cell.border = bottom_border
                else:
                    # 其他列：無邊框
                    cell.border = Border()

    # 儲存
    wb.save(file_path)
    print(f"\n✓ 已套用格式設定")

def main():
    # 讀取 CSV 檔案
    csv_path = Path('/home/justin/lab/hcesvm/docs/reports/SVOR_NPSVOR_HCESVM_TEST3_COMPARISON.csv')
    df_comparison = pd.read_csv(csv_path)

    # 取得唯一的 datasets
    datasets = df_comparison['Dataset'].unique()

    print(f"處理 {len(datasets)} 個 datasets...")
    print(f"時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 80)

    # 建立 Excel writer
    output_path = Path('/home/justin/lab/hcesvm/docs/reports/weights_comparison.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        for dataset in datasets:
            print(f"\n處理 Dataset: {dataset}")

            # 取得該 dataset 的資料
            dataset_df = df_comparison[df_comparison['Dataset'] == dataset]

            # 取得每個方法的 source file
            svor_row = dataset_df[dataset_df['Method'] == 'SVOR']
            npsvor_row = dataset_df[dataset_df['Method'] == 'NPSVOR']
            hcesvm_row = dataset_df[dataset_df['Method'] == 'HCESVM(test3)']

            # 確定特徵數量（從 SVOR/NPSVOR 日誌推斷）
            if not svor_row.empty:
                svor_file = Path('/home/justin/lab/hcesvm') / svor_row.iloc[0]['Source File']
                print(f"  SVOR source: {svor_file.name}")

                # 從日誌讀取特徵數量
                with open(svor_file, 'r') as f:
                    content = f.read()
                features_match = re.search(rf"Dataset: {re.escape(dataset)}.*?Features: (\d+)", content, re.DOTALL)
                if features_match:
                    n_features = int(features_match.group(1))
                else:
                    n_features = 6  # 預設值
            else:
                n_features = 6

            print(f"  特徵數量: {n_features}")

            # 提取 SVOR weights
            svor_data = None
            if not svor_row.empty:
                svor_file = Path('/home/justin/lab/hcesvm') / svor_row.iloc[0]['Source File']
                svor_data = extract_svor_weights(svor_file, dataset)
                if svor_data:
                    print(f"  ✓ SVOR weights 提取成功")
                else:
                    print(f"  ✗ SVOR weights 提取失敗")

            # 提取 NPSVOR weights
            npsvor_data = None
            if not npsvor_row.empty:
                npsvor_file = Path('/home/justin/lab/hcesvm') / npsvor_row.iloc[0]['Source File']
                npsvor_data = extract_npsvor_weights(npsvor_file, dataset)
                if npsvor_data:
                    print(f"  ✓ NPSVOR weights 提取成功")
                else:
                    print(f"  ✗ NPSVOR weights 提取失敗")

            # 提取 HCESVM weights
            hcesvm_data = None
            if not hcesvm_row.empty:
                hcesvm_file = Path('/home/justin/lab/hcesvm') / hcesvm_row.iloc[0]['Source File']
                hcesvm_data = extract_hcesvm_weights(hcesvm_file, dataset)
                if hcesvm_data:
                    print(f"  ✓ HCESVM weights 提取成功")
                else:
                    print(f"  ✗ HCESVM weights 提取失敗")

            # 建立表格
            weights_table = create_weights_table(svor_data, npsvor_data, hcesvm_data, n_features)

            # 寫入 Excel sheet
            weights_table.to_excel(writer, sheet_name=dataset)
            print(f"  ✓ 已寫入 sheet: {dataset}")

    # 套用格式設定
    apply_formatting(output_path)

    print("\n" + "=" * 80)
    print(f"✓ 完成！輸出檔案: {output_path}")
    print(f"時間: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

if __name__ == '__main__':
    main()
