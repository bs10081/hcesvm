#!/usr/bin/env python3
"""Wait for teaching-data deadline runs, then analyze accuracy/runtime tradeoffs."""

from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from hcesvm.utils.postrun_analysis import (
    WorkbookRunData,
    class_accuracy_values,
    evaluate_significant_improvement,
    format_duration_label,
    historical_three_model_per_classifier_limit,
    latest_dataset_deadline_report,
    latest_three_model_baseline_workbook,
    load_deadline_run_workbook,
    load_three_model_baseline,
    parse_compact_timestamp,
    recommend_time_limit_seconds,
    zero_accuracy_count,
)


DEFAULT_DATASETS = ["skill", "californiahousing"]


class TeeStream:
    """Mirror output to stdout/stderr and the run log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    print(f"\n{line}", file=tee)
    print(f"[{human_timestamp(utc_now())}] {title}", file=tee)
    print(line, file=tee)


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_excel(
    output_path: Path,
    *,
    summary_rows: list[dict[str, Any]],
    class_rows: list[dict[str, Any]],
    baseline_parameter_rows: list[dict[str, Any]],
    current_parameter_rows: list[dict[str, Any]],
    current_progress_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_headers = list(summary_rows[0].keys()) if summary_rows else []
    summary_sheet.append(summary_headers)
    for row in summary_rows:
        summary_sheet.append([row.get(header) for header in summary_headers])

    class_sheet = workbook.create_sheet("Class_Deltas")
    class_headers = list(class_rows[0].keys()) if class_rows else []
    class_sheet.append(class_headers)
    for row in class_rows:
        class_sheet.append([row.get(header) for header in class_headers])

    baseline_params_sheet = workbook.create_sheet("Baseline_Parameters")
    baseline_param_headers = list(baseline_parameter_rows[0].keys()) if baseline_parameter_rows else []
    baseline_params_sheet.append(baseline_param_headers)
    for row in baseline_parameter_rows:
        baseline_params_sheet.append([row.get(header) for header in baseline_param_headers])

    current_params_sheet = workbook.create_sheet("Current_Parameters")
    current_param_headers = list(current_parameter_rows[0].keys()) if current_parameter_rows else []
    current_params_sheet.append(current_param_headers)
    for row in current_parameter_rows:
        current_params_sheet.append([row.get(header) for header in current_param_headers])

    current_progress_sheet = workbook.create_sheet("Current_Progress")
    progress_headers = list(current_progress_rows[0].keys()) if current_progress_rows else []
    current_progress_sheet.append(progress_headers)
    for row in current_progress_rows:
        current_progress_sheet.append([row.get(header) for header in progress_headers])

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in workbook.worksheets:
        if sheet.max_row >= 1:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Wait for teaching-data deadline runs to finish, then compare them against the old HCESVM baseline."
    )
    parser.add_argument(
        "--datasets",
        nargs="+",
        default=DEFAULT_DATASETS,
        help="Datasets to watch. Default: skill californiahousing",
    )
    parser.add_argument(
        "--started-after",
        required=True,
        help="Only consider deadline reports generated at or after this compact timestamp (YYYYMMDD_HHMMSS).",
    )
    parser.add_argument(
        "--poll-seconds",
        type=int,
        default=60,
        help="Polling interval while waiting for final reports.",
    )
    parser.add_argument(
        "--timeout-hours",
        type=float,
        default=24.0,
        help="Stop waiting after this many hours.",
    )
    return parser.parse_args()


def safe_load_deadline_report(path: Path) -> WorkbookRunData | None:
    try:
        return load_deadline_run_workbook(path)
    except Exception:
        return None


def wait_for_final_reports(
    tee: TeeStream,
    *,
    datasets: list[str],
    reports_dir: Path,
    started_after: datetime,
    poll_seconds: int,
    timeout_hours: float,
) -> dict[str, WorkbookRunData]:
    deadline = utc_now() + timedelta(hours=timeout_hours)
    last_snapshot: dict[str, str] = {}

    while True:
        results: dict[str, WorkbookRunData] = {}
        snapshot: dict[str, str] = {}

        for dataset in datasets:
            path = latest_dataset_deadline_report(
                reports_dir,
                dataset=dataset,
                started_after=started_after,
            )
            if path is None:
                snapshot[dataset] = "pending"
                continue

            run_data = safe_load_deadline_report(path)
            if run_data is None:
                snapshot[dataset] = f"updating:{path.name}"
                continue

            snapshot[dataset] = f"{run_data.final_status}:{path.name}"
            if run_data.is_final:
                results[dataset] = run_data

        if snapshot != last_snapshot:
            print(f"Watch status: {json.dumps(snapshot, ensure_ascii=False, sort_keys=True)}", file=tee)
            last_snapshot = snapshot

        if len(results) == len(datasets):
            return results

        if utc_now() >= deadline:
            missing = sorted(set(datasets) - set(results))
            raise TimeoutError(f"Timed out waiting for final reports: {', '.join(missing)}")

        time.sleep(max(5, poll_seconds))


def metric_delta(
    current_metrics: dict[str, Any] | None,
    baseline_metrics: dict[str, Any],
    key: str,
) -> float | None:
    if current_metrics is None:
        return None
    current_value = current_metrics.get(key)
    baseline_value = baseline_metrics.get(key)
    if current_value is None or baseline_value is None:
        return None
    return float(current_value) - float(baseline_value)


def current_metrics_available(run_data: WorkbookRunData) -> bool:
    if run_data.metrics_row is None:
        return False
    return run_data.metrics_row.get("test_total_accuracy") is not None


def current_test_macro_accuracy(run_data: WorkbookRunData) -> float | None:
    if not current_metrics_available(run_data):
        return None
    values = class_accuracy_values(run_data.metrics_row, split="test", n_classes=run_data.n_classes)
    return None if not values else float(sum(values) / len(values))


def baseline_test_macro_accuracy(run_data: WorkbookRunData) -> float:
    values = class_accuracy_values(run_data.metrics_row, split="test", n_classes=run_data.n_classes)
    return float(sum(values) / len(values))


def build_summary_row(baseline: WorkbookRunData, current: WorkbookRunData) -> dict[str, Any]:
    assert baseline.metrics_row is not None

    dataset = baseline.dataset or ""
    baseline_limit_seconds = historical_three_model_per_classifier_limit(
        baseline.n_classes,
        total_budget_seconds=1800,
    )
    current_limit_seconds = current.configured_time_limit_seconds

    significant_improvement = False
    significance_reason = "latest run has no completed accuracy metrics"
    comparison_details = {
        "baseline_test_macro_accuracy": baseline_test_macro_accuracy(baseline),
        "candidate_test_macro_accuracy": None,
        "delta_test_macro_accuracy": None,
        "baseline_zero_test_classes": zero_accuracy_count(
            baseline.metrics_row, split="test", n_classes=baseline.n_classes
        ),
        "candidate_zero_test_classes": None,
        "zero_test_classes_reduced": None,
        "classes_improved_by_5pp": None,
        "largest_test_class_gain": None,
        "largest_test_class_drop": None,
    }

    if current.metrics_row is not None and current_metrics_available(current):
        significant_improvement, significance_reason, comparison_details = evaluate_significant_improvement(
            baseline_metrics=baseline.metrics_row,
            candidate_metrics=current.metrics_row,
            n_classes=baseline.n_classes,
        )

    delta_test_total = metric_delta(current.metrics_row, baseline.metrics_row, "test_total_accuracy")
    delta_train_total = metric_delta(current.metrics_row, baseline.metrics_row, "train_total_accuracy")
    recommendation_seconds, recommendation_reason = recommend_time_limit_seconds(
        baseline_limit_seconds=baseline_limit_seconds,
        current_limit_seconds=current_limit_seconds,
        progress_rows=current.progress_rows,
        final_status=current.final_status,
        significant_improvement=significant_improvement,
        delta_test_total_accuracy=delta_test_total,
        delta_test_macro_accuracy=comparison_details["delta_test_macro_accuracy"],
        zero_test_classes_reduced=0
        if comparison_details["zero_test_classes_reduced"] is None
        else int(comparison_details["zero_test_classes_reduced"]),
    )
    estimated_model_budget_seconds = recommendation_seconds * max(baseline.n_classes - 1, 1)

    return {
        "dataset": dataset,
        "baseline_workbook": baseline.path.name,
        "current_workbook": current.path.name,
        "baseline_status": baseline.final_status,
        "current_status": current.final_status,
        "baseline_fit_seconds": baseline.metrics_row.get("fit_seconds"),
        "current_fit_seconds": None if current.metrics_row is None else current.metrics_row.get("fit_seconds"),
        "baseline_train_total_accuracy": baseline.metrics_row.get("train_total_accuracy"),
        "current_train_total_accuracy": None
        if current.metrics_row is None
        else current.metrics_row.get("train_total_accuracy"),
        "delta_train_total_accuracy": delta_train_total,
        "baseline_test_total_accuracy": baseline.metrics_row.get("test_total_accuracy"),
        "current_test_total_accuracy": None if current.metrics_row is None else current.metrics_row.get("test_total_accuracy"),
        "delta_test_total_accuracy": delta_test_total,
        "baseline_test_macro_accuracy": comparison_details["baseline_test_macro_accuracy"],
        "current_test_macro_accuracy": comparison_details["candidate_test_macro_accuracy"],
        "delta_test_macro_accuracy": comparison_details["delta_test_macro_accuracy"],
        "baseline_zero_test_classes": comparison_details["baseline_zero_test_classes"],
        "current_zero_test_classes": comparison_details["candidate_zero_test_classes"],
        "zero_test_classes_reduced": comparison_details["zero_test_classes_reduced"],
        "classes_improved_by_5pp": comparison_details["classes_improved_by_5pp"],
        "largest_test_class_gain": comparison_details["largest_test_class_gain"],
        "largest_test_class_drop": comparison_details["largest_test_class_drop"],
        "significant_improvement": significant_improvement,
        "significance_reason": significance_reason,
        "baseline_effective_per_classifier_limit_seconds": baseline_limit_seconds,
        "baseline_effective_per_classifier_limit": format_duration_label(baseline_limit_seconds),
        "current_configured_per_classifier_limit_seconds": current_limit_seconds,
        "current_configured_per_classifier_limit": format_duration_label(current_limit_seconds),
        "recommended_per_classifier_limit_seconds": recommendation_seconds,
        "recommended_per_classifier_limit": format_duration_label(recommendation_seconds),
        "recommended_model_budget_seconds": estimated_model_budget_seconds,
        "recommended_model_budget": format_duration_label(estimated_model_budget_seconds),
        "recommendation_reason": recommendation_reason,
    }


def build_class_rows(baseline: WorkbookRunData, current: WorkbookRunData) -> list[dict[str, Any]]:
    assert baseline.metrics_row is not None
    rows: list[dict[str, Any]] = []
    current_metrics = current.metrics_row or {}

    for split in ("train", "test"):
        for class_index in range(1, baseline.n_classes + 1):
            baseline_accuracy = baseline.metrics_row.get(f"{split}_class_{class_index}_accuracy")
            current_accuracy = current_metrics.get(f"{split}_class_{class_index}_accuracy")
            delta_accuracy = None
            if baseline_accuracy is not None and current_accuracy is not None:
                delta_accuracy = float(current_accuracy) - float(baseline_accuracy)

            rows.append(
                {
                    "dataset": baseline.dataset,
                    "split": split,
                    "class_index": class_index,
                    "baseline_count": baseline.metrics_row.get(f"{split}_class_{class_index}_count"),
                    "current_count": current_metrics.get(f"{split}_class_{class_index}_count"),
                    "baseline_accuracy": baseline_accuracy,
                    "current_accuracy": current_accuracy,
                    "delta_accuracy": delta_accuracy,
                }
            )
    return rows


def annotate_rows(
    rows: list[dict[str, Any]],
    *,
    dataset: str | None,
    source_workbook: Path,
) -> list[dict[str, Any]]:
    annotated: list[dict[str, Any]] = []
    for row in rows:
        enriched = {
            "dataset": dataset,
            "source_workbook": source_workbook.name,
        }
        enriched.update(row)
        annotated.append(enriched)
    return annotated


def build_metadata_rows(
    *,
    output_path: Path,
    baseline_workbook: Path,
    datasets: list[str],
    started_after: datetime,
    poll_seconds: int,
    timeout_hours: float,
) -> list[tuple[str, Any]]:
    return [
        ("generated_at_utc", human_timestamp(utc_now())),
        ("output_path", str(output_path)),
        ("baseline_workbook", str(baseline_workbook)),
        ("datasets", ", ".join(datasets)),
        ("started_after", started_after.strftime("%Y%m%d_%H%M%S")),
        ("poll_seconds", poll_seconds),
        ("timeout_hours", timeout_hours),
        ("significant_total_accuracy_threshold", 0.02),
        ("significant_macro_accuracy_threshold", 0.03),
        ("baseline_hcesvm_total_budget_seconds", 1800),
        ("recommendation_min_balanced_limit_seconds", 900),
        ("recommendation_buffer_rule", "slowest_completed_classifier * 1.15, rounded up to 5 minutes"),
    ]


def main() -> int:
    args = parse_args()
    started_after = parse_compact_timestamp(args.started_after)
    started_at = utc_now()
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")

    archive_dir = ROOT / "results" / "archive" / f"{date_prefix}_postrun_analysis"
    archive_dir.mkdir(parents=True, exist_ok=True)
    log_path = archive_dir / f"teaching_data_postrun_analysis_{timestamp}.log"
    excel_path = ROOT / "docs" / "reports" / f"TEACHING_DATA_POSTRUN_TIME_LIMIT_ANALYSIS_{timestamp}.xlsx"
    reports_dir = ROOT / "docs" / "reports"

    with log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, "Teaching-data post-run watcher")
        print(f"Reports dir: {reports_dir}", file=tee)
        print(f"Datasets: {', '.join(args.datasets)}", file=tee)
        print(f"Started after: {args.started_after}", file=tee)
        print(f"Poll seconds: {args.poll_seconds}", file=tee)
        print(f"Timeout hours: {args.timeout_hours}", file=tee)

        final_reports = wait_for_final_reports(
            tee,
            datasets=args.datasets,
            reports_dir=reports_dir,
            started_after=started_after,
            poll_seconds=args.poll_seconds,
            timeout_hours=args.timeout_hours,
        )

        banner(tee, "Final reports detected", major=False)
        for dataset, run_data in sorted(final_reports.items()):
            print(f"{dataset}: {run_data.final_status} -> {run_data.path}", file=tee)

        baseline_workbook = latest_three_model_baseline_workbook(reports_dir)
        print(f"Baseline workbook: {baseline_workbook}", file=tee)

        summary_rows: list[dict[str, Any]] = []
        class_rows: list[dict[str, Any]] = []
        baseline_parameter_rows: list[dict[str, Any]] = []
        current_parameter_rows: list[dict[str, Any]] = []
        current_progress_rows: list[dict[str, Any]] = []

        for dataset in args.datasets:
            baseline = load_three_model_baseline(baseline_workbook, dataset=dataset)
            current = final_reports[dataset]

            summary_rows.append(build_summary_row(baseline, current))
            class_rows.extend(build_class_rows(baseline, current))
            baseline_parameter_rows.extend(
                annotate_rows(baseline.parameter_rows, dataset=dataset, source_workbook=baseline.path)
            )
            current_parameter_rows.extend(
                annotate_rows(current.parameter_rows, dataset=dataset, source_workbook=current.path)
            )
            current_progress_rows.extend(
                annotate_rows(current.progress_rows, dataset=dataset, source_workbook=current.path)
            )

        metadata_rows = build_metadata_rows(
            output_path=excel_path,
            baseline_workbook=baseline_workbook,
            datasets=args.datasets,
            started_after=started_after,
            poll_seconds=args.poll_seconds,
            timeout_hours=args.timeout_hours,
        )

        write_excel(
            excel_path,
            summary_rows=summary_rows,
            class_rows=class_rows,
            baseline_parameter_rows=baseline_parameter_rows,
            current_parameter_rows=current_parameter_rows,
            current_progress_rows=current_progress_rows,
            metadata_rows=metadata_rows,
        )

        banner(tee, "Recommendation summary", major=False)
        for row in summary_rows:
            print(
                f"{row['dataset']}: significant={row['significant_improvement']} | "
                f"recommended limit={row['recommended_per_classifier_limit']} "
                f"(~model budget {row['recommended_model_budget']})",
                file=tee,
            )
            print(f"  Reason: {row['recommendation_reason']}", file=tee)
            print(f"  Signal: {row['significance_reason']}", file=tee)

        print(f"\nAnalysis log saved to: {log_path}", file=tee)
        print(f"Analysis Excel saved to: {excel_path}", file=tee)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
