#!/usr/bin/env python3
"""Build a simple Excel report for baseline vs reduced 1000-sample accuracy."""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
REPORTS_DIR = ROOT / "docs" / "reports"
HCESVM_MODEL_NAME = "HCESVM(test3)"
FULL_DATASET_MODELS = [HCESVM_MODEL_NAME, "SVOR", "NPSVOR"]
DEFAULT_CURRENT_PATTERN = "HCESVM_TEST3_1000_VALIDATION_*.xlsx"
DEFAULT_BASELINE_PATTERN = "TEACHING_DATA_THREE_MODEL_COMPARISON_*.xlsx"

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def latest_matching(pattern: str) -> Path:
    matches = sorted(REPORTS_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"No workbook matches {pattern!r} in {REPORTS_DIR}")
    return matches[-1]


def sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def metric_record(workbook_path: Path, *, dataset: str, model: str) -> dict[str, Any]:
    for row in sheet_records(workbook_path, "Metrics"):
        if row.get("dataset") == dataset and row.get("model") == model:
            return row
    raise KeyError(f"{dataset} / {model} not found in {workbook_path}")


def weighted_total_for_classes(
    metrics_row: dict[str, Any],
    *,
    split: str,
    class_indices: tuple[int, ...],
) -> float:
    weighted_sum = 0.0
    total_count = 0
    for class_index in class_indices:
        count = int(metrics_row[f"{split}_class_{class_index}_count"])
        accuracy = float(metrics_row[f"{split}_class_{class_index}_accuracy"])
        weighted_sum += count * accuracy
        total_count += count
    return weighted_sum / total_count


def full_dataset_model_rows(metrics_by_model: dict[str, dict[str, Any]], *, n_classes: int) -> list[list[Any]]:
    rows = []
    total_row = ["Total Accuracy"]
    for model in FULL_DATASET_MODELS:
        metrics = metrics_by_model[model]
        total_row.extend(
            [
                float(metrics["train_total_accuracy"]),
                float(metrics["test_total_accuracy"]),
            ]
        )
    rows.append(total_row)

    for class_index in range(1, n_classes + 1):
        row = [f"Class {class_index}"]
        for model in FULL_DATASET_MODELS:
            metrics = metrics_by_model[model]
            row.extend(
                [
                    metrics[f"train_class_{class_index}_accuracy"],
                    metrics[f"test_class_{class_index}_accuracy"],
                ]
            )
        rows.append(row)
    return rows


def compare_flag(baseline_value: float | None, current_value: float | None) -> str:
    if baseline_value is None or current_value is None:
        return ""
    delta = float(current_value) - float(baseline_value)
    if abs(delta) <= 1e-12:
        return "same"
    return "better" if delta > 0 else "worse"


def fill_for_flag(flag: str) -> PatternFill | None:
    if flag == "better":
        return GREEN_FILL
    if flag == "worse":
        return RED_FILL
    if flag == "same":
        return YELLOW_FILL
    return None


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_header(worksheet, row_index: int, title: str) -> int:
    worksheet.cell(row=row_index, column=1, value=title).font = Font(bold=True)
    return row_index + 1


def write_table(
    worksheet,
    row_index: int,
    *,
    headers: list[str],
    rows: list[list[Any]],
) -> int:
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=row_index, column=column_index, value=header).font = Font(bold=True)
    row_index += 1

    flag_column = headers.index("Change") + 1 if "Change" in headers else None
    for row in rows:
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)
        if flag_column is not None:
            flag_value = worksheet.cell(row=row_index, column=flag_column).value
            fill = fill_for_flag("" if flag_value is None else str(flag_value))
            if fill is not None:
                worksheet.cell(row=row_index, column=flag_column).fill = fill
        row_index += 1
    return row_index + 1


def build_overview_sheet(workbook: Workbook, current_workbook: Path, baseline_workbook: Path) -> None:
    worksheet = workbook.active
    worksheet.title = "Overview"
    row_index = 1
    row_index = write_header(worksheet, row_index, "Simple Accuracy Comparison")
    worksheet.cell(row=row_index, column=1, value="Generated at")
    worksheet.cell(row=row_index, column=2, value=human_timestamp(utc_now()))
    row_index += 1
    worksheet.cell(row=row_index, column=1, value="Current workbook")
    worksheet.cell(row=row_index, column=2, value=str(current_workbook))
    row_index += 1
    worksheet.cell(row=row_index, column=1, value="Baseline workbook")
    worksheet.cell(row=row_index, column=2, value=str(baseline_workbook))
    row_index += 2
    worksheet.cell(
        row=row_index,
        column=1,
        value="Reduced columns are only available for HCESVM(test3). SVOR/NPSVOR below are full-dataset baseline only.",
    )
    row_index += 2

    current_skill = metric_record(
        current_workbook,
        dataset="skill_method3_4class_1000",
        model=HCESVM_MODEL_NAME,
    )
    baseline_skill = {
        model: metric_record(baseline_workbook, dataset="skill", model=model)
        for model in FULL_DATASET_MODELS
    }
    current_california = metric_record(
        current_workbook,
        dataset="californiahousing_1000",
        model=HCESVM_MODEL_NAME,
    )
    baseline_california = {
        model: metric_record(baseline_workbook, dataset="californiahousing", model=model)
        for model in FULL_DATASET_MODELS
    }

    overview_rows = [
        [
            "skill (classes 4-7 comparable)",
            HCESVM_MODEL_NAME,
            weighted_total_for_classes(baseline_skill[HCESVM_MODEL_NAME], split="train", class_indices=(4, 5, 6, 7)),
            weighted_total_for_classes(baseline_skill[HCESVM_MODEL_NAME], split="test", class_indices=(4, 5, 6, 7)),
            float(current_skill["train_total_accuracy"]),
            float(current_skill["test_total_accuracy"]),
        ],
        [
            "skill",
            "SVOR",
            float(baseline_skill["SVOR"]["train_total_accuracy"]),
            float(baseline_skill["SVOR"]["test_total_accuracy"]),
            None,
            None,
        ],
        [
            "skill",
            "NPSVOR",
            float(baseline_skill["NPSVOR"]["train_total_accuracy"]),
            float(baseline_skill["NPSVOR"]["test_total_accuracy"]),
            None,
            None,
        ],
        [
            "californiahousing",
            HCESVM_MODEL_NAME,
            float(baseline_california[HCESVM_MODEL_NAME]["train_total_accuracy"]),
            float(baseline_california[HCESVM_MODEL_NAME]["test_total_accuracy"]),
            float(current_california["train_total_accuracy"]),
            float(current_california["test_total_accuracy"]),
        ],
        [
            "californiahousing",
            "SVOR",
            float(baseline_california["SVOR"]["train_total_accuracy"]),
            float(baseline_california["SVOR"]["test_total_accuracy"]),
            None,
            None,
        ],
        [
            "californiahousing",
            "NPSVOR",
            float(baseline_california["NPSVOR"]["train_total_accuracy"]),
            float(baseline_california["NPSVOR"]["test_total_accuracy"]),
            None,
            None,
        ],
    ]
    write_table(
        worksheet,
        row_index,
        headers=["Dataset", "Model", "Baseline Train", "Baseline Test", "Reduced Train", "Reduced Test"],
        rows=overview_rows,
    )
    autosize_worksheet(worksheet)


def build_skill_sheet(workbook: Workbook, current_workbook: Path, baseline_workbook: Path) -> None:
    worksheet = workbook.create_sheet("skill_simple")
    baseline = {
        model: metric_record(baseline_workbook, dataset="skill", model=model)
        for model in FULL_DATASET_MODELS
    }
    current = metric_record(
        current_workbook,
        dataset="skill_method3_4class_1000",
        model=HCESVM_MODEL_NAME,
    )

    row_index = 1
    row_index = write_header(worksheet, row_index, "skill baseline vs reduced")
    worksheet.cell(
        row=row_index,
        column=1,
        value="Current reduced classes map to original skill classes 4/5/6/7.",
    )
    row_index += 2

    baseline_rows = full_dataset_model_rows(baseline, n_classes=7)
    row_index = write_header(worksheet, row_index, "Baseline Full Dataset (All Models)")
    row_index = write_table(
        worksheet,
        row_index,
        headers=[
            "Metric",
            "HCESVM Train",
            "HCESVM Test",
            "SVOR Train",
            "SVOR Test",
            "NPSVOR Train",
            "NPSVOR Test",
        ],
        rows=baseline_rows,
    )

    comparable_baseline_train = weighted_total_for_classes(
        baseline[HCESVM_MODEL_NAME],
        split="train",
        class_indices=(4, 5, 6, 7),
    )
    comparable_baseline_test = weighted_total_for_classes(
        baseline[HCESVM_MODEL_NAME],
        split="test",
        class_indices=(4, 5, 6, 7),
    )
    reduced_rows = [
        [
            "Comparable Total Accuracy",
            comparable_baseline_train,
            comparable_baseline_test,
            float(current["train_total_accuracy"]),
            float(current["test_total_accuracy"]),
            compare_flag(comparable_baseline_test, float(current["test_total_accuracy"])),
        ],
        [
            "Class 4 vs C1",
            baseline[HCESVM_MODEL_NAME]["train_class_4_accuracy"],
            baseline[HCESVM_MODEL_NAME]["test_class_4_accuracy"],
            current["train_class_1_accuracy"],
            current["test_class_1_accuracy"],
            compare_flag(
                float(baseline[HCESVM_MODEL_NAME]["test_class_4_accuracy"]),
                float(current["test_class_1_accuracy"]),
            ),
        ],
        [
            "Class 5 vs C2",
            baseline[HCESVM_MODEL_NAME]["train_class_5_accuracy"],
            baseline[HCESVM_MODEL_NAME]["test_class_5_accuracy"],
            current["train_class_2_accuracy"],
            current["test_class_2_accuracy"],
            compare_flag(
                float(baseline[HCESVM_MODEL_NAME]["test_class_5_accuracy"]),
                float(current["test_class_2_accuracy"]),
            ),
        ],
        [
            "Class 6 vs C3",
            baseline[HCESVM_MODEL_NAME]["train_class_6_accuracy"],
            baseline[HCESVM_MODEL_NAME]["test_class_6_accuracy"],
            current["train_class_3_accuracy"],
            current["test_class_3_accuracy"],
            compare_flag(
                float(baseline[HCESVM_MODEL_NAME]["test_class_6_accuracy"]),
                float(current["test_class_3_accuracy"]),
            ),
        ],
        [
            "Class 7 vs C4",
            baseline[HCESVM_MODEL_NAME]["train_class_7_accuracy"],
            baseline[HCESVM_MODEL_NAME]["test_class_7_accuracy"],
            current["train_class_4_accuracy"],
            current["test_class_4_accuracy"],
            compare_flag(
                float(baseline[HCESVM_MODEL_NAME]["test_class_7_accuracy"]),
                float(current["test_class_4_accuracy"]),
            ),
        ],
    ]
    row_index = write_header(worksheet, row_index, "Reduced 1000-Sample HCESVM vs HCESVM Baseline")
    row_index = write_table(
        worksheet,
        row_index,
        headers=[
            "Metric",
            "Baseline Train",
            "Baseline Test",
            "Reduced Train",
            "Reduced Test",
            "Change",
        ],
        rows=reduced_rows,
    )
    autosize_worksheet(worksheet)


def build_california_sheet(workbook: Workbook, current_workbook: Path, baseline_workbook: Path) -> None:
    worksheet = workbook.create_sheet("california_simple")
    baseline = {
        model: metric_record(baseline_workbook, dataset="californiahousing", model=model)
        for model in FULL_DATASET_MODELS
    }
    current = metric_record(
        current_workbook,
        dataset="californiahousing_1000",
        model=HCESVM_MODEL_NAME,
    )

    row_index = 1
    row_index = write_header(worksheet, row_index, "californiahousing baseline vs reduced")
    worksheet.cell(row=row_index, column=1, value="Classes are unchanged, so this is a direct comparison.")
    row_index += 2

    row_index = write_header(worksheet, row_index, "Baseline Full Dataset (All Models)")
    row_index = write_table(
        worksheet,
        row_index,
        headers=[
            "Metric",
            "HCESVM Train",
            "HCESVM Test",
            "SVOR Train",
            "SVOR Test",
            "NPSVOR Train",
            "NPSVOR Test",
        ],
        rows=full_dataset_model_rows(baseline, n_classes=6),
    )

    rows = [
        [
            "Total Accuracy",
            float(baseline[HCESVM_MODEL_NAME]["train_total_accuracy"]),
            float(baseline[HCESVM_MODEL_NAME]["test_total_accuracy"]),
            float(current["train_total_accuracy"]),
            float(current["test_total_accuracy"]),
            compare_flag(
                float(baseline[HCESVM_MODEL_NAME]["test_total_accuracy"]),
                float(current["test_total_accuracy"]),
            ),
        ]
    ]
    for class_index in range(1, 7):
        rows.append(
            [
                f"Class {class_index}",
                baseline[HCESVM_MODEL_NAME][f"train_class_{class_index}_accuracy"],
                baseline[HCESVM_MODEL_NAME][f"test_class_{class_index}_accuracy"],
                current[f"train_class_{class_index}_accuracy"],
                current[f"test_class_{class_index}_accuracy"],
                compare_flag(
                    float(baseline[HCESVM_MODEL_NAME][f"test_class_{class_index}_accuracy"]),
                    float(current[f"test_class_{class_index}_accuracy"]),
                ),
            ]
        )

    row_index = write_header(worksheet, row_index, "Reduced 1000-Sample HCESVM vs HCESVM Baseline")
    write_table(
        worksheet,
        row_index,
        headers=[
            "Metric",
            "Baseline Train",
            "Baseline Test",
            "Reduced Train",
            "Reduced Test",
            "Change",
        ],
        rows=rows,
    )
    autosize_worksheet(worksheet)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a simplified accuracy-only Excel comparison for the 1000-sample runs."
    )
    parser.add_argument(
        "--current-workbook",
        type=Path,
        default=None,
        help=f"Current workbook. Default: latest {DEFAULT_CURRENT_PATTERN}.",
    )
    parser.add_argument(
        "--baseline-workbook",
        type=Path,
        default=None,
        help=f"Baseline workbook. Default: latest {DEFAULT_BASELINE_PATTERN}.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output path for the simple workbook.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    current_workbook = args.current_workbook or latest_matching(DEFAULT_CURRENT_PATTERN)
    baseline_workbook = args.baseline_workbook or latest_matching(DEFAULT_BASELINE_PATTERN)
    output_path = args.output or (
        REPORTS_DIR / f"HCESVM_TEST3_1000_SIMPLE_ACCURACY_{short_timestamp(utc_now())}.xlsx"
    )

    workbook = Workbook()
    build_overview_sheet(workbook, current_workbook, baseline_workbook)
    build_skill_sheet(workbook, current_workbook, baseline_workbook)
    build_california_sheet(workbook, current_workbook, baseline_workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"Current workbook: {current_workbook}")
    print(f"Baseline workbook: {baseline_workbook}")
    print(f"Simple workbook: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
