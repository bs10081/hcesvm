#!/usr/bin/env python3
"""Finalize Boston HCESVM by merging rerun H1/H2 with prior optimal H3/H4."""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "src"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from examples import run_teaching_data_hcesvm_full as full_runner
from hcesvm.utils.evaluator import evaluate_multiclass


DATASET_NAME = "bostonhousing_ord"
MODEL_NAME = "HCESVM(test3)"
MERGED_WORKBOOK_STEM = "BOSTONHOUSING_ORD_HCESVM_TEST3_FULL_MERGED"


@dataclass(frozen=True, slots=True)
class WorkbookData:
    path: Path
    metrics: list[dict[str, Any]]
    parameters: list[dict[str, Any]]
    progress: list[dict[str, Any]]
    metadata: dict[str, Any]


def _sheet_dicts(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{workbook.properties.title or 'workbook'} missing sheet {sheet_name!r}")
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header) for header in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value is not None for value in row)]


def read_workbook(path: Path) -> WorkbookData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    metadata_rows = _sheet_dicts(workbook, "Run_Metadata")
    metadata = {str(row["key"]): row.get("value") for row in metadata_rows}
    return WorkbookData(
        path=path,
        metrics=_sheet_dicts(workbook, "Metrics"),
        parameters=_sheet_dicts(workbook, "Parameters"),
        progress=_sheet_dicts(workbook, "Progress"),
        metadata=metadata,
    )


def _json_list(value: Any, *, key: str) -> list[int]:
    if isinstance(value, str):
        parsed = json.loads(value)
    elif isinstance(value, (list, tuple)):
        parsed = value
    else:
        raise ValueError(f"Metadata {key!r} is not a JSON/list value: {value!r}")
    return [int(item) for item in parsed]


def _weights(value: Any, *, classifier: str) -> np.ndarray:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{classifier} weights are missing")
    parsed = json.loads(value)
    return np.asarray(parsed, dtype=float)


def _rows_by_classifier(rows: list[dict[str, Any]], *, dataset: str) -> dict[str, dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("dataset") != dataset:
            continue
        classifier = str(row.get("classifier"))
        if classifier in selected:
            raise ValueError(f"Duplicate {dataset} {classifier} rows in Parameters sheet")
        selected[classifier] = dict(row)
    return selected


def _progress_by_classifier(rows: list[dict[str, Any]], *, dataset: str) -> dict[str, dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("dataset") != dataset:
            continue
        selected[str(row.get("classifier"))] = dict(row)
    return selected


def _first_metric(data: WorkbookData, *, dataset: str) -> dict[str, Any] | None:
    for row in data.metrics:
        if row.get("dataset") == dataset:
            return row
    return None


def _metadata_key(dataset: str, suffix: str) -> str:
    return f"{dataset}_{suffix}"


def validate_split_metadata(
    h1_h2: WorkbookData,
    h3_h4: WorkbookData,
    split: full_runner.FullDatasetSplit,
    *,
    dataset: str = DATASET_NAME,
) -> None:
    keys = [
        "source_url",
        "source_type",
        "split_rule",
        "random_state",
        "train_counts",
        "test_counts",
        "train_row_fingerprint",
        "test_row_fingerprint",
        "expected_classifiers",
    ]
    for suffix in keys:
        key = _metadata_key(dataset, suffix)
        left = h1_h2.metadata.get(key)
        right = h3_h4.metadata.get(key)
        if left != right:
            raise ValueError(f"Split metadata mismatch for {key}: {left!r} != {right!r}")

    expected_train_counts = _json_list(
        h1_h2.metadata[_metadata_key(dataset, "train_counts")],
        key=_metadata_key(dataset, "train_counts"),
    )
    expected_test_counts = _json_list(
        h1_h2.metadata[_metadata_key(dataset, "test_counts")],
        key=_metadata_key(dataset, "test_counts"),
    )
    if expected_train_counts != split.train_counts:
        raise ValueError(f"Current train counts mismatch: {split.train_counts} != {expected_train_counts}")
    if expected_test_counts != split.test_counts:
        raise ValueError(f"Current test counts mismatch: {split.test_counts} != {expected_test_counts}")

    train_fingerprint = full_runner.row_fingerprint(split.X_train, split.y_train)
    test_fingerprint = full_runner.row_fingerprint(split.X_test, split.y_test)
    if train_fingerprint != h1_h2.metadata[_metadata_key(dataset, "train_row_fingerprint")]:
        raise ValueError("Current train row fingerprint does not match source workbooks")
    if test_fingerprint != h1_h2.metadata[_metadata_key(dataset, "test_row_fingerprint")]:
        raise ValueError("Current test row fingerprint does not match source workbooks")

    expected_classifiers = int(h1_h2.metadata[_metadata_key(dataset, "expected_classifiers")])
    if expected_classifiers != split.n_classes - 1:
        raise ValueError(
            f"Expected classifier count mismatch: workbook={expected_classifiers}, split={split.n_classes - 1}"
        )

    for data in (h1_h2, h3_h4):
        metric = _first_metric(data, dataset=dataset)
        if metric is None:
            continue
        if int(metric["n_classes"]) != split.n_classes:
            raise ValueError(f"{data.path} n_classes mismatch")
        if int(metric["n_features"]) != split.n_features:
            raise ValueError(f"{data.path} n_features mismatch")


def _expected_sample_counts(split: full_runner.FullDatasetSplit, hk: int) -> tuple[int, int]:
    train_counts = split.train_counts
    positive = sum(train_counts[:hk])
    negative = sum(train_counts[hk:])
    return positive, negative


def validate_classifier_rows(
    rows_by_classifier: dict[str, dict[str, Any]],
    split: full_runner.FullDatasetSplit,
    *,
    required: tuple[str, ...],
    source_name: str,
) -> None:
    for classifier in required:
        if classifier not in rows_by_classifier:
            raise ValueError(f"{source_name} is missing {classifier}")
        row = rows_by_classifier[classifier]
        if row.get("solver_status_label") != "optimal":
            raise ValueError(
                f"{source_name} {classifier} must be optimal, got {row.get('solver_status_label')!r}"
            )
        hk = int(classifier.removeprefix("H"))
        expected_positive, expected_negative = _expected_sample_counts(split, hk)
        if int(row["positive_sample_count"]) != expected_positive:
            raise ValueError(f"{source_name} {classifier} positive_sample_count mismatch")
        if int(row["negative_sample_count"]) != expected_negative:
            raise ValueError(f"{source_name} {classifier} negative_sample_count mismatch")
        if len(_weights(row.get("weights"), classifier=classifier)) != split.n_features:
            raise ValueError(f"{source_name} {classifier} weight count does not match n_features")
        if row.get("b") is None:
            raise ValueError(f"{source_name} {classifier} b is missing")


def recompute_predictions(
    X: np.ndarray,
    parameter_rows: list[dict[str, Any]],
    *,
    n_classes: int,
) -> np.ndarray:
    rows = {str(row["classifier"]): row for row in parameter_rows}
    predictions = np.zeros(len(X), dtype=int)
    remaining_mask = np.ones(len(X), dtype=bool)
    for hk in range(1, n_classes):
        row = rows[f"H{hk}"]
        weights = _weights(row.get("weights"), classifier=f"H{hk}")
        intercept = float(row["b"])
        remaining_indices = np.where(remaining_mask)[0]
        if len(remaining_indices) == 0:
            break
        decision_values = X[remaining_indices] @ weights + intercept
        positive_mask = decision_values >= 0
        classified_indices = remaining_indices[positive_mask]
        predictions[classified_indices] = hk
        remaining_mask[classified_indices] = False
    predictions[remaining_mask] = n_classes
    return predictions


def _copy_progress_row(
    progress_rows: dict[str, dict[str, Any]],
    parameter_row: dict[str, Any],
) -> dict[str, Any]:
    classifier = str(parameter_row["classifier"])
    if classifier in progress_rows:
        row = dict(progress_rows[classifier])
    else:
        row = dict(parameter_row)
        row["started_at_utc"] = None
        row["finished_at_utc"] = None
    return row


def _recompute_cumulative_elapsed(rows: list[dict[str, Any]]) -> None:
    cumulative = 0.0
    for row in rows:
        elapsed = row.get("elapsed_seconds")
        if elapsed is None:
            row["cumulative_elapsed_seconds"] = None
            continue
        cumulative += float(elapsed)
        row["cumulative_elapsed_seconds"] = cumulative


def git_output_safe(*args: str) -> str:
    try:
        return full_runner.git_output(*args)
    except Exception:
        return "unknown"


def finalize_merge(
    *,
    h1_h2_workbook: Path,
    h3_h4_workbook: Path,
    output_root: Path = ROOT,
    timestamp: str | None = None,
    split: full_runner.FullDatasetSplit | None = None,
    dataset: str = DATASET_NAME,
) -> dict[str, Path]:
    generated_at = full_runner.utc_now()
    timestamp = timestamp or full_runner.short_timestamp(generated_at)
    date_prefix = timestamp.split("_", 1)[0]
    split = split or full_runner.load_dataset(dataset)

    h1_h2 = read_workbook(h1_h2_workbook)
    h3_h4 = read_workbook(h3_h4_workbook)
    validate_split_metadata(h1_h2, h3_h4, split, dataset=dataset)

    h1_h2_rows = _rows_by_classifier(h1_h2.parameters, dataset=dataset)
    h3_h4_rows = _rows_by_classifier(h3_h4.parameters, dataset=dataset)
    validate_classifier_rows(h1_h2_rows, split, required=("H1", "H2"), source_name="H1/H2 workbook")
    validate_classifier_rows(h3_h4_rows, split, required=("H3", "H4"), source_name="H3/H4 workbook")

    parameter_rows = [
        dict(h1_h2_rows["H1"]),
        dict(h1_h2_rows["H2"]),
        dict(h3_h4_rows["H3"]),
        dict(h3_h4_rows["H4"]),
    ]
    _recompute_cumulative_elapsed(parameter_rows)

    h1_h2_progress = _progress_by_classifier(h1_h2.progress, dataset=dataset)
    h3_h4_progress = _progress_by_classifier(h3_h4.progress, dataset=dataset)
    progress_rows = [
        _copy_progress_row(h1_h2_progress, parameter_rows[0]),
        _copy_progress_row(h1_h2_progress, parameter_rows[1]),
        _copy_progress_row(h3_h4_progress, parameter_rows[2]),
        _copy_progress_row(h3_h4_progress, parameter_rows[3]),
    ]
    _recompute_cumulative_elapsed(progress_rows)

    train_predictions = recompute_predictions(split.X_train, parameter_rows, n_classes=split.n_classes)
    test_predictions = recompute_predictions(split.X_test, parameter_rows, n_classes=split.n_classes)
    train_metrics = evaluate_multiclass(split.y_train, train_predictions, n_classes=split.n_classes)
    test_metrics = evaluate_multiclass(split.y_test, test_predictions, n_classes=split.n_classes)

    outcome = full_runner.DatasetRunOutcome(
        split=split,
        status="completed",
        stop_reason=None,
        fit_seconds=sum(float(row["elapsed_seconds"]) for row in parameter_rows if row.get("elapsed_seconds") is not None),
        train_metrics=train_metrics,
        test_metrics=test_metrics,
        diagnostics_rows=parameter_rows,
        per_classifier_time_limit=None,
        time_limit_message="HCESVM per-classifier time limit: none",
    )
    metric_headers = full_runner.build_metric_headers(split.n_classes)
    parameter_headers = full_runner.build_parameter_headers()
    progress_headers = full_runner.build_progress_headers()
    metric_rows = [full_runner.build_metric_row(outcome, max_classes=split.n_classes)]

    archive_dir = output_root / "results" / "archive" / f"{date_prefix}_test3_boston_h1h2_merge"
    report_dir = output_root / "docs" / "reports"
    log_path = archive_dir / f"test3_{dataset}_h1h2_merge_{timestamp}.log"
    xlsx_path = report_dir / f"{MERGED_WORKBOOK_STEM}_{timestamp}.xlsx"
    report_path = report_dir / f"{MERGED_WORKBOOK_STEM}_{timestamp}.md"

    h1_h2_meta = h1_h2.metadata
    h3_h4_meta = h3_h4.metadata
    metadata: dict[str, Any] = {
        "generated_at_utc": full_runner.human_timestamp(generated_at),
        "branch": git_output_safe("git", "rev-parse", "--abbrev-ref", "HEAD"),
        "commit": git_output_safe("git", "rev-parse", "HEAD"),
        "worktree": str(ROOT),
        "log_path": str(log_path),
        "excel_path": str(xlsx_path),
        "report_path": str(report_path),
        "dataset_base_url": full_runner.DATASET_BASE_URL,
        "datasets": dataset,
        "split_rule": "per-dataset; see <dataset>_split_rule",
        "merge_strategy": "H1/H2 from rerun workbook, H3/H4 from prior optimal workbook",
        "h1_h2_source_workbook": str(h1_h2.path),
        "h3_h4_source_workbook": str(h3_h4.path),
        "h1_h2_source_generated_at_utc": h1_h2_meta.get("generated_at_utc"),
        "h3_h4_source_generated_at_utc": h3_h4_meta.get("generated_at_utc"),
        "h1_h2_source_commit": h1_h2_meta.get("commit"),
        "h3_h4_source_commit": h3_h4_meta.get("commit"),
        "h1_h2_source_mip_focus": h1_h2_meta.get("mip_focus"),
        "h1_h2_source_mip_gap": h1_h2_meta.get("mip_gap"),
        "h1_h2_source_threads": h1_h2_meta.get("threads"),
        "h1_h2_source_soft_mem_limit_gb": h1_h2_meta.get("soft_mem_limit_gb"),
        "h1_h2_source_nodefile_start_gb": h1_h2_meta.get("nodefile_start_gb"),
        "h1_h2_source_nodefile_dir": h1_h2_meta.get("nodefile_dir"),
        "merged_classifiers": "H1,H2,H3,H4",
        "hcesvm_time_limit_seconds": h1_h2_meta.get("hcesvm_time_limit_seconds"),
        "hcesvm_time_limit_label": h1_h2_meta.get("hcesvm_time_limit_label"),
        "max_classifiers_per_dataset": None,
        "threads": h1_h2_meta.get("threads"),
        "soft_mem_limit_gb": h1_h2_meta.get("soft_mem_limit_gb"),
        "mip_gap": h1_h2_meta.get("mip_gap"),
        "mip_focus": h1_h2_meta.get("mip_focus"),
        "nodefile_start_gb": h1_h2_meta.get("nodefile_start_gb"),
        "nodefile_dir_requested": h1_h2_meta.get("nodefile_dir_requested"),
        "nodefile_dir": h1_h2_meta.get("nodefile_dir"),
        "heartbeat_interval_seconds": h1_h2_meta.get("heartbeat_interval_seconds"),
        "feat_lower_bound": h1_h2_meta.get("feat_lower_bound"),
        f"{dataset}_source_url": split.source_url,
        f"{dataset}_source_type": h1_h2_meta.get(f"{dataset}_source_type"),
        f"{dataset}_split_rule": split.split_rule,
        f"{dataset}_random_state": split.random_state,
        f"{dataset}_train_counts": json.dumps(split.train_counts, ensure_ascii=False),
        f"{dataset}_test_counts": json.dumps(split.test_counts, ensure_ascii=False),
        f"{dataset}_train_row_fingerprint": full_runner.row_fingerprint(split.X_train, split.y_train),
        f"{dataset}_test_row_fingerprint": full_runner.row_fingerprint(split.X_test, split.y_test),
        f"{dataset}_expected_classifiers": split.n_classes - 1,
        f"{dataset}_final_status": "completed",
        f"{dataset}_stop_reason": None,
    }

    archive_dir.mkdir(parents=True, exist_ok=True)
    full_runner.write_excel(
        xlsx_path,
        metric_headers=metric_headers,
        metric_rows=metric_rows,
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        progress_headers=progress_headers,
        progress_rows=progress_rows,
        metadata_rows=full_runner.build_metadata_rows(metadata),
    )

    report_content = full_runner.render_markdown_report(
        outcomes=[outcome],
        generated_at=full_runner.human_timestamp(generated_at) or "",
        workbook_path=xlsx_path,
        log_path=log_path,
        run_config={
            "threads": metadata["threads"],
            "soft_mem_limit_gb": metadata["soft_mem_limit_gb"],
            "time_limit_label": metadata["hcesvm_time_limit_label"],
            "mip_gap": metadata["mip_gap"],
            "mip_focus": metadata["mip_focus"],
            "nodefile_start_gb": metadata["nodefile_start_gb"],
            "nodefile_dir": metadata["nodefile_dir"],
        },
    )
    report_content += (
        "\n## Merge Sources\n\n"
        f"- H1/H2 workbook: `{h1_h2.path}`\n"
        f"- H3/H4 workbook: `{h3_h4.path}`\n"
        "- Accuracy was recomputed from merged weights and b.\n"
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(report_content, encoding="utf-8")

    with log_path.open("w", encoding="utf-8") as handle:
        print(f"[{full_runner.human_timestamp(generated_at)}] Boston H1/H2 merge finalize", file=handle)
        print(f"H1/H2 workbook: {h1_h2.path}", file=handle)
        print(f"H3/H4 workbook: {h3_h4.path}", file=handle)
        print("Validated split metadata, class counts, feature count, and optimal solver statuses.", file=handle)
        full_runner.log_metrics(handle, "Training", train_metrics, split.n_classes)
        full_runner.log_metrics(handle, "Testing", test_metrics, split.n_classes)
        for row in parameter_rows:
            print(
                f"{row['classifier']} {row['solver_status_label']} "
                f"b={row['b']} weights={row['weights']}",
                file=handle,
            )
        print(f"Excel saved to: {xlsx_path}", file=handle)
        print(f"Markdown report saved to: {report_path}", file=handle)

    return {"workbook": xlsx_path, "report": report_path, "log": log_path}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge optimal Boston H1/H2 rerun rows with prior optimal H3/H4 rows."
    )
    parser.add_argument("--h1-h2-workbook", required=True, type=Path)
    parser.add_argument("--h3-h4-workbook", required=True, type=Path)
    parser.add_argument("--timestamp", default=None)
    parser.add_argument("--dataset", default=DATASET_NAME, choices=[DATASET_NAME])
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    outputs = finalize_merge(
        h1_h2_workbook=args.h1_h2_workbook,
        h3_h4_workbook=args.h3_h4_workbook,
        timestamp=args.timestamp,
        dataset=args.dataset,
    )
    print(f"Workbook: {outputs['workbook']}")
    print(f"Report: {outputs['report']}")
    print(f"Log: {outputs['log']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
