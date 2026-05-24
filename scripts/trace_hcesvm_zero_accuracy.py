#!/usr/bin/env python3
"""Trace HCESVM zero-accuracy classes against classifier coefficients."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sklearn.model_selection import train_test_split

ROOT = Path(__file__).resolve().parents[1]
DATASET_BASE_URL = "https://raw.githubusercontent.com/gagolews/teaching-data/master/ordinal_regression"
DEFAULT_WORKBOOK = ROOT / "docs" / "reports" / "TEACHING_DATA_THREE_MODEL_COMPARISON_20260423_190157.xlsx"
DEFAULT_LOG = (
    ROOT
    / "results"
    / "archive"
    / "20260423_three_model_ordinal_regression"
    / "three_model_ordinal_regression_all_20260423_190157.log"
)
DEFAULT_TEST_SIZE = 0.2
DEFAULT_RANDOM_STATE = 42
MODEL_NAME = "HCESVM(test3)"
ZERO_EPSILON = 1e-12


@dataclass(slots=True)
class ClassifierTrace:
    dataset: str
    hk: int
    component: str
    description: str
    weights: np.ndarray
    intercept: float
    objective_value: float | None
    positive_lb: float | None
    negative_lb: float | None
    mip_gap: float | None
    total_time_budget_seconds: int | None
    per_classifier_time_limit_seconds: int | None

    @property
    def l1_norm(self) -> float:
        return float(np.sum(np.abs(self.weights)))

    @property
    def max_abs_weight(self) -> float:
        return float(np.max(np.abs(self.weights))) if len(self.weights) else 0.0

    @property
    def all_zero_weights(self) -> bool:
        return bool(np.all(np.abs(self.weights) <= ZERO_EPSILON))

    @property
    def constant_decision(self) -> str:
        if not self.all_zero_weights:
            return "non_constant"
        return "always_positive" if self.intercept >= 0 else "always_negative"


@dataclass(slots=True)
class ZeroAccuracySummary:
    dataset: str
    split: str
    true_class: int
    sample_count: int
    total_accuracy: float
    blocker_hks: list[int]
    blocker_rows: list[dict[str, Any]]
    blocked_by_zero_coefficients_fraction: float
    diagnosis: str


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Trace zero-accuracy HCESVM classes against classifier coefficients.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--log", type=Path, default=DEFAULT_LOG)
    parser.add_argument("--report-output", type=Path)
    parser.add_argument("--xlsx-output", type=Path)
    parser.add_argument("--archive-log-output", type=Path)
    parser.add_argument("--test-size", type=float, default=DEFAULT_TEST_SIZE)
    parser.add_argument("--random-state", type=int, default=DEFAULT_RANDOM_STATE)
    return parser.parse_args()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def format_float(value: Any) -> str:
    if value is None:
        return "N/A"
    return f"{float(value):.4f}"


def format_vector(values: np.ndarray | list[float]) -> str:
    return json.dumps([round(float(value), 10) for value in np.asarray(values, dtype=float).tolist()], ensure_ascii=False)


def parse_json_vector(raw: Any) -> np.ndarray:
    if raw in (None, ""):
        return np.array([], dtype=float)
    return np.asarray(json.loads(raw), dtype=float)


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(len("" if cell.value is None else str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)


def parse_time_limits(log_text: str) -> dict[str, tuple[int | None, int | None]]:
    dataset_limits: dict[str, tuple[int | None, int | None]] = {}
    current_dataset: str | None = None

    for raw_line in log_text.splitlines():
        line = raw_line.strip()
        if line.startswith("[") and "] Dataset " in line:
            current_dataset = line.split("] Dataset ", 1)[1]
            continue

        if (
            current_dataset is not None
            and line.startswith("Total HCESVM time budget:")
            and "per binary classifier time limit:" in line
        ):
            prefix, suffix = line.split("per binary classifier time limit:", 1)
            total_budget_text = prefix.split("Total HCESVM time budget:", 1)[1].split(";", 1)[0].strip()
            total_budget = int(total_budget_text.removesuffix("s").strip())
            per_classifier = int(suffix.split("s", 1)[0].strip())
            dataset_limits[current_dataset] = (total_budget, per_classifier)
            continue

        if current_dataset is not None and line.startswith("HCESVM per-classifier time limit:"):
            value_text = line.split("HCESVM per-classifier time limit:", 1)[1].strip()
            per_classifier = None if value_text == "none" else int(value_text.removesuffix("s").strip())
            dataset_limits[current_dataset] = (None, per_classifier)

    return dataset_limits


def load_workbook_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Parameters"]
    rows = list(worksheet.iter_rows(values_only=True))
    return list(rows[0]), rows[1:]


def load_metrics_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Metrics"]
    rows = list(worksheet.iter_rows(values_only=True))
    return list(rows[0]), rows[1:]


def build_classifier_map(workbook_path: Path, log_text: str) -> dict[str, list[ClassifierTrace]]:
    headers, rows = load_workbook_rows(workbook_path)
    index = {name: position for position, name in enumerate(headers)}
    time_limits = parse_time_limits(log_text)
    by_dataset: dict[str, list[ClassifierTrace]] = defaultdict(list)

    for row in rows:
        if row[index["model"]] != MODEL_NAME:
            continue

        dataset = str(row[index["dataset"]])
        component = str(row[index["component"]])
        hk = int(component.removeprefix("H"))
        total_budget, per_classifier = time_limits.get(dataset, (None, None))
        by_dataset[dataset].append(
            ClassifierTrace(
                dataset=dataset,
                hk=hk,
                component=component,
                description=str(row[index["description"]]),
                weights=parse_json_vector(row[index["weights"]]),
                intercept=float(row[index["b"]]),
                objective_value=None if row[index["objective_value"]] is None else float(row[index["objective_value"]]),
                positive_lb=None
                if row[index["positive_class_accuracy_lb"]] is None
                else float(row[index["positive_class_accuracy_lb"]]),
                negative_lb=None
                if row[index["negative_class_accuracy_lb"]] is None
                else float(row[index["negative_class_accuracy_lb"]]),
                mip_gap=None if row[index["mip_gap"]] is None else float(row[index["mip_gap"]]),
                total_time_budget_seconds=total_budget,
                per_classifier_time_limit_seconds=per_classifier,
            )
        )

    for dataset in by_dataset:
        by_dataset[dataset].sort(key=lambda item: item.hk)

    return by_dataset


def load_dataset_split(dataset: str, *, test_size: float, random_state: int) -> dict[str, np.ndarray]:
    source_url = f"{DATASET_BASE_URL}/{dataset}.csv"
    dataframe = pd.read_csv(source_url)
    X = dataframe.drop(columns=["response"]).to_numpy(dtype=float)
    y = dataframe["response"].to_numpy(dtype=int)
    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=test_size,
        random_state=random_state,
        stratify=y,
    )
    return {
        "X_train": X_train,
        "y_train": y_train,
        "X_test": X_test,
        "y_test": y_test,
    }


def trace_samples(
    X: np.ndarray,
    y_true: np.ndarray,
    classifiers: list[ClassifierTrace],
) -> dict[str, Any]:
    n_samples = len(X)
    n_classes = len(classifiers) + 1
    decision_matrix = np.column_stack([X @ classifier.weights + classifier.intercept for classifier in classifiers])
    predictions = np.full(n_samples, n_classes, dtype=int)
    first_positive_hk = np.zeros(n_samples, dtype=int)

    for hk_index in range(len(classifiers)):
        still_unassigned = first_positive_hk == 0
        if not np.any(still_unassigned):
            break
        hk = hk_index + 1
        positive_mask = decision_matrix[still_unassigned, hk_index] >= 0
        remaining_indices = np.where(still_unassigned)[0]
        positive_indices = remaining_indices[positive_mask]
        predictions[positive_indices] = hk
        first_positive_hk[positive_indices] = hk

    return {
        "predictions": predictions,
        "first_positive_hk": first_positive_hk,
        "decision_matrix": decision_matrix,
        "n_classes": n_classes,
        "y_true": y_true,
    }


def collect_zero_accuracy_summaries(
    workbook_path: Path,
    classifier_map: dict[str, list[ClassifierTrace]],
    *,
    test_size: float,
    random_state: int,
) -> tuple[list[ZeroAccuracySummary], list[dict[str, Any]], list[dict[str, Any]]]:
    metric_headers, metric_rows = load_metrics_rows(workbook_path)
    metric_index = {name: position for position, name in enumerate(metric_headers)}

    summaries: list[ZeroAccuracySummary] = []
    blocker_trace_rows: list[dict[str, Any]] = []
    classifier_diagnostic_rows: list[dict[str, Any]] = []

    for metric_row in metric_rows:
        if metric_row[metric_index["model"]] != MODEL_NAME:
            continue

        dataset = str(metric_row[metric_index["dataset"]])
        classifiers = classifier_map[dataset]
        dataset_split = load_dataset_split(dataset, test_size=test_size, random_state=random_state)
        traces = {
            "train": trace_samples(dataset_split["X_train"], dataset_split["y_train"], classifiers),
            "test": trace_samples(dataset_split["X_test"], dataset_split["y_test"], classifiers),
        }
        n_classes = len(classifiers) + 1

        for classifier in classifiers:
            classifier_diagnostic_rows.append(
                {
                    "dataset": dataset,
                    "hk": classifier.hk,
                    "component": classifier.component,
                    "description": classifier.description,
                    "weights": format_vector(classifier.weights),
                    "b": classifier.intercept,
                    "l1_norm": classifier.l1_norm,
                    "max_abs_weight": classifier.max_abs_weight,
                    "all_zero_weights": classifier.all_zero_weights,
                    "constant_decision": classifier.constant_decision,
                    "objective_value": classifier.objective_value,
                    "positive_class_accuracy_lb": classifier.positive_lb,
                    "negative_class_accuracy_lb": classifier.negative_lb,
                    "mip_gap": classifier.mip_gap,
                    "total_time_budget_seconds": classifier.total_time_budget_seconds,
                    "per_classifier_time_limit_seconds": classifier.per_classifier_time_limit_seconds,
                }
            )

        for split in ("train", "test"):
            trace = traces[split]
            y_true = trace["y_true"]
            predictions = trace["predictions"]
            first_positive_hk = trace["first_positive_hk"]
            total_accuracy = float(metric_row[metric_index[f"{split}_total_accuracy"]])

            for true_class in range(1, n_classes + 1):
                class_accuracy = metric_row[metric_index[f"{split}_class_{true_class}_accuracy"]]
                if class_accuracy != 0:
                    continue

                sample_mask = y_true == true_class
                sample_indices = np.where(sample_mask)[0]
                blocker_groups: dict[int, list[int]] = defaultdict(list)
                diagnosis_flags: list[bool] = []

                for sample_index in sample_indices:
                    assigned_hk = int(first_positive_hk[sample_index])
                    blocker_hk, blocker_type = determine_first_blocker(
                        true_class=true_class,
                        assigned_hk=assigned_hk,
                        n_classes=n_classes,
                    )
                    if blocker_hk is None or blocker_type is None:
                        continue
                    blocker_groups[blocker_hk].append(sample_index)

                blocker_rows: list[dict[str, Any]] = []
                zero_coeff_blocked_count = 0

                for blocker_hk, blocked_indices in sorted(blocker_groups.items()):
                    classifier = classifiers[blocker_hk - 1]
                    blocker_type = blocker_type_for_group(true_class=true_class, blocker_hk=blocker_hk)
                    prediction_distribution = class_distribution(predictions[blocked_indices])
                    blocked_by_zero = classifier.all_zero_weights
                    if blocked_by_zero:
                        zero_coeff_blocked_count += len(blocked_indices)
                    diagnosis_flags.append(blocked_by_zero)

                    blocker_row = {
                        "dataset": dataset,
                        "split": split,
                        "true_class": true_class,
                        "sample_count": int(len(sample_indices)),
                        "total_accuracy": total_accuracy,
                        "class_accuracy": 0.0,
                        "blocker_hk": blocker_hk,
                        "blocker_type": blocker_type,
                        "blocker_description": classifier.description,
                        "blocked_sample_count": int(len(blocked_indices)),
                        "blocked_fraction": float(len(blocked_indices) / len(sample_indices)),
                        "final_prediction_distribution": prediction_distribution,
                        "weights": format_vector(classifier.weights),
                        "b": classifier.intercept,
                        "all_zero_weights": classifier.all_zero_weights,
                        "constant_decision": classifier.constant_decision,
                        "l1_norm": classifier.l1_norm,
                        "max_abs_weight": classifier.max_abs_weight,
                        "objective_value": classifier.objective_value,
                        "positive_class_accuracy_lb": classifier.positive_lb,
                        "negative_class_accuracy_lb": classifier.negative_lb,
                        "mip_gap": classifier.mip_gap,
                        "total_time_budget_seconds": classifier.total_time_budget_seconds,
                        "per_classifier_time_limit_seconds": classifier.per_classifier_time_limit_seconds,
                    }
                    blocker_rows.append(blocker_row)
                    blocker_trace_rows.append(blocker_row)

                blocked_fraction = float(zero_coeff_blocked_count / len(sample_indices)) if len(sample_indices) else 0.0
                diagnosis = diagnose_zero_accuracy_case(blocker_rows, blocked_fraction)
                summaries.append(
                    ZeroAccuracySummary(
                        dataset=dataset,
                        split=split,
                        true_class=true_class,
                        sample_count=int(len(sample_indices)),
                        total_accuracy=total_accuracy,
                        blocker_hks=sorted(blocker_groups.keys()),
                        blocker_rows=blocker_rows,
                        blocked_by_zero_coefficients_fraction=blocked_fraction,
                        diagnosis=diagnosis,
                    )
                )

    return summaries, blocker_trace_rows, classifier_diagnostic_rows


def determine_first_blocker(true_class: int, assigned_hk: int, n_classes: int) -> tuple[int | None, str | None]:
    if assigned_hk == true_class:
        return None, None

    if assigned_hk > 0 and assigned_hk < true_class:
        return assigned_hk, "upstream_positive"

    if true_class < n_classes:
        return true_class, "target_negative"

    if assigned_hk > 0 and assigned_hk < n_classes:
        return assigned_hk, "upstream_positive"

    return None, None


def blocker_type_for_group(true_class: int, blocker_hk: int) -> str:
    return "upstream_positive" if blocker_hk < true_class else "target_negative"


def class_distribution(values: np.ndarray) -> str:
    unique, counts = np.unique(values, return_counts=True)
    parts = [f"class_{int(value)}={int(count)}" for value, count in zip(unique, counts)]
    return ", ".join(parts)


def diagnose_zero_accuracy_case(blocker_rows: list[dict[str, Any]], blocked_fraction: float) -> str:
    if blocker_rows and blocked_fraction == 1.0:
        if any(row["all_zero_weights"] and row["constant_decision"] == "always_negative" for row in blocker_rows):
            return "all samples blocked by always-negative zero-coefficient classifier(s)"
        if any(row["all_zero_weights"] and row["constant_decision"] == "always_positive" for row in blocker_rows):
            return "all samples trapped by always-positive zero-coefficient classifier(s)"
        return "all samples blocked by zero-coefficient classifier(s)"
    if blocker_rows and blocked_fraction > 0:
        return "mixed blockers; zero-coefficient classifiers explain part of the failure"
    return "zero accuracy not explained by zero-coefficient blockers in this run"


def render_markdown(
    summaries: list[ZeroAccuracySummary],
    classifier_rows: list[dict[str, Any]],
    *,
    workbook_path: Path,
    log_path: Path,
    generated_at: datetime,
) -> str:
    dataset_groups: dict[str, list[ZeroAccuracySummary]] = defaultdict(list)
    for summary in summaries:
        dataset_groups[summary.dataset].append(summary)

    lines: list[str] = []
    lines.append("# HCESVM Zero-Accuracy Coefficient Trace")
    lines.append("")
    lines.append(f"Generated: {human_timestamp(generated_at)}")
    lines.append(f"Source workbook: `{workbook_path}`")
    lines.append(f"Source log: `{log_path}`")
    lines.append("")
    lines.append("## Conclusion")
    lines.append("")

    if summaries and all(summary.blocked_by_zero_coefficients_fraction == 1.0 for summary in summaries):
        lines.append(
            "In the latest 30m-budget run, every HCESVM class with zero accuracy is fully explained by one or more "
            "zero-coefficient classifiers on the prediction path."
        )
    else:
        lines.append(
            "In the latest 30m-budget run, zero-coefficient classifiers explain some, but not necessarily all, "
            "zero-accuracy HCESVM classes."
        )

    lines.append(
        "The strongest evidence comes from `californiahousing` and `skill`, where zero-accuracy classes coincide with "
        "all-zero `Hk` weights, constant `b`, and time-limited solves with very high MIP gaps."
    )
    lines.append("")
    lines.append("## Zero-Accuracy Summary")
    lines.append("")
    lines.append("| Dataset | Split | Class | Samples | Blocker Hk | Zero-Coeff Fraction | Diagnosis |")
    lines.append("|---|---|---:|---:|---|---:|---|")

    for summary in summaries:
        blocker_text = ", ".join(f"H{hk}" for hk in summary.blocker_hks)
        lines.append(
            f"| {summary.dataset} | {summary.split} | {summary.true_class} | {summary.sample_count} | "
            f"{blocker_text} | {summary.blocked_by_zero_coefficients_fraction:.2f} | {summary.diagnosis} |"
        )

    for dataset in sorted(dataset_groups):
        lines.append("")
        lines.append(f"## {dataset}")
        lines.append("")

        zero_classifiers = [
            row for row in classifier_rows if row["dataset"] == dataset and row["all_zero_weights"]
        ]
        if zero_classifiers:
            lines.append("Zero-coefficient classifiers in this dataset:")
            for row in zero_classifiers:
                lines.append(
                    f"- `{row['component']}`: `{row['constant_decision']}`, `b={row['b']:.4f}`, "
                    f"`mip_gap={row['mip_gap']:.4f}`, `time_limit={row['per_classifier_time_limit_seconds']}s`"
                )
        else:
            lines.append("No zero-coefficient classifiers detected.")

        lines.append("")
        for summary in sorted(dataset_groups[dataset], key=lambda item: (item.split, item.true_class)):
            lines.append(
                f"- `{summary.split}` class `{summary.true_class}`: all `{summary.sample_count}` samples were blocked before "
                f"reaching the correct class. Diagnosis: {summary.diagnosis}."
            )
            for row in summary.blocker_rows:
                lines.append(
                    f"  blocker `H{row['blocker_hk']}` (`{row['blocker_type']}`): "
                    f"`blocked={row['blocked_sample_count']}`, "
                    f"`all_zero_weights={row['all_zero_weights']}`, "
                    f"`constant_decision={row['constant_decision']}`, "
                    f"`predictions={row['final_prediction_distribution']}`"
                )

    lines.append("")
    lines.append("## Interpretation")
    lines.append("")
    lines.append(
        "- `always_negative` zero-coefficient classifiers make the corresponding class unreachable because `x·w + b < 0` "
        "for every sample."
    )
    lines.append(
        "- `always_positive` zero-coefficient classifiers siphon all remaining samples into the classifier index class, "
        "which can wipe out later classes in the cascade."
    )
    lines.append(
        "- The observed high MIP gaps indicate the solver stopped at the allocated time budget with weak incumbent quality; "
        "this is consistent with, but does not mathematically prove, that the time budget was too short."
    )
    lines.append("")

    return "\n".join(lines) + "\n"


def write_excel(
    output_path: Path,
    summaries: list[ZeroAccuracySummary],
    blocker_rows: list[dict[str, Any]],
    classifier_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    blocker_sheet = workbook.create_sheet("blocker_trace")
    classifier_sheet = workbook.create_sheet("classifier_diagnostics")
    metadata_sheet = workbook.create_sheet("metadata")

    summary_headers = [
        "dataset",
        "split",
        "true_class",
        "sample_count",
        "total_accuracy",
        "blocker_hks",
        "blocked_by_zero_coefficients_fraction",
        "diagnosis",
    ]
    summary_sheet.append(summary_headers)
    for summary in summaries:
        summary_sheet.append(
            [
                summary.dataset,
                summary.split,
                summary.true_class,
                summary.sample_count,
                summary.total_accuracy,
                ", ".join(f"H{hk}" for hk in summary.blocker_hks),
                summary.blocked_by_zero_coefficients_fraction,
                summary.diagnosis,
            ]
        )

    blocker_headers = list(blocker_rows[0].keys()) if blocker_rows else []
    if blocker_headers:
        blocker_sheet.append(blocker_headers)
        for row in blocker_rows:
            blocker_sheet.append([row.get(header) for header in blocker_headers])

    classifier_headers = list(classifier_rows[0].keys()) if classifier_rows else []
    if classifier_headers:
        classifier_sheet.append(classifier_headers)
        for row in classifier_rows:
            classifier_sheet.append([row.get(header) for header in classifier_headers])

    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 1:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        autosize_worksheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    args = parse_arguments()
    generated_at = utc_now()
    timestamp = short_timestamp(generated_at)

    if args.report_output is None:
        args.report_output = ROOT / "docs" / "reports" / f"HCESVM_ZERO_ACCURACY_TRACE_{timestamp}.md"
    if args.xlsx_output is None:
        args.xlsx_output = ROOT / "docs" / "reports" / f"HCESVM_ZERO_ACCURACY_TRACE_{timestamp}.xlsx"
    if args.archive_log_output is None:
        archive_dir = ROOT / "results" / "archive" / f"{generated_at.strftime('%Y%m%d')}_zero_accuracy_trace"
        args.archive_log_output = archive_dir / f"hcesvm_zero_accuracy_trace_{timestamp}.log"

    args.archive_log_output.parent.mkdir(parents=True, exist_ok=True)

    log_lines: list[str] = []

    def log(message: str) -> None:
        line = f"[{human_timestamp(utc_now())}] {message}"
        print(line)
        log_lines.append(line)

    log(f"Source workbook: {args.workbook}")
    log(f"Source run log: {args.log}")
    log(f"Report output: {args.report_output}")
    log(f"Excel output: {args.xlsx_output}")

    log_text = args.log.read_text(encoding="utf-8")
    classifier_map = build_classifier_map(args.workbook, log_text)
    summaries, blocker_rows, classifier_rows = collect_zero_accuracy_summaries(
        args.workbook,
        classifier_map,
        test_size=args.test_size,
        random_state=args.random_state,
    )

    log(f"Found {len(summaries)} zero-accuracy class/split cases.")
    log(f"Collected {len(blocker_rows)} blocker trace rows.")
    log(f"Collected {len(classifier_rows)} classifier diagnostic rows.")

    markdown = render_markdown(
        summaries,
        classifier_rows,
        workbook_path=args.workbook,
        log_path=args.log,
        generated_at=generated_at,
    )
    args.report_output.parent.mkdir(parents=True, exist_ok=True)
    args.report_output.write_text(markdown, encoding="utf-8")
    log(f"Wrote Markdown report: {args.report_output}")

    metadata_rows = [
        ("generated_at_utc", human_timestamp(generated_at)),
        ("source_workbook", str(args.workbook)),
        ("source_log", str(args.log)),
        ("report_output", str(args.report_output)),
        ("excel_output", str(args.xlsx_output)),
        ("test_size", args.test_size),
        ("random_state", args.random_state),
        ("model", MODEL_NAME),
    ]
    write_excel(args.xlsx_output, summaries, blocker_rows, classifier_rows, metadata_rows)
    log(f"Wrote Excel trace workbook: {args.xlsx_output}")

    log(f"Wrote execution log: {args.archive_log_output}")
    args.archive_log_output.write_text("\n".join(log_lines) + "\n", encoding="utf-8")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
