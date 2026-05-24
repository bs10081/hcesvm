#!/usr/bin/env python3
"""Generate an SVOR/NPSVOR/HCESVM(test3) comparison report from existing logs."""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = ROOT / "results"
DOCS_DIR = ROOT / "docs" / "reports"
DEFAULT_SVOR_NPSVOR_LOG = RESULTS_DIR / "svor_npsvor_all_datasets_20260323_035717.log"
DEFAULT_REPORT_OUTPUT = DOCS_DIR / "SVOR_NPSVOR_HCESVM_TEST3_COMPARISON.md"
DEFAULT_CSV_OUTPUT = DOCS_DIR / "SVOR_NPSVOR_HCESVM_TEST3_COMPARISON.csv"
DEFAULT_XLSX_OUTPUT = DOCS_DIR / "SVOR_NPSVOR_HCESVM_TEST3_COMPARISON.xlsx"

DATASET_ORDER = [
    "Car_Evaluation",
    "Balance",
    "Contraceptive",
    "Hayes_Roth",
    "New_Thyroid",
    "TAE",
    "Thyroid",
    "Wine",
]
METHOD_ORDER = ["SVOR", "NPSVOR", "HCESVM(test3)"]

NUMBER_RE = re.compile(r"[-+]?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][-+]?\d+)?")


@dataclass(slots=True)
class SVORRecord:
    dataset: str
    train_total: float
    test_total: float
    train_class: list[float]
    test_class: list[float]
    weights: list[float]
    thresholds: list[float]
    source_file: Path
    source_timestamp: datetime | None


@dataclass(slots=True)
class NPSVORRecord:
    dataset: str
    train_total: float
    test_total: float
    train_class: list[float]
    test_class: list[float]
    hyperplanes: dict[int, tuple[list[float], float]]
    source_file: Path
    source_timestamp: datetime | None


@dataclass(slots=True)
class HCESVMCandidate:
    dataset: str
    source_file: Path
    source_timestamp: datetime | None
    train_total: float | None = None
    test_total: float | None = None
    train_class: list[float | None] = field(default_factory=lambda: [None, None, None])
    test_class: list[float | None] = field(default_factory=lambda: [None, None, None])
    h1_description: str | None = None
    h2_description: str | None = None
    h1_weights: list[float] | None = None
    h2_weights: list[float] | None = None
    h1_b: float | None = None
    h2_b: float | None = None
    format_name: str = "unknown"

    @property
    def complete(self) -> bool:
        return (
            self.train_total is not None
            and self.test_total is not None
            and all(value is not None for value in self.train_class)
            and all(value is not None for value in self.test_class)
            and self.h1_weights is not None
            and self.h2_weights is not None
            and self.h1_b is not None
            and self.h2_b is not None
        )


@dataclass(slots=True)
class HCESVMSelection:
    dataset: str
    selected: HCESVMCandidate
    skipped_newer_incomplete: list[HCESVMCandidate] = field(default_factory=list)


@dataclass(slots=True)
class DatasetSizes:
    dataset: str
    train_class: list[int]
    test_class: list[int]


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate the SVOR/NPSVOR/HCESVM(test3) comparison report."
    )
    parser.add_argument(
        "--svor-npsvor-log",
        type=Path,
        default=DEFAULT_SVOR_NPSVOR_LOG,
        help="Path to the combined SVOR/NPSVOR log.",
    )
    parser.add_argument(
        "--report-output",
        type=Path,
        default=DEFAULT_REPORT_OUTPUT,
        help="Markdown report output path.",
    )
    parser.add_argument(
        "--csv-output",
        type=Path,
        default=DEFAULT_CSV_OUTPUT,
        help="Summary CSV output path.",
    )
    parser.add_argument(
        "--xlsx-output",
        type=Path,
        default=DEFAULT_XLSX_OUTPUT,
        help="Summary Excel output path.",
    )
    parser.add_argument(
        "--archive-root",
        type=Path,
        default=RESULTS_DIR / "archive",
        help="Root directory for timestamped execution logs.",
    )
    return parser.parse_args()


def parse_timestamp(raw: str | None) -> datetime | None:
    if not raw:
        return None

    candidate = raw.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y%m%d_%H%M%S", "%Y%m%d %H%M%S"):
        try:
            return datetime.strptime(candidate, fmt)
        except ValueError:
            continue
    return None


def parse_timestamp_from_text(text: str) -> datetime | None:
    match = re.search(r"Timestamp:\s*([^\n]+)", text)
    if match:
        return parse_timestamp(match.group(1))
    return None


def parse_timestamp_from_path(path: Path) -> datetime | None:
    match = re.search(r"(\d{8}_\d{6})", path.name)
    if match:
        return parse_timestamp(match.group(1))
    return None


def candidate_timestamp(path: Path, text: str, raw_timestamp: str | None = None) -> datetime | None:
    return parse_timestamp(raw_timestamp) or parse_timestamp_from_text(text) or parse_timestamp_from_path(path)


def parse_float(value: str) -> float:
    return float(value.strip())


def parse_vector(text: str) -> list[float]:
    return [float(value) for value in NUMBER_RE.findall(text)]


def format_float(value: float | None) -> str:
    return "N/A" if value is None else f"{value:.4f}"


def format_vector(values: Iterable[float] | None, precision: int = 4) -> str:
    if values is None:
        return "N/A"
    return "[" + ", ".join(f"{value:.{precision}f}" for value in values) + "]"


def relpath(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def extract_between(text: str, start: str, end: str) -> str | None:
    pattern = re.escape(start) + r"(?P<body>.*?)" + re.escape(end)
    match = re.search(pattern, text, re.DOTALL)
    if not match:
        return None
    return match.group("body")


def dataset_sort_key(dataset: str) -> int:
    try:
        return DATASET_ORDER.index(dataset)
    except ValueError:
        return len(DATASET_ORDER)


def parse_svor_npsvor_records(
    log_path: Path,
) -> tuple[dict[str, SVORRecord], dict[str, NPSVORRecord], dict[str, DatasetSizes], datetime | None]:
    content = log_path.read_text(encoding="utf-8")
    start_time_match = re.search(r"Start time:\s*([^\n]+)", content)
    start_timestamp = parse_timestamp(start_time_match.group(1)) if start_time_match else None
    dataset_pattern = re.compile(
        r"Dataset: (?P<dataset>[^\n]+)\n=+\n(?P<body>.*?)(?=\n=+\nDataset: |\n=+\nAll Tests Completed!|\Z)",
        re.DOTALL,
    )
    accuracy_pattern = re.compile(
        r"Training accuracy:\s*([\d.]+)\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)\s*"
        r"\n\s*Testing accuracy:\s*([\d.]+)\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)",
        re.DOTALL,
    )

    svor_records: dict[str, SVORRecord] = {}
    npsvor_records: dict[str, NPSVORRecord] = {}
    dataset_sizes: dict[str, DatasetSizes] = {}

    for dataset_match in dataset_pattern.finditer(content):
        dataset = dataset_match.group("dataset").strip()
        body = dataset_match.group("body")
        sizes_match = re.search(
            r"Dataset Info:\s*"
            r"\n\s*Training:\s*Class1=(\d+),\s*Class2=(\d+),\s*Class3=(\d+)\s*"
            r"\n\s*Testing:\s*Class1=(\d+),\s*Class2=(\d+),\s*Class3=(\d+)",
            body,
        )
        if not sizes_match:
            raise ValueError(f"Missing dataset sizes for {dataset}")
        dataset_sizes[dataset] = DatasetSizes(
            dataset=dataset,
            train_class=[int(sizes_match.group(index)) for index in (1, 2, 3)],
            test_class=[int(sizes_match.group(index)) for index in (4, 5, 6)],
        )

        svor_section = extract_between(body, "SVOR Results:\n", "\n================================================================================\nTesting NPSVOR on")
        if not svor_section:
            raise ValueError(f"Could not parse SVOR section for {dataset}")

        svor_accuracy = accuracy_pattern.search(svor_section)
        if not svor_accuracy:
            raise ValueError(f"Missing SVOR accuracy block for {dataset}")

        svor_weights_match = re.search(r"Weights:\s*(\[[^\n]+\])", svor_section)
        svor_thresholds_match = re.search(r"Thresholds:\s*(\[[^\n]+\])", svor_section)
        if not svor_weights_match or not svor_thresholds_match:
            raise ValueError(f"Missing SVOR parameters for {dataset}")

        svor_records[dataset] = SVORRecord(
            dataset=dataset,
            train_total=parse_float(svor_accuracy.group(1)),
            test_total=parse_float(svor_accuracy.group(5)),
            train_class=[parse_float(svor_accuracy.group(index)) for index in (2, 3, 4)],
            test_class=[parse_float(svor_accuracy.group(index)) for index in (6, 7, 8)],
            weights=parse_vector(svor_weights_match.group(1)),
            thresholds=parse_vector(svor_thresholds_match.group(1)),
            source_file=log_path,
            source_timestamp=start_timestamp,
        )

        npsvor_section_match = re.search(r"NPSVOR Results:\n(?P<section>.*)", body, re.DOTALL)
        if not npsvor_section_match:
            raise ValueError(f"Could not parse NPSVOR section for {dataset}")
        npsvor_section = npsvor_section_match.group("section")

        npsvor_accuracy = accuracy_pattern.search(npsvor_section)
        if not npsvor_accuracy:
            raise ValueError(f"Missing NPSVOR accuracy block for {dataset}")

        hyperplanes: dict[int, tuple[list[float], float]] = {}
        for match in re.finditer(r"Class (\d+): w=(\[[^\]]+\]), b=([-\d.eE]+)", npsvor_section):
            hyperplanes[int(match.group(1))] = (
                parse_vector(match.group(2)),
                parse_float(match.group(3)),
            )
        if len(hyperplanes) != 3:
            raise ValueError(f"Missing NPSVOR hyperplanes for {dataset}")

        npsvor_records[dataset] = NPSVORRecord(
            dataset=dataset,
            train_total=parse_float(npsvor_accuracy.group(1)),
            test_total=parse_float(npsvor_accuracy.group(5)),
            train_class=[parse_float(npsvor_accuracy.group(index)) for index in (2, 3, 4)],
            test_class=[parse_float(npsvor_accuracy.group(index)) for index in (6, 7, 8)],
            hyperplanes=hyperplanes,
            source_file=log_path,
            source_timestamp=start_timestamp,
        )

    return svor_records, npsvor_records, dataset_sizes, start_timestamp


def parse_standard_hcesvm_test3(path: Path, text: str) -> list[HCESVMCandidate]:
    dataset_match = re.search(r"Dataset:\s*([A-Za-z0-9_]+)", text)
    if not dataset_match:
        return []

    train_match = re.search(
        r"Training Set Evaluation.*?Total Accuracy:\s*([\d.]+).*?Per-Class Accuracy:\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)\s*"
        r"\n\nClass Distribution:",
        text,
        re.DOTALL,
    )
    test_match = re.search(
        r"Test Set Evaluation.*?Total Accuracy:\s*([\d.]+).*?Per-Class Accuracy:\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)\s*"
        r"\n\nClass Distribution:",
        text,
        re.DOTALL,
    )

    h1_section = extract_between(text, "H1 Complete Weights and Bias", "H2 Complete Weights and Bias")
    h2_section = extract_between(text, "H2 Complete Weights and Bias", "Training Set Evaluation")
    if not h1_section or not h2_section:
        return []

    h1_b_match = re.search(r"Intercept \(b\):\s*([-\d.eE]+)", h1_section)
    h2_b_match = re.search(r"Intercept \(b\):\s*([-\d.eE]+)", h2_section)
    h1_description = re.search(r"Description:\s*(.+)", h1_section)
    h2_description = re.search(r"Description:\s*(.+)", h2_section)
    h1_weights = re.findall(r"w\[\d+\]\s*=\s*([-\d.eE]+)", h1_section)
    h2_weights = re.findall(r"w\[\d+\]\s*=\s*([-\d.eE]+)", h2_section)

    candidate = HCESVMCandidate(
        dataset=dataset_match.group(1),
        source_file=path,
        source_timestamp=candidate_timestamp(path, text),
        train_total=parse_float(train_match.group(1)) if train_match else None,
        test_total=parse_float(test_match.group(1)) if test_match else None,
        train_class=[parse_float(train_match.group(index)) for index in (2, 3, 4)] if train_match else [None, None, None],
        test_class=[parse_float(test_match.group(index)) for index in (2, 3, 4)] if test_match else [None, None, None],
        h1_description=h1_description.group(1).strip() if h1_description else None,
        h2_description=h2_description.group(1).strip() if h2_description else None,
        h1_weights=[float(value) for value in h1_weights] if h1_weights else None,
        h2_weights=[float(value) for value in h2_weights] if h2_weights else None,
        h1_b=parse_float(h1_b_match.group(1)) if h1_b_match else None,
        h2_b=parse_float(h2_b_match.group(1)) if h2_b_match else None,
        format_name="standard_single_dataset",
    )
    return [candidate]


def parse_summary_hcesvm_test3(path: Path, text: str) -> list[HCESVMCandidate]:
    dataset_match = re.search(r"Dataset:\s*([A-Za-z0-9_]+)", text)
    if not dataset_match:
        return []

    train_match = re.search(
        r"TRAINING SET EVALUATION.*?Total Accuracy:\s*([\d.]+).*?Per-Class Accuracy:\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)\s*"
        r"\n\nClass Distribution:",
        text,
        re.DOTALL,
    )
    test_match = re.search(
        r"TEST SET EVALUATION.*?Total Accuracy:\s*([\d.]+).*?Per-Class Accuracy:\s*"
        r"\n\s*Class 1:\s*([\d.]+)\s*"
        r"\n\s*Class 2:\s*([\d.]+)\s*"
        r"\n\s*Class 3:\s*([\d.]+)\s*"
        r"\n\nClass Distribution:",
        text,
        re.DOTALL,
    )
    h1_match = re.search(
        r"H1 Classifier:\s*\n\s*Description:\s*(.+?)\n\s*Weights:\s*(\[[^\]]+\])\n\s*Intercept:\s*([-\d.eE]+)",
        text,
        re.DOTALL,
    )
    h2_match = re.search(
        r"H2 Classifier:\s*\n\s*Description:\s*(.+?)\n\s*Weights:\s*(\[[^\]]+\])\n\s*Intercept:\s*([-\d.eE]+)",
        text,
        re.DOTALL,
    )

    candidate = HCESVMCandidate(
        dataset=dataset_match.group(1),
        source_file=path,
        source_timestamp=candidate_timestamp(path, text),
        train_total=parse_float(train_match.group(1)) if train_match else None,
        test_total=parse_float(test_match.group(1)) if test_match else None,
        train_class=[parse_float(train_match.group(index)) for index in (2, 3, 4)] if train_match else [None, None, None],
        test_class=[parse_float(test_match.group(index)) for index in (2, 3, 4)] if test_match else [None, None, None],
        h1_description=h1_match.group(1).strip() if h1_match else None,
        h2_description=h2_match.group(1).strip() if h2_match else None,
        h1_weights=parse_vector(h1_match.group(2)) if h1_match else None,
        h2_weights=parse_vector(h2_match.group(2)) if h2_match else None,
        h1_b=parse_float(h1_match.group(3)) if h1_match else None,
        h2_b=parse_float(h2_match.group(3)) if h2_match else None,
        format_name="summary_single_dataset",
    )
    return [candidate]


def parse_refactor_validation(path: Path, text: str) -> list[HCESVMCandidate]:
    file_timestamp = parse_timestamp_from_text(text)
    block_pattern = re.compile(
        r"Testing Dataset:\s*(?P<dataset>[A-Za-z0-9_]+)\n=+\n(?P<body>.*?)(?=\n=+\nTesting Dataset: |\n=+\nVALIDATION SUMMARY|\Z)",
        re.DOTALL,
    )
    candidates: list[HCESVMCandidate] = []
    for match in block_pattern.finditer(text):
        dataset = match.group("dataset")
        body = match.group("body")

        train_match = re.search(
            r"TRAINING RESULTS:\s*\n\s*Overall Accuracy:\s*([\d.]+)\s*"
            r"\n\s*Class 1 Accuracy:\s*([\d.]+)"
            r".*?\n\s*Class 2 Accuracy:\s*([\d.]+)"
            r".*?\n\s*Class 3 Accuracy:\s*([\d.]+)",
            body,
            re.DOTALL,
        )
        test_match = re.search(
            r"TESTING RESULTS:\s*\n\s*Overall Accuracy:\s*([\d.]+)\s*"
            r"\n\s*Class 1 Accuracy:\s*([\d.]+)"
            r".*?\n\s*Class 2 Accuracy:\s*([\d.]+)"
            r".*?\n\s*Class 3 Accuracy:\s*([\d.]+)",
            body,
            re.DOTALL,
        )
        h1_description = re.search(r"H1 Description:\s*(.+)", body)
        h2_description = re.search(r"H2 Description:\s*(.+)", body)
        h1_match = re.search(r"H1 WEIGHTS:\s*\n\s*w = (\[[^\]]+\])\s*\n\s*b = ([-\d.eE]+)", body, re.DOTALL)
        h2_match = re.search(r"H2 WEIGHTS:\s*\n\s*w = (\[[^\]]+\])\s*\n\s*b = ([-\d.eE]+)", body, re.DOTALL)

        candidates.append(
            HCESVMCandidate(
                dataset=dataset,
                source_file=path,
                source_timestamp=file_timestamp or parse_timestamp_from_path(path),
                train_total=parse_float(train_match.group(1)) if train_match else None,
                test_total=parse_float(test_match.group(1)) if test_match else None,
                train_class=[parse_float(train_match.group(index)) for index in (2, 3, 4)] if train_match else [None, None, None],
                test_class=[parse_float(test_match.group(index)) for index in (2, 3, 4)] if test_match else [None, None, None],
                h1_description=h1_description.group(1).strip() if h1_description else None,
                h2_description=h2_description.group(1).strip() if h2_description else None,
                h1_weights=parse_vector(h1_match.group(1)) if h1_match else None,
                h2_weights=parse_vector(h2_match.group(1)) if h2_match else None,
                h1_b=parse_float(h1_match.group(2)) if h1_match else None,
                h2_b=parse_float(h2_match.group(2)) if h2_match else None,
                format_name="refactor_validation",
            )
        )
    return candidates


def parse_revalidation(path: Path, text: str) -> list[HCESVMCandidate]:
    block_pattern = re.compile(
        r"Test3 Strategy Revalidation:\s*(?P<dataset>[A-Za-z0-9_]+)\nTimestamp:\s*(?P<timestamp>[^\n]+)\n=+\n(?P<body>.*?)(?=\n\nProcessing |\n=+\nSUMMARY:|\Z)",
        re.DOTALL,
    )
    candidates: list[HCESVMCandidate] = []
    for match in block_pattern.finditer(text):
        dataset = match.group("dataset")
        body = match.group("body")
        h1_match = re.search(
            r"H1 Solution:\s*\n\s*Weights \(w\):\s*(\[[^\]]+\])\s*\n\s*Intercept \(b\):\s*([-\d.eE]+)",
            body,
            re.DOTALL,
        )
        h2_match = re.search(
            r"H2 Solution:\s*\n\s*Weights \(w\):\s*(\[[^\]]+\])\s*\n\s*Intercept \(b\):\s*([-\d.eE]+)",
            body,
            re.DOTALL,
        )
        h1_description = re.search(r"Training H1:\s*(.+)", body)
        h2_description = re.search(r"Training H2:\s*(.+)", body)

        candidates.append(
            HCESVMCandidate(
                dataset=dataset,
                source_file=path,
                source_timestamp=candidate_timestamp(path, text, match.group("timestamp")),
                h1_description=h1_description.group(1).strip() if h1_description else None,
                h2_description=h2_description.group(1).strip() if h2_description else None,
                h1_weights=parse_vector(h1_match.group(1)) if h1_match else None,
                h2_weights=parse_vector(h2_match.group(1)) if h2_match else None,
                h1_b=parse_float(h1_match.group(2)) if h1_match else None,
                h2_b=parse_float(h2_match.group(2)) if h2_match else None,
                format_name="revalidation_incomplete",
            )
        )
    return candidates


def load_hcesvm_candidates() -> dict[str, list[HCESVMCandidate]]:
    candidates: dict[str, list[HCESVMCandidate]] = {}
    for path in sorted(ROOT.rglob("*.log")):
        if "test3" not in path.name.lower():
            continue
        text = path.read_text(encoding="utf-8")
        parsed: list[HCESVMCandidate]
        if "TEST3 STRATEGY REFACTOR VALIDATION" in text:
            parsed = parse_refactor_validation(path, text)
        elif "Test3 Strategy Revalidation" in text:
            parsed = parse_revalidation(path, text)
        elif "MODEL SUMMARY" in text and "TRAINING SET EVALUATION" in text:
            parsed = parse_summary_hcesvm_test3(path, text)
        elif "Training Set Evaluation" in text and "H1 Complete Weights and Bias" in text:
            parsed = parse_standard_hcesvm_test3(path, text)
        else:
            parsed = []

        for candidate in parsed:
            if candidate.dataset not in DATASET_ORDER:
                continue
            candidates.setdefault(candidate.dataset, []).append(candidate)
    return candidates


def select_latest_complete_hcesvm(candidates_by_dataset: dict[str, list[HCESVMCandidate]]) -> dict[str, HCESVMSelection]:
    selections: dict[str, HCESVMSelection] = {}

    def sort_key(candidate: HCESVMCandidate) -> tuple[datetime, str]:
        return (
            candidate.source_timestamp or datetime.min,
            candidate.source_file.name,
        )

    for dataset in DATASET_ORDER:
        candidates = candidates_by_dataset.get(dataset, [])
        if not candidates:
            raise ValueError(f"No HCESVM(test3) candidates found for {dataset}")

        ordered = sorted(candidates, key=sort_key, reverse=True)
        selected: HCESVMCandidate | None = None
        skipped_newer_incomplete: list[HCESVMCandidate] = []
        for candidate in ordered:
            if candidate.complete:
                selected = candidate
                break
            skipped_newer_incomplete.append(candidate)

        if selected is None:
            raise ValueError(f"No complete HCESVM(test3) result found for {dataset}")

        selections[dataset] = HCESVMSelection(
            dataset=dataset,
            selected=selected,
            skipped_newer_incomplete=skipped_newer_incomplete,
        )
    return selections


def build_summary_rows(
    svor_records: dict[str, SVORRecord],
    npsvor_records: dict[str, NPSVORRecord],
    hcesvm_selections: dict[str, HCESVMSelection],
    dataset_sizes: dict[str, DatasetSizes],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for dataset in DATASET_ORDER:
        sizes = dataset_sizes[dataset]
        svor_record = svor_records[dataset]
        rows.append(
            {
                "Dataset": dataset,
                "Method": "SVOR",
                "Train Acc": format_float(svor_record.train_total),
                "Test Acc": format_float(svor_record.test_total),
                "Train C1": format_float(svor_record.train_class[0]),
                "Train C2": format_float(svor_record.train_class[1]),
                "Train C3": format_float(svor_record.train_class[2]),
                "Test C1": format_float(svor_record.test_class[0]),
                "Test C2": format_float(svor_record.test_class[1]),
                "Test C3": format_float(svor_record.test_class[2]),
                "Train Size C1": str(sizes.train_class[0]),
                "Train Size C2": str(sizes.train_class[1]),
                "Train Size C3": str(sizes.train_class[2]),
                "Test Size C1": str(sizes.test_class[0]),
                "Test Size C2": str(sizes.test_class[1]),
                "Test Size C3": str(sizes.test_class[2]),
                "Source Timestamp": svor_record.source_timestamp.isoformat(sep=" ") if svor_record.source_timestamp else "",
                "Source File": relpath(svor_record.source_file),
            }
        )

        npsvor_record = npsvor_records[dataset]
        rows.append(
            {
                "Dataset": dataset,
                "Method": "NPSVOR",
                "Train Acc": format_float(npsvor_record.train_total),
                "Test Acc": format_float(npsvor_record.test_total),
                "Train C1": format_float(npsvor_record.train_class[0]),
                "Train C2": format_float(npsvor_record.train_class[1]),
                "Train C3": format_float(npsvor_record.train_class[2]),
                "Test C1": format_float(npsvor_record.test_class[0]),
                "Test C2": format_float(npsvor_record.test_class[1]),
                "Test C3": format_float(npsvor_record.test_class[2]),
                "Train Size C1": str(sizes.train_class[0]),
                "Train Size C2": str(sizes.train_class[1]),
                "Train Size C3": str(sizes.train_class[2]),
                "Test Size C1": str(sizes.test_class[0]),
                "Test Size C2": str(sizes.test_class[1]),
                "Test Size C3": str(sizes.test_class[2]),
                "Source Timestamp": npsvor_record.source_timestamp.isoformat(sep=" ") if npsvor_record.source_timestamp else "",
                "Source File": relpath(npsvor_record.source_file),
            }
        )

        hcesvm_record = hcesvm_selections[dataset].selected
        rows.append(
            {
                "Dataset": dataset,
                "Method": "HCESVM(test3)",
                "Train Acc": format_float(hcesvm_record.train_total),
                "Test Acc": format_float(hcesvm_record.test_total),
                "Train C1": format_float(hcesvm_record.train_class[0]),
                "Train C2": format_float(hcesvm_record.train_class[1]),
                "Train C3": format_float(hcesvm_record.train_class[2]),
                "Test C1": format_float(hcesvm_record.test_class[0]),
                "Test C2": format_float(hcesvm_record.test_class[1]),
                "Test C3": format_float(hcesvm_record.test_class[2]),
                "Train Size C1": str(sizes.train_class[0]),
                "Train Size C2": str(sizes.train_class[1]),
                "Train Size C3": str(sizes.train_class[2]),
                "Test Size C1": str(sizes.test_class[0]),
                "Test Size C2": str(sizes.test_class[1]),
                "Test Size C3": str(sizes.test_class[2]),
                "Source Timestamp": hcesvm_record.source_timestamp.isoformat(sep=" ") if hcesvm_record.source_timestamp else "",
                "Source File": relpath(hcesvm_record.source_file),
            }
        )
    return rows


def render_markdown(
    generation_time: datetime,
    summary_rows: list[dict[str, str]],
    svor_records: dict[str, SVORRecord],
    npsvor_records: dict[str, NPSVORRecord],
    hcesvm_selections: dict[str, HCESVMSelection],
) -> str:
    lines: list[str] = []
    lines.append("# SVOR / NPSVOR / HCESVM(test3) Comparison Report")
    lines.append("")
    lines.append(f"**Generated**: {generation_time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("**Datasets**: 8")
    lines.append("**HCESVM selection rule**: latest complete `test3` result per dataset")
    lines.append("")
    lines.append("## Summary Table")
    lines.append("")
    lines.append(
        "| Dataset | Method | Train Acc | Test Acc | Train C1 | Train C2 | Train C3 | Test C1 | Test C2 | Test C3 | Train Size C1 | Train Size C2 | Train Size C3 | Test Size C1 | Test Size C2 | Test Size C3 |"
    )
    lines.append(
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |"
    )
    for row in summary_rows:
        lines.append(
            f"| {row['Dataset']} | {row['Method']} | {row['Train Acc']} | {row['Test Acc']} | "
            f"{row['Train C1']} | {row['Train C2']} | {row['Train C3']} | "
            f"{row['Test C1']} | {row['Test C2']} | {row['Test C3']} | "
            f"{row['Train Size C1']} | {row['Train Size C2']} | {row['Train Size C3']} | "
            f"{row['Test Size C1']} | {row['Test Size C2']} | {row['Test Size C3']} |"
        )

    lines.append("")
    lines.append("## HCESVM Source Selection")
    lines.append("")
    lines.append("| Dataset | Selected Source | Source Timestamp | Selection Note |")
    lines.append("| --- | --- | --- | --- |")
    for dataset in DATASET_ORDER:
        selection = hcesvm_selections[dataset]
        note = "Latest complete test3 result"
        if selection.skipped_newer_incomplete:
            skipped = ", ".join(
                f"`{candidate.source_file.name}`"
                for candidate in selection.skipped_newer_incomplete
            )
            note = f"Skipped newer incomplete logs: {skipped}"
        lines.append(
            f"| {dataset} | `{relpath(selection.selected.source_file)}` | "
            f"{selection.selected.source_timestamp.strftime('%Y-%m-%d %H:%M:%S') if selection.selected.source_timestamp else 'N/A'} | "
            f"{note} |"
        )

    lines.append("")
    lines.append("## Parameter Details")
    lines.append("")
    for dataset in DATASET_ORDER:
        svor_record = svor_records[dataset]
        npsvor_record = npsvor_records[dataset]
        hcesvm_record = hcesvm_selections[dataset].selected

        npsvor_params = "<br>".join(
            f"Class {class_id}: `w={format_vector(weights)}`, `b={bias:.4f}`"
            for class_id, (weights, bias) in sorted(npsvor_record.hyperplanes.items())
        )
        hcesvm_source_note = (
            f"`{relpath(hcesvm_record.source_file)}`"
            f" ({hcesvm_record.source_timestamp.strftime('%Y-%m-%d %H:%M:%S') if hcesvm_record.source_timestamp else 'N/A'})"
        )
        lines.append(f"### {dataset}")
        lines.append("")
        lines.append("| Method | Source | Parameters |")
        lines.append("| --- | --- | --- |")
        lines.append(
            "| SVOR | `{source}` | `w={weights}`<br>`thresholds={thresholds}` |".format(
                source=relpath(svor_record.source_file),
                weights=format_vector(svor_record.weights),
                thresholds=format_vector(svor_record.thresholds),
            )
        )
        lines.append(
            "| NPSVOR | `{source}` | {params} |".format(
                source=relpath(npsvor_record.source_file),
                params=npsvor_params,
            )
        )
        lines.append(
            "| HCESVM(test3) | {source} | "
            "`H1 {h1_desc}`: `w={h1_weights}`, `b={h1_b:.4f}`<br>"
            "`H2 {h2_desc}`: `w={h2_weights}`, `b={h2_b:.4f}` |".format(
                source=hcesvm_source_note,
                h1_desc=hcesvm_record.h1_description or "",
                h1_weights=format_vector(hcesvm_record.h1_weights),
                h1_b=hcesvm_record.h1_b if hcesvm_record.h1_b is not None else 0.0,
                h2_desc=hcesvm_record.h2_description or "",
                h2_weights=format_vector(hcesvm_record.h2_weights),
                h2_b=hcesvm_record.h2_b if hcesvm_record.h2_b is not None else 0.0,
            )
        )
        lines.append("")

    return "\n".join(lines) + "\n"


def write_summary_csv(output_path: Path, summary_rows: list[dict[str, str]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "Dataset",
        "Method",
        "Train Acc",
        "Test Acc",
        "Train C1",
        "Train C2",
        "Train C3",
        "Test C1",
        "Test C2",
        "Test C3",
        "Train Size C1",
        "Train Size C2",
        "Train Size C3",
        "Test Size C1",
        "Test Size C2",
        "Test Size C3",
        "Source Timestamp",
        "Source File",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(summary_rows)


def write_summary_xlsx(output_path: Path, summary_rows: list[dict[str, str]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SVOR_NPSVOR_HCESVM_TEST3_COMPAR"

    headers = [
        "Dataset",
        "Method",
        "Train Acc",
        "Test Acc",
        "Train C1",
        "Train C2",
        "Train C3",
        "Test C1",
        "Test C2",
        "Test C3",
        "Train Size C1",
        "Train Size C2",
        "Train Size C3",
        "Test Size C1",
        "Test Size C2",
        "Test Size C3",
        "Source Timestamp",
        "Source File",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    def excel_value(header: str, value: str) -> object:
        if value == "":
            return None
        if header in {"Train Acc", "Test Acc", "Train C1", "Train C2", "Train C3", "Test C1", "Test C2", "Test C3"}:
            return float(value)
        if header in {"Train Size C1", "Train Size C2", "Train Size C3", "Test Size C1", "Test Size C2", "Test Size C3"}:
            return int(value)
        if header == "Source Timestamp":
            return parse_timestamp(value)
        return value

    last_dataset = None
    for row in summary_rows:
        if last_dataset is not None and row["Dataset"] != last_dataset:
            worksheet.append([None] * len(headers))
        worksheet.append([excel_value(header, row[header]) for header in headers])
        last_dataset = row["Dataset"]

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    width_map = {
        "A": 18,
        "B": 16,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 21,
        "R": 72,
    }
    for column_letter, width in width_map.items():
        worksheet.column_dimensions[column_letter].width = width

    numeric_columns = {"C", "D", "E", "F", "G", "H", "I", "J"}
    integer_columns = {"K", "L", "M", "N", "O", "P"}
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if all(cell.value is None for cell in row):
            continue
        for cell in row:
            if cell.column_letter in numeric_columns and cell.value not in (None, ""):
                cell.number_format = "0.0000"
            elif cell.column_letter in integer_columns and cell.value not in (None, ""):
                cell.number_format = "0"
    workbook.save(output_path)


def render_execution_log(
    start_time: datetime,
    end_time: datetime,
    summary_rows: list[dict[str, str]],
    svor_records: dict[str, SVORRecord],
    npsvor_records: dict[str, NPSVORRecord],
    hcesvm_selections: dict[str, HCESVMSelection],
    report_output: Path,
    csv_output: Path,
    xlsx_output: Path,
) -> str:
    lines: list[str] = []
    lines.append("=" * 80)
    lines.append("Three-Model Comparison Report Generation")
    lines.append("=" * 80)
    lines.append(f"Start time: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"End time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Duration: {end_time - start_time}")
    lines.append("")
    lines.append(f"[{end_time.strftime('%Y-%m-%d %H:%M:%S')}] Wrote Markdown report: {relpath(report_output)}")
    lines.append(f"[{end_time.strftime('%Y-%m-%d %H:%M:%S')}] Wrote CSV summary: {relpath(csv_output)}")
    lines.append(f"[{end_time.strftime('%Y-%m-%d %H:%M:%S')}] Wrote Excel summary: {relpath(xlsx_output)}")
    lines.append("")
    lines.append("HCESVM source selection")
    lines.append("-" * 80)
    for dataset in DATASET_ORDER:
        selection = hcesvm_selections[dataset]
        selected = selection.selected
        lines.append(
            f"[{end_time.strftime('%Y-%m-%d %H:%M:%S')}] {dataset}: selected "
            f"{relpath(selected.source_file)} ({selected.source_timestamp.strftime('%Y-%m-%d %H:%M:%S') if selected.source_timestamp else 'N/A'})"
        )
        for skipped in selection.skipped_newer_incomplete:
            lines.append(
                f"[{end_time.strftime('%Y-%m-%d %H:%M:%S')}] {dataset}: skipped newer incomplete "
                f"{relpath(skipped.source_file)} ({skipped.source_timestamp.strftime('%Y-%m-%d %H:%M:%S') if skipped.source_timestamp else 'N/A'})"
            )

    lines.append("")
    lines.append("Summary table")
    lines.append("-" * 80)
    header = (
        f"{'Dataset':<18} {'Method':<14} {'Train':>8} {'Test':>8} "
        f"{'Tr C1':>8} {'Tr C2':>8} {'Tr C3':>8} {'Te C1':>8} {'Te C2':>8} {'Te C3':>8} "
        f"{'TrS1':>6} {'TrS2':>6} {'TrS3':>6} {'TeS1':>6} {'TeS2':>6} {'TeS3':>6}"
    )
    lines.append(header)
    lines.append("-" * len(header))
    for row in summary_rows:
        lines.append(
            f"{row['Dataset']:<18} {row['Method']:<14} {row['Train Acc']:>8} {row['Test Acc']:>8} "
            f"{row['Train C1']:>8} {row['Train C2']:>8} {row['Train C3']:>8} "
            f"{row['Test C1']:>8} {row['Test C2']:>8} {row['Test C3']:>8} "
            f"{row['Train Size C1']:>6} {row['Train Size C2']:>6} {row['Train Size C3']:>6} "
            f"{row['Test Size C1']:>6} {row['Test Size C2']:>6} {row['Test Size C3']:>6}"
        )

    lines.append("")
    lines.append("Detailed parameters")
    lines.append("-" * 80)
    for dataset in DATASET_ORDER:
        svor_record = svor_records[dataset]
        npsvor_record = npsvor_records[dataset]
        hcesvm_record = hcesvm_selections[dataset].selected

        lines.append(f"Dataset: {dataset}")
        lines.append(
            f"  SVOR train/test: {svor_record.train_total:.4f} / {svor_record.test_total:.4f}"
        )
        lines.append(
            f"  SVOR class train: {format_vector(svor_record.train_class)}"
        )
        lines.append(
            f"  SVOR class test: {format_vector(svor_record.test_class)}"
        )
        lines.append(f"  SVOR w: {format_vector(svor_record.weights, precision=6)}")
        lines.append(f"  SVOR thresholds: {format_vector(svor_record.thresholds, precision=6)}")

        lines.append(
            f"  NPSVOR train/test: {npsvor_record.train_total:.4f} / {npsvor_record.test_total:.4f}"
        )
        lines.append(
            f"  NPSVOR class train: {format_vector(npsvor_record.train_class)}"
        )
        lines.append(
            f"  NPSVOR class test: {format_vector(npsvor_record.test_class)}"
        )
        for class_id, (weights, bias) in sorted(npsvor_record.hyperplanes.items()):
            lines.append(
                f"  NPSVOR class {class_id}: w={format_vector(weights, precision=6)}, b={bias:.6f}"
            )

        lines.append(
            f"  HCESVM(test3) source: {relpath(hcesvm_record.source_file)}"
        )
        lines.append(
            f"  HCESVM(test3) train/test: {hcesvm_record.train_total:.4f} / {hcesvm_record.test_total:.4f}"
        )
        lines.append(
            f"  HCESVM(test3) class train: {format_vector(hcesvm_record.train_class)}"
        )
        lines.append(
            f"  HCESVM(test3) class test: {format_vector(hcesvm_record.test_class)}"
        )
        lines.append(
            f"  HCESVM(test3) H1: w={format_vector(hcesvm_record.h1_weights, precision=6)}, b={hcesvm_record.h1_b:.6f}"
        )
        lines.append(
            f"  HCESVM(test3) H2: w={format_vector(hcesvm_record.h2_weights, precision=6)}, b={hcesvm_record.h2_b:.6f}"
        )
        lines.append("")

    return "\n".join(lines) + "\n"


def write_execution_log(
    archive_root: Path,
    generation_time: datetime,
    content: str,
) -> Path:
    archive_dir = archive_root / f"{generation_time.strftime('%Y%m%d')}_three_model_comparison"
    archive_dir.mkdir(parents=True, exist_ok=True)
    log_path = archive_dir / f"three_model_comparison_all_{generation_time.strftime('%Y%m%d_%H%M%S')}.log"
    log_path.write_text(content, encoding="utf-8")
    return log_path


def main() -> int:
    args = parse_arguments()
    start_time = datetime.now()

    svor_records, npsvor_records, dataset_sizes, _ = parse_svor_npsvor_records(args.svor_npsvor_log)
    hcesvm_candidates = load_hcesvm_candidates()
    hcesvm_selections = select_latest_complete_hcesvm(hcesvm_candidates)

    summary_rows = build_summary_rows(svor_records, npsvor_records, hcesvm_selections, dataset_sizes)

    report_output = args.report_output
    report_output.parent.mkdir(parents=True, exist_ok=True)
    report_output.write_text(
        render_markdown(start_time, summary_rows, svor_records, npsvor_records, hcesvm_selections),
        encoding="utf-8",
    )
    write_summary_csv(args.csv_output, summary_rows)
    write_summary_xlsx(args.xlsx_output, summary_rows)

    end_time = datetime.now()
    execution_log = render_execution_log(
        start_time=start_time,
        end_time=end_time,
        summary_rows=summary_rows,
        svor_records=svor_records,
        npsvor_records=npsvor_records,
        hcesvm_selections=hcesvm_selections,
        report_output=report_output,
        csv_output=args.csv_output,
        xlsx_output=args.xlsx_output,
    )
    log_path = write_execution_log(args.archive_root, end_time, execution_log)

    print(f"Markdown report: {report_output}")
    print(f"CSV summary: {args.csv_output}")
    print(f"Excel summary: {args.xlsx_output}")
    print(f"Execution log: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
