"""Run SVOR then NPSVOR on the exact archived skill 1000-sample split."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import traceback
from contextlib import redirect_stderr, redirect_stdout
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
from gurobipy import GRB
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models.npsvor import NPSVORQP
from .models.svor import SVORImplicitQP
from .utils.evaluator import evaluate_multiclass
from .utils.ordinal_data import OrdinalDatasetSplit, load_tabular_dataset_split


ROOT = Path(__file__).resolve().parents[2]
DOCS_DIR = ROOT / "docs" / "reports"
RESULTS_ARCHIVE_DIR = ROOT / "results" / "archive"
DATASET_NAME = "skill_method3_4class_1000"
TARGET_COLUMN = "response"
DEFAULT_SVOR_PARAMS = {
    "C": 1.0,
    "add_order_constraints": True,
    "solver_params": {
        "TimeLimit": 1800,
        "MIPGap": 1e-4,
        "OutputFlag": 0,
    },
}
DEFAULT_NPSVOR_PARAMS = {
    "C1": 1.0,
    "C2": 1.0,
    "epsilon": 0.2,
    "prediction_rule": "min_distance",
    "solver_params": {
        "TimeLimit": 1800,
        "MIPGap": 1e-4,
        "OutputFlag": 0,
    },
}
NEAR_ZERO_ACCURACY = 0.05
SUBSTANTIAL_OUTPERFORM_DELTA = 0.10


class TeeStream:
    """Mirror output to stdout/stderr and the timestamped log."""

    def __init__(self, *streams: Any) -> None:
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()


@dataclass(frozen=True, slots=True)
class ManifestSplit:
    dataset: str
    source_dataset: str
    source_url: str
    description: str
    class_map: dict[int, int]
    train_original_counts: list[int]
    test_original_counts: list[int]
    train_sampled_counts: list[int]
    test_sampled_counts: list[int]
    train_target_size: int
    test_target_size: int
    feature_names: list[str]
    manifest_path: Path
    train_csv: Path
    test_csv: Path

    @property
    def n_classes(self) -> int:
        return len(self.train_sampled_counts)

    @property
    def relabel_summary(self) -> str:
        return ", ".join(f"{source}->{target}" for source, target in sorted(self.class_map.items()))


@dataclass(slots=True)
class DatasetBundle:
    split: ManifestSplit
    dataset: OrdinalDatasetSplit
    X_train: np.ndarray
    y_train: np.ndarray
    X_test: np.ndarray
    y_test: np.ndarray

    @property
    def n_classes(self) -> int:
        return len(sorted(set(self.y_train.tolist())))

    @property
    def n_features(self) -> int:
        return int(self.X_train.shape[1])


@dataclass(slots=True)
class ModelRunResult:
    dataset: str
    model: str
    status: str
    fit_seconds: float | None
    train_metrics: dict[str, Any] | None
    test_metrics: dict[str, Any] | None
    parameter_rows: list[dict[str, Any]]
    error: str | None = None


@dataclass(frozen=True, slots=True)
class ArtifactPaths:
    log_path: Path
    workbook_path: Path
    markdown_path: Path


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run SVOR then NPSVOR on the exact archived skill_method3_4class_1000 split."
    )
    parser.add_argument("--manifest-json", required=True, help="Path to the archived manifest JSON.")
    parser.add_argument(
        "--svor-time-limit",
        type=int,
        default=DEFAULT_SVOR_PARAMS["solver_params"]["TimeLimit"],
        help="SVOR total model time limit in seconds.",
    )
    parser.add_argument(
        "--npsvor-time-limit",
        type=int,
        default=DEFAULT_NPSVOR_PARAMS["solver_params"]["TimeLimit"],
        help="NPSVOR total model time limit in seconds before per-rank splitting.",
    )
    return parser.parse_args(argv)


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def short_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y%m%d_%H%M%S")


def human_timestamp(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S UTC")


def git_output(*args: str) -> str:
    return subprocess.check_output(args, cwd=ROOT, text=True).strip()


def format_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (np.floating, float)):
        return float(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    return value


def format_vector(values: Any) -> str:
    if values is None:
        return ""
    array = np.asarray(values, dtype=float).tolist()
    return json.dumps([round(float(value), 10) for value in array], ensure_ascii=False)


def format_optional_float(value: Any, *, decimals: int = 4) -> str:
    if value is None:
        return ""
    return f"{float(value):.{decimals}f}"


def format_report_table_row(columns: list[str]) -> str:
    return "| " + " | ".join(columns) + " |"


def banner(tee: TeeStream, title: str, *, major: bool = True) -> None:
    line = "=" * 100 if major else "-" * 100
    print(f"\n{line}", file=tee)
    print(f"[{human_timestamp(utc_now())}] {title}", file=tee)
    print(line, file=tee)


def model_status_label(status_code: int | None, *, has_solution: bool) -> str:
    if status_code == GRB.OPTIMAL:
        return "optimal"
    if status_code == GRB.SUBOPTIMAL and has_solution:
        return "suboptimal_with_solution"
    if status_code == GRB.TIME_LIMIT and has_solution:
        return "time_limit_with_solution"
    return "failed_or_no_solution"


def derive_csv_path(manifest_path: Path, manifest: dict[str, Any], key: str, suffix: str) -> Path:
    raw_value = manifest.get(key)
    if raw_value:
        return Path(str(raw_value)).expanduser().resolve()
    return manifest_path.with_name(f"{manifest_path.stem.replace('_manifest_', f'_{suffix}_')}.csv").resolve()


def load_manifest_split(manifest_path: str | Path) -> ManifestSplit:
    resolved_path = Path(manifest_path).expanduser().resolve()
    manifest = json.loads(resolved_path.read_text(encoding="utf-8"))
    dataset_name = str(manifest["dataset"])
    if dataset_name != DATASET_NAME:
        raise ValueError(f"Expected dataset {DATASET_NAME!r}, got {dataset_name!r}.")

    class_map_raw = manifest.get("class_map")
    if not isinstance(class_map_raw, dict) or not class_map_raw:
        raise ValueError("Manifest must contain a non-empty class_map.")

    class_map = {int(source): int(target) for source, target in class_map_raw.items()}
    train_csv = derive_csv_path(resolved_path, manifest, "train_csv", "train")
    test_csv = derive_csv_path(resolved_path, manifest, "test_csv", "test")

    if not train_csv.exists():
        raise FileNotFoundError(f"Train CSV not found: {train_csv}")
    if not test_csv.exists():
        raise FileNotFoundError(f"Test CSV not found: {test_csv}")

    return ManifestSplit(
        dataset=dataset_name,
        source_dataset=str(manifest["source_dataset"]),
        source_url=str(manifest["source_url"]),
        description=str(manifest["description"]),
        class_map=class_map,
        train_original_counts=[int(value) for value in manifest["train_original_counts"]],
        test_original_counts=[int(value) for value in manifest["test_original_counts"]],
        train_sampled_counts=[int(value) for value in manifest["train_sampled_counts"]],
        test_sampled_counts=[int(value) for value in manifest["test_sampled_counts"]],
        train_target_size=int(manifest["train_target_size"]),
        test_target_size=int(manifest["test_target_size"]),
        feature_names=[str(value) for value in manifest["feature_names"]],
        manifest_path=resolved_path,
        train_csv=train_csv,
        test_csv=test_csv,
    )


def load_bundle(split: ManifestSplit) -> DatasetBundle:
    dataset = load_tabular_dataset_split(
        split.train_csv,
        test_path=split.test_csv,
        target_column=TARGET_COLUMN,
    )
    return DatasetBundle(
        split=split,
        dataset=dataset,
        X_train=np.asarray(dataset.X_train, dtype=float),
        y_train=np.asarray(dataset.y_train, dtype=int),
        X_test=np.asarray(dataset.X_test, dtype=float),
        y_test=np.asarray(dataset.y_test, dtype=int),
    )


def build_metric_headers(n_classes: int) -> list[str]:
    headers = [
        "dataset",
        "source_dataset",
        "model",
        "status",
        "n_classes",
        "n_features",
        "fit_seconds",
        "source_manifest_json",
        "train_csv",
        "test_csv",
        "source_url",
        "relabel_map",
        "train_total_accuracy",
        "test_total_accuracy",
    ]
    for class_index in range(1, n_classes + 1):
        headers.extend(
            [
                f"train_class_{class_index}_count",
                f"train_class_{class_index}_accuracy",
                f"test_class_{class_index}_count",
                f"test_class_{class_index}_accuracy",
            ]
        )
    headers.append("error")
    return headers


def build_parameter_headers() -> list[str]:
    return [
        "dataset",
        "source_dataset",
        "model",
        "component",
        "description",
        "status",
        "fit_seconds",
        "solver_status_code",
        "solver_status_label",
        "weights",
        "b",
        "objective_value",
        "prediction_rule",
    ]


def build_metadata_rows(metadata: dict[str, Any]) -> list[tuple[str, Any]]:
    return [(key, value) for key, value in metadata.items()]


def collect_metric_row(bundle: DatasetBundle, result: ModelRunResult) -> dict[str, Any]:
    row: dict[str, Any] = {
        "dataset": bundle.split.dataset,
        "source_dataset": bundle.split.source_dataset,
        "model": result.model,
        "status": result.status,
        "n_classes": bundle.n_classes,
        "n_features": bundle.n_features,
        "fit_seconds": result.fit_seconds,
        "source_manifest_json": str(bundle.split.manifest_path),
        "train_csv": str(bundle.split.train_csv),
        "test_csv": str(bundle.split.test_csv),
        "source_url": bundle.split.source_url,
        "relabel_map": bundle.split.relabel_summary,
        "train_total_accuracy": None,
        "test_total_accuracy": None,
        "error": result.error,
    }

    if result.train_metrics is not None:
        row["train_total_accuracy"] = format_scalar(result.train_metrics["total_accuracy"])
    if result.test_metrics is not None:
        row["test_total_accuracy"] = format_scalar(result.test_metrics["total_accuracy"])

    for class_index in range(1, bundle.n_classes + 1):
        row[f"train_class_{class_index}_count"] = (
            int(result.train_metrics[f"class_{class_index}_count"])
            if result.train_metrics
            else int(bundle.split.train_sampled_counts[class_index - 1])
        )
        row[f"train_class_{class_index}_accuracy"] = (
            format_scalar(result.train_metrics[f"class_{class_index}_accuracy"]) if result.train_metrics else None
        )
        row[f"test_class_{class_index}_count"] = (
            int(result.test_metrics[f"class_{class_index}_count"])
            if result.test_metrics
            else int(bundle.split.test_sampled_counts[class_index - 1])
        )
        row[f"test_class_{class_index}_accuracy"] = (
            format_scalar(result.test_metrics[f"class_{class_index}_accuracy"]) if result.test_metrics else None
        )

    return row


def log_metrics(tee: TeeStream, label: str, metrics: dict[str, Any], n_classes: int) -> None:
    print(f"{label} total accuracy: {metrics['total_accuracy']:.4f}", file=tee)
    for class_index in range(1, n_classes + 1):
        print(
            f"  Class {class_index}: {metrics[f'class_{class_index}_accuracy']:.4f} "
            f"({metrics[f'class_{class_index}_count']} samples)",
            file=tee,
        )


def run_svor(bundle: DatasetBundle, tee: TeeStream, svor_params: dict[str, Any]) -> ModelRunResult:
    start = utc_now()
    model_name = "SVOR"
    banner(tee, f"{bundle.split.dataset} | {model_name}", major=False)

    try:
        print(
            f"Total SVOR time budget: {svor_params['solver_params']['TimeLimit']}s (single optimization problem)",
            file=tee,
        )
        model = SVORImplicitQP(**svor_params)
        with redirect_stdout(tee), redirect_stderr(tee):
            model.fit(bundle.X_train.tolist(), bundle.y_train.tolist())

        train_predictions = np.asarray(model.predict(bundle.X_train.tolist()), dtype=int)
        test_predictions = np.asarray(model.predict(bundle.X_test.tolist()), dtype=int)
        train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
        test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)
        fit_seconds = (utc_now() - start).total_seconds()
        status_label = model_status_label(getattr(model, "model_status_", None), has_solution=True)

        log_metrics(tee, "Training", train_metrics, bundle.n_classes)
        log_metrics(tee, "Testing", test_metrics, bundle.n_classes)
        print(f"SVOR status: {status_label}", file=tee)
        print(f"Weights (w): {format_vector(model.weights_)}", file=tee)
        print(f"Thresholds (b): {format_vector(model.thresholds_)}", file=tee)

        return ModelRunResult(
            dataset=bundle.split.dataset,
            model=model_name,
            status="completed",
            fit_seconds=fit_seconds,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=[
                {
                    "dataset": bundle.split.dataset,
                    "source_dataset": bundle.split.source_dataset,
                    "model": model_name,
                    "component": "GLOBAL",
                    "description": "Shared hyperplane with ordered thresholds",
                    "status": "completed",
                    "fit_seconds": fit_seconds,
                    "solver_status_code": int(model.model_status_),
                    "solver_status_label": status_label,
                    "weights": format_vector(model.weights_),
                    "b": format_vector(model.thresholds_),
                    "objective_value": format_scalar(model.objective_value_),
                    "prediction_rule": "ordered_thresholds",
                }
            ],
        )
    except Exception:
        error = traceback.format_exc()
        fit_seconds = (utc_now() - start).total_seconds()
        print(error, file=tee)
        return ModelRunResult(
            dataset=bundle.split.dataset,
            model=model_name,
            status="failed",
            fit_seconds=fit_seconds,
            train_metrics=None,
            test_metrics=None,
            parameter_rows=[],
            error=error,
        )


def run_npsvor(bundle: DatasetBundle, tee: TeeStream, npsvor_params: dict[str, Any]) -> ModelRunResult:
    start = utc_now()
    model_name = "NPSVOR"
    banner(tee, f"{bundle.split.dataset} | {model_name}", major=False)

    try:
        total_time_limit = int(npsvor_params["solver_params"]["TimeLimit"])
        subproblem_count = max(bundle.n_classes, 1)
        per_rank_time_limit = max(1, total_time_limit // subproblem_count)
        actual_params = dict(npsvor_params)
        actual_params["solver_params"] = dict(npsvor_params["solver_params"])
        actual_params["solver_params"]["TimeLimit"] = per_rank_time_limit

        print(
            f"Total NPSVOR time budget: {total_time_limit}s; "
            f"per rank time limit: {per_rank_time_limit}s across {subproblem_count} ranks",
            file=tee,
        )
        model = NPSVORQP(**actual_params)
        with redirect_stdout(tee), redirect_stderr(tee):
            model.fit(bundle.X_train.tolist(), bundle.y_train.tolist())

        train_predictions = np.asarray(model.predict(bundle.X_train.tolist()), dtype=int)
        test_predictions = np.asarray(model.predict(bundle.X_test.tolist()), dtype=int)
        train_metrics = evaluate_multiclass(bundle.y_train, train_predictions, n_classes=bundle.n_classes)
        test_metrics = evaluate_multiclass(bundle.y_test, test_predictions, n_classes=bundle.n_classes)
        fit_seconds = (utc_now() - start).total_seconds()

        log_metrics(tee, "Training", train_metrics, bundle.n_classes)
        log_metrics(tee, "Testing", test_metrics, bundle.n_classes)

        parameter_rows: list[dict[str, Any]] = []
        for rank_index, (weights, bias, objective_value, status_code) in enumerate(
            zip(
                model.hyperplanes_,
                model.biases_,
                model.objective_values_,
                model.model_statuses_,
            ),
            start=1,
        ):
            status_label = model_status_label(int(status_code), has_solution=True)
            print(f"Rank {rank_index} status: {status_label}", file=tee)
            print(f"Rank {rank_index} weights: {format_vector(weights)}", file=tee)
            print(f"Rank {rank_index} b: {bias:.10f}", file=tee)
            parameter_rows.append(
                {
                    "dataset": bundle.split.dataset,
                    "source_dataset": bundle.split.source_dataset,
                    "model": model_name,
                    "component": f"RANK_{rank_index}",
                    "description": f"Non-parallel hyperplane for class {rank_index}",
                    "status": "completed",
                    "fit_seconds": fit_seconds,
                    "solver_status_code": int(status_code),
                    "solver_status_label": status_label,
                    "weights": format_vector(weights),
                    "b": format_scalar(bias),
                    "objective_value": format_scalar(objective_value),
                    "prediction_rule": actual_params["prediction_rule"],
                }
            )

        return ModelRunResult(
            dataset=bundle.split.dataset,
            model=model_name,
            status="completed",
            fit_seconds=fit_seconds,
            train_metrics=train_metrics,
            test_metrics=test_metrics,
            parameter_rows=parameter_rows,
        )
    except Exception:
        error = traceback.format_exc()
        fit_seconds = (utc_now() - start).total_seconds()
        print(error, file=tee)
        return ModelRunResult(
            dataset=bundle.split.dataset,
            model=model_name,
            status="failed",
            fit_seconds=fit_seconds,
            train_metrics=None,
            test_metrics=None,
            parameter_rows=[],
            error=error,
        )


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = min(max(len(value) for value in values), 100)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2


def write_excel(
    output_path: Path,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
) -> None:
    workbook = Workbook()

    metrics_sheet = workbook.active
    metrics_sheet.title = "Metrics"
    metrics_sheet.append(metric_headers)
    for row in metric_rows:
        metrics_sheet.append([row.get(header) for header in metric_headers])

    params_sheet = workbook.create_sheet("Parameters")
    params_sheet.append(parameter_headers)
    for row in parameter_rows:
        params_sheet.append([row.get(header) for header in parameter_headers])

    metadata_sheet = workbook.create_sheet("Run_Metadata")
    metadata_sheet.append(["key", "value"])
    for key, value in metadata_rows:
        metadata_sheet.append([key, value])

    for sheet in (metrics_sheet, params_sheet, metadata_sheet):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        autosize_worksheet(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = rows[0]
    return [dict(zip(headers, row)) for row in rows[1:]]


def load_matching_hcesvm_metrics(split: ManifestSplit) -> dict[str, Any] | None:
    manifest_name = split.manifest_path.name
    if "_manifest_" not in manifest_name:
        return None

    timestamp = manifest_name.rsplit("_manifest_", 1)[1].removesuffix(".json")
    workbook_path = DOCS_DIR / f"HCESVM_TEST3_1000_VALIDATION_{timestamp}.xlsx"
    if not workbook_path.exists():
        return None

    for row in sheet_records(workbook_path, "Metrics"):
        if row.get("dataset") == split.dataset and row.get("model") == "HCESVM(test3)":
            row["workbook_path"] = str(workbook_path)
            return row
    return None


def render_accuracy_table(result: ModelRunResult, n_classes: int) -> str:
    if result.train_metrics is None or result.test_metrics is None:
        return "_No accuracy table because the run failed._"

    lines = [
        format_report_table_row(["Split", "Total Accuracy", "Per-Class Accuracy"]),
        format_report_table_row(["---", "---:", "---"]),
    ]
    for split_name, metrics in [("Train", result.train_metrics), ("Test", result.test_metrics)]:
        per_class = ", ".join(
            f"C{class_index}={metrics[f'class_{class_index}_accuracy']:.4f} "
            f"(n={metrics[f'class_{class_index}_count']})"
            for class_index in range(1, n_classes + 1)
        )
        lines.append(
            format_report_table_row(
                [
                    split_name,
                    f"{metrics['total_accuracy']:.4f}",
                    per_class,
                ]
            )
        )
    return "\n".join(lines)


def render_parameter_table(result: ModelRunResult) -> str:
    if not result.parameter_rows:
        return "_No parameter rows were recorded._"

    lines = [
        format_report_table_row(
            ["Component", "Description", "Status", "Objective", "b", "weights"]
        ),
        format_report_table_row(["---", "---", "---", "---:", "---", "---"]),
    ]
    for row in result.parameter_rows:
        lines.append(
            format_report_table_row(
                [
                    str(row["component"]),
                    str(row["description"]),
                    str(row["solver_status_label"]),
                    format_optional_float(row["objective_value"], decimals=6),
                    str(row["b"]),
                    str(row["weights"]),
                ]
            )
        )
    return "\n".join(lines)


def substantially_outperforms(
    result: ModelRunResult,
    hcesvm_metrics: dict[str, Any] | None,
) -> bool:
    if (
        hcesvm_metrics is None
        or result.train_metrics is None
        or result.test_metrics is None
    ):
        return False

    train_delta = float(result.train_metrics["total_accuracy"]) - float(hcesvm_metrics["train_total_accuracy"])
    test_delta = float(result.test_metrics["total_accuracy"]) - float(hcesvm_metrics["test_total_accuracy"])
    return max(train_delta, test_delta) >= SUBSTANTIAL_OUTPERFORM_DELTA


def c4_near_zero(result: ModelRunResult) -> bool:
    if result.train_metrics is None or result.test_metrics is None:
        return False
    return (
        float(result.train_metrics["class_4_accuracy"]) <= NEAR_ZERO_ACCURACY
        and float(result.test_metrics["class_4_accuracy"]) <= NEAR_ZERO_ACCURACY
    )


def render_interpretation(
    results: list[ModelRunResult],
    split: ManifestSplit,
    hcesvm_metrics: dict[str, Any] | None,
) -> str:
    if not results:
        return "\n".join(
            [
                "## Interpretation",
                "",
                f"- Exact split checked: train `{split.train_sampled_counts}` and test `{split.test_sampled_counts}`.",
                "- Baseline runs have not completed yet.",
            ]
        )

    result_by_name = {result.model: result for result in results}
    svor = result_by_name.get("SVOR")
    npsvor = result_by_name.get("NPSVOR")

    lines = [
        "## Interpretation",
        "",
        f"- Exact split checked: train `{split.train_sampled_counts}` and test `{split.test_sampled_counts}`.",
    ]

    if svor and svor.train_metrics and svor.test_metrics:
        lines.append(
            f"- SVOR C4 accuracy: train `{svor.train_metrics['class_4_accuracy']:.4f}`, "
            f"test `{svor.test_metrics['class_4_accuracy']:.4f}`."
        )
    if npsvor and npsvor.train_metrics and npsvor.test_metrics:
        lines.append(
            f"- NPSVOR C4 accuracy: train `{npsvor.train_metrics['class_4_accuracy']:.4f}`, "
            f"test `{npsvor.test_metrics['class_4_accuracy']:.4f}`."
        )

    if hcesvm_metrics is not None:
        lines.append(
            f"- Current HCESVM(test3) on the same split: train total `{float(hcesvm_metrics['train_total_accuracy']):.4f}`, "
            f"test total `{float(hcesvm_metrics['test_total_accuracy']):.4f}`, "
            f"C4 train/test `{float(hcesvm_metrics['train_class_4_accuracy']):.4f}` / "
            f"`{float(hcesvm_metrics['test_class_4_accuracy']):.4f}`."
        )

        outperforming = [result.model for result in results if substantially_outperforms(result, hcesvm_metrics)]
        if outperforming:
            lines.append(
                f"- Ordinal baselines that substantially outperform HCESVM on total accuracy: `{', '.join(outperforming)}`."
            )
        else:
            lines.append(
                "- Neither ordinal baseline substantially outperforms the current HCESVM result on total accuracy."
            )

        if all(c4_near_zero(result) for result in results if result.model in {"SVOR", "NPSVOR"}):
            if outperforming:
                lines.append(
                    "- Minority-only explanation is incomplete: C4 still collapses for both baselines, but at least one ordinal baseline still improves total accuracy noticeably over HCESVM."
                )
            else:
                lines.append(
                    "- Minority-only explanation still looks plausible on this exact split because C4 stays near zero for both baselines and overall accuracy does not move much above HCESVM."
                )
        else:
            lines.append(
                "- Minority-only explanation does not hold by itself on this exact split because at least one ordinal baseline recovers non-trivial C4 accuracy."
            )
    else:
        lines.append(
            "- Matching HCESVM workbook for the same timestamp was not found, so the minority-hypothesis comparison is limited to SVOR/NPSVOR only."
        )

    return "\n".join(lines)


def render_markdown_report(
    *,
    generated_at: str,
    artifact_paths: ArtifactPaths,
    split: ManifestSplit,
    results: list[ModelRunResult],
    hcesvm_metrics: dict[str, Any] | None,
) -> str:
    lines = [
        "# SKILL Method3 4-Class 1000 Split Baselines",
        "",
        f"Generated at: `{generated_at}`",
        f"Workbook: `{artifact_paths.workbook_path}`",
        f"Log: `{artifact_paths.log_path}`",
        f"Manifest: `{split.manifest_path}`",
        "",
        f"- Source dataset: `{split.source_dataset}`",
        f"- Source train CSV: `{split.train_csv}`",
        f"- Source test CSV: `{split.test_csv}`",
        f"- Relabel map: `{split.relabel_summary}`",
        f"- Train counts: `{split.train_original_counts} -> {split.train_sampled_counts}`",
        f"- Test counts: `{split.test_original_counts} -> {split.test_sampled_counts}`",
        "",
    ]

    for result in results:
        lines.extend(
            [
                f"## {result.model}",
                "",
                f"- Final status: `{result.status}`",
                f"- Fit seconds: `{format_optional_float(result.fit_seconds, decimals=3)}`",
                "",
                render_accuracy_table(result, split.n_classes),
                "",
                render_parameter_table(result),
                "",
            ]
        )
        if result.error:
            lines.extend(
                [
                    "```text",
                    result.error.strip(),
                    "```",
                    "",
                ]
            )

    lines.append(render_interpretation(results, split, hcesvm_metrics))
    lines.append("")
    return "\n".join(lines)


def checkpoint_artifacts(
    artifact_paths: ArtifactPaths,
    *,
    metric_headers: list[str],
    metric_rows: list[dict[str, Any]],
    parameter_headers: list[str],
    parameter_rows: list[dict[str, Any]],
    metadata_rows: list[tuple[str, Any]],
    split: ManifestSplit,
    results: list[ModelRunResult],
    hcesvm_metrics: dict[str, Any] | None,
    tee: TeeStream,
    label: str,
    generated_at: str,
) -> None:
    write_excel(
        artifact_paths.workbook_path,
        metric_headers=metric_headers,
        metric_rows=metric_rows,
        parameter_headers=parameter_headers,
        parameter_rows=parameter_rows,
        metadata_rows=metadata_rows,
    )
    artifact_paths.markdown_path.write_text(
        render_markdown_report(
            generated_at=generated_at,
            artifact_paths=artifact_paths,
            split=split,
            results=results,
            hcesvm_metrics=hcesvm_metrics,
        ),
        encoding="utf-8",
    )
    print(f"Checkpoint saved after {label}: {artifact_paths.workbook_path}", file=tee)
    print(f"Checkpoint markdown after {label}: {artifact_paths.markdown_path}", file=tee)


def build_artifact_paths(started_at: datetime) -> ArtifactPaths:
    timestamp = short_timestamp(started_at)
    date_prefix = started_at.strftime("%Y%m%d")
    archive_dir = RESULTS_ARCHIVE_DIR / f"{date_prefix}_svor_npsvor_skill_1000"
    archive_dir.mkdir(parents=True, exist_ok=True)
    return ArtifactPaths(
        log_path=archive_dir / f"svor_npsvor_skill_method3_4class_1000_{timestamp}.log",
        workbook_path=DOCS_DIR / f"SKILL_METHOD3_4CLASS_1000_SVOR_NPSVOR_{timestamp}.xlsx",
        markdown_path=DOCS_DIR / f"SKILL_METHOD3_4CLASS_1000_SVOR_NPSVOR_{timestamp}.md",
    )


def run_manifest_experiment(
    *,
    manifest_json: str | Path,
    svor_time_limit: int,
    npsvor_time_limit: int,
    started_at: datetime | None = None,
) -> dict[str, Any]:
    started = utc_now() if started_at is None else started_at
    split = load_manifest_split(manifest_json)
    bundle = load_bundle(split)
    artifact_paths = build_artifact_paths(started)
    metric_headers = build_metric_headers(bundle.n_classes)
    parameter_headers = build_parameter_headers()
    metric_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    results: list[ModelRunResult] = []
    generated_at = human_timestamp(started)
    hcesvm_metrics = load_matching_hcesvm_metrics(split)

    artifact_paths.log_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_paths.workbook_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_paths.markdown_path.parent.mkdir(parents=True, exist_ok=True)

    svor_params = dict(DEFAULT_SVOR_PARAMS)
    svor_params["solver_params"] = dict(DEFAULT_SVOR_PARAMS["solver_params"])
    svor_params["solver_params"]["TimeLimit"] = svor_time_limit
    npsvor_params = dict(DEFAULT_NPSVOR_PARAMS)
    npsvor_params["solver_params"] = dict(DEFAULT_NPSVOR_PARAMS["solver_params"])
    npsvor_params["solver_params"]["TimeLimit"] = npsvor_time_limit

    metadata = {
        "generated_at_utc": generated_at,
        "branch": git_output("git", "rev-parse", "--abbrev-ref", "HEAD"),
        "commit": git_output("git", "rev-parse", "HEAD"),
        "worktree": str(ROOT),
        "log_path": str(artifact_paths.log_path),
        "excel_path": str(artifact_paths.workbook_path),
        "report_path": str(artifact_paths.markdown_path),
        "manifest_json": str(split.manifest_path),
        "source_dataset": split.source_dataset,
        "source_url": split.source_url,
        "train_csv": str(split.train_csv),
        "test_csv": str(split.test_csv),
        "train_sampled_counts": json.dumps(split.train_sampled_counts, ensure_ascii=False),
        "test_sampled_counts": json.dumps(split.test_sampled_counts, ensure_ascii=False),
        "relabel_map": split.relabel_summary,
        "svor_params": json.dumps(svor_params, ensure_ascii=False),
        "npsvor_params": json.dumps(npsvor_params, ensure_ascii=False),
        "hcesvm_reference_workbook": "" if hcesvm_metrics is None else hcesvm_metrics["workbook_path"],
    }
    metadata_rows = build_metadata_rows(metadata)

    with artifact_paths.log_path.open("w", encoding="utf-8", buffering=1) as log_handle:
        tee = TeeStream(sys.stdout, log_handle)
        banner(tee, "SKILL Method3 4-class 1000 split baselines")
        print(f"Branch: {metadata['branch']}", file=tee)
        print(f"Commit: {metadata['commit']}", file=tee)
        print(f"Manifest: {split.manifest_path}", file=tee)
        print(f"Source dataset: {split.source_dataset}", file=tee)
        print(f"Train CSV: {split.train_csv}", file=tee)
        print(f"Test CSV: {split.test_csv}", file=tee)
        print(f"Target column: {TARGET_COLUMN}", file=tee)
        print(f"Features inferred from CSV; manifest listed {len(split.feature_names)} feature columns", file=tee)
        print(f"Train counts: {split.train_sampled_counts}", file=tee)
        print(f"Test counts: {split.test_sampled_counts}", file=tee)
        print(f"Relabel map: {split.relabel_summary}", file=tee)
        print(f"SVOR params: {json.dumps(svor_params, ensure_ascii=False)}", file=tee)
        print(f"NPSVOR params: {json.dumps(npsvor_params, ensure_ascii=False)}", file=tee)
        if hcesvm_metrics is not None:
            print(
                f"HCESVM comparison workbook: {hcesvm_metrics['workbook_path']} "
                f"(train={float(hcesvm_metrics['train_total_accuracy']):.4f}, "
                f"test={float(hcesvm_metrics['test_total_accuracy']):.4f})",
                file=tee,
            )
        checkpoint_artifacts(
            artifact_paths,
            metric_headers=metric_headers,
            metric_rows=metric_rows,
            parameter_headers=parameter_headers,
            parameter_rows=parameter_rows,
            metadata_rows=metadata_rows,
            split=split,
            results=results,
            hcesvm_metrics=hcesvm_metrics,
            tee=tee,
            label="initialization",
            generated_at=generated_at,
        )

        for runner, params in ((run_svor, svor_params), (run_npsvor, npsvor_params)):
            result = runner(bundle, tee, params)
            results.append(result)
            metric_rows.append(collect_metric_row(bundle, result))
            parameter_rows.extend(result.parameter_rows)
            checkpoint_artifacts(
                artifact_paths,
                metric_headers=metric_headers,
                metric_rows=metric_rows,
                parameter_headers=parameter_headers,
                parameter_rows=parameter_rows,
                metadata_rows=metadata_rows,
                split=split,
                results=results,
                hcesvm_metrics=hcesvm_metrics,
                tee=tee,
                label=result.model,
                generated_at=generated_at,
            )

        banner(tee, "Run complete")
        print(f"Log saved to: {artifact_paths.log_path}", file=tee)
        print(f"Workbook saved to: {artifact_paths.workbook_path}", file=tee)
        print(f"Markdown saved to: {artifact_paths.markdown_path}", file=tee)

    return {
        "artifacts": artifact_paths,
        "split": split,
        "results": results,
        "hcesvm_metrics": hcesvm_metrics,
    }


def main(argv: list[str] | None = None) -> int:
    args = parse_arguments(argv)
    run_manifest_experiment(
        manifest_json=args.manifest_json,
        svor_time_limit=args.svor_time_limit,
        npsvor_time_limit=args.npsvor_time_limit,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
