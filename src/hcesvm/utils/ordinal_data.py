"""Ordinal data loaders for LINGO workbooks and generic tabular files."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


_LINGO_METADATA_NAMES = {
    "class",
    "actual",
    "predict",
    "predict ",
    "ksi1",
    "ksi2",
    "ksi+",
    "ksi-",
    "eta21/12/13",
    "eta31/32/23",
    "alpha+",
    "alpha-",
    "beta+",
    "beta-",
    "k1",
    "k2",
    "k3",
}


@dataclass
class OrdinalDatasetSplit:
    workbook_path: Path
    feature_names: list[str]
    X_train: list[list[float]]
    y_train: list[int]
    X_test: list[list[float]]
    y_test: list[int]
    class_token_to_label: dict[str, int]
    reported_train_accuracy: float | None = None
    reported_test_accuracy: float | None = None

    @property
    def classes(self) -> list[int]:
        return sorted(set(self.y_train))


@dataclass
class _SheetData:
    feature_names: list[str]
    X: list[list[float]]
    y: list[int]
    token_to_label: dict[str, int]


def _take_first_per_class(
    X: list[list[float]],
    y: list[int],
    limit: int | None,
) -> tuple[list[list[float]], list[int]]:
    if limit is None:
        return list(X), list(y)

    selected_X: list[list[float]] = []
    selected_y: list[int] = []
    per_class: dict[int, int] = {}

    for row, label in zip(X, y):
        if per_class.get(label, 0) >= limit:
            continue

        per_class[label] = per_class.get(label, 0) + 1
        selected_X.append(row)
        selected_y.append(label)

    return selected_X, selected_y


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_header(value: Any) -> str:
    return value.strip().lower() if isinstance(value, str) else ""


def _resolve_column_name(columns: list[str], requested_name: str) -> str:
    normalized = _normalize_header(requested_name)
    for column in columns:
        if _normalize_header(column) == normalized:
            return column
    raise ValueError(f"Column {requested_name!r} was not found. Available columns: {columns!r}.")


def _coerce_float(value: Any, *, column_name: str) -> float:
    if _is_blank(value):
        raise ValueError(f"Column {column_name!r} contains a blank value.")
    if _is_number(value):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        try:
            return float(stripped)
        except ValueError as exc:
            raise ValueError(f"Column {column_name!r} must be numeric, got {value!r}.") from exc
    raise ValueError(f"Column {column_name!r} must be numeric, got {value!r}.")


def _can_coerce_float(value: Any) -> bool:
    if _is_blank(value):
        return False
    if _is_number(value):
        return True
    if isinstance(value, str):
        try:
            float(value.strip())
            return True
        except ValueError:
            return False
    return False


def _infer_target_column(columns: list[str]) -> str:
    preferred_names = ["target", "label", "labels", "y", "class", "actual", "rating"]
    for preferred in preferred_names:
        for column in columns:
            if _normalize_header(column) == preferred:
                return column

    raise ValueError(
        "Could not infer the target column. Pass target_column explicitly. "
        f"Available columns: {columns!r}."
    )


def _infer_feature_columns(rows: list[dict[str, Any]], target_column: str) -> list[str]:
    if not rows:
        raise ValueError("At least one data row is required.")

    feature_columns: list[str] = []
    for column in rows[0]:
        if column == target_column:
            continue
        if all(_is_blank(row.get(column)) or _can_coerce_float(row.get(column)) for row in rows):
            feature_columns.append(column)

    if not feature_columns:
        raise ValueError(
            "Could not infer numeric feature columns. Pass feature_columns explicitly or provide numeric columns."
        )

    return feature_columns


def _extract_reported_accuracy(worksheet) -> float | None:
    for row in worksheet.iter_rows(min_row=1, max_row=min(8, worksheet.max_row), values_only=True):
        values = list(row)
        for index, value in enumerate(values[:-1]):
            if _normalize_header(value) in {"total", "total accuracy"} and _is_number(values[index + 1]):
                return float(values[index + 1])
    return None


def _convert_label(value: Any) -> int:
    if _is_number(value):
        numeric = float(value)
    elif isinstance(value, str):
        stripped = value.strip()
        try:
            numeric = float(stripped)
        except ValueError as exc:
            raise ValueError(
                f"Expected a numeric ordinal label, got {value!r}. "
                "For generic CSV/Excel input, labels must be integral numbers."
            ) from exc
    else:
        raise ValueError(f"Expected a numeric ordinal label, got {value!r}.")

    rounded = round(numeric)
    if abs(numeric - rounded) > 1e-9:
        raise ValueError(f"Ordinal labels must be integral, got {value!r}.")
    return int(rounded)


def _rows_from_csv(path: Path, *, delimiter: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        if delimiter is None:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.get_dialect("excel")
            reader = csv.DictReader(handle, dialect=dialect)
        else:
            reader = csv.DictReader(handle, delimiter=delimiter)

        if reader.fieldnames is None:
            raise ValueError(f"CSV file {path} must have a header row.")

        headers = [str(name) for name in reader.fieldnames]
        rows = [{str(key): value for key, value in row.items()} for row in reader]

    return headers, rows


def _rows_from_excel(path: Path, *, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name is None:
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {sheet_name!r} was not found in workbook {path}.")
        worksheet = workbook[sheet_name]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Worksheet {worksheet.title!r} is empty.")

    header_row = rows[0]
    headers = [
        str(value).strip() if value is not None else f"column_{index + 1}"
        for index, value in enumerate(header_row)
    ]
    dict_rows: list[dict[str, Any]] = []

    for row in rows[1:]:
        values = list(row)
        if not any(value is not None for value in values):
            continue
        padded = values + [None] * (len(headers) - len(values))
        dict_rows.append({headers[index]: padded[index] for index in range(len(headers))})

    return headers, dict_rows


def _load_tabular_rows(
    path: Path,
    *,
    sheet_name: str | None = None,
    delimiter: str | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(path, delimiter=delimiter)
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_excel(path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported file type {suffix!r}. Supported types are .csv, .xlsx, and .xlsm.")


def _parse_tabular_rows(
    rows: list[dict[str, Any]],
    *,
    target_column: str | None,
    feature_columns: list[str] | None,
) -> tuple[list[str], list[list[float]], list[int]]:
    if not rows:
        raise ValueError("At least one data row is required.")

    columns = list(rows[0].keys())
    resolved_target = _resolve_column_name(columns, target_column) if target_column else _infer_target_column(columns)
    resolved_features = (
        [_resolve_column_name(columns, column_name) for column_name in feature_columns]
        if feature_columns is not None
        else _infer_feature_columns(rows, resolved_target)
    )

    X: list[list[float]] = []
    y: list[int] = []

    for row in rows:
        target_value = row.get(resolved_target)
        if _is_blank(target_value):
            continue

        features = []
        skip_row = False
        for feature_name in resolved_features:
            value = row.get(feature_name)
            if _is_blank(value):
                skip_row = True
                break
            features.append(_coerce_float(value, column_name=feature_name))
        if skip_row:
            continue

        X.append(features)
        y.append(_convert_label(target_value))

    if not X:
        raise ValueError("No usable tabular rows were found after filtering blank rows.")

    return resolved_features, X, y


def _find_header_row(worksheet) -> tuple[int, list[Any]]:
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        normalized = [_normalize_header(value) for value in values]
        if "class" not in normalized and "actual" not in normalized:
            continue

        metadata_positions = [
            index for index, value in enumerate(normalized) if value in _LINGO_METADATA_NAMES
        ]
        if metadata_positions and min(metadata_positions) > 0:
            return row_index, values

    raise ValueError(f"Could not find a LINGO feature header row in worksheet {worksheet.title!r}.")


def _find_column_index(header: list[Any], candidates: list[str]) -> int | None:
    normalized = [_normalize_header(value) for value in header]
    for candidate in candidates:
        for index, value in enumerate(normalized):
            if value == candidate:
                return index
    return None


def _find_feature_end_index(header: list[Any]) -> int:
    normalized = [_normalize_header(value) for value in header]
    metadata_positions = [
        index for index, value in enumerate(normalized) if value in _LINGO_METADATA_NAMES
    ]
    if not metadata_positions:
        raise ValueError("Could not identify where the feature columns stop in the LINGO worksheet.")

    feature_end = min(metadata_positions)
    if feature_end == 0:
        raise ValueError("LINGO worksheet does not appear to contain any feature columns.")
    return feature_end


def _parse_sheet(worksheet, token_to_label: dict[str, int] | None = None) -> _SheetData:
    header_row, header = _find_header_row(worksheet)
    feature_end = _find_feature_end_index(header)
    class_col = _find_column_index(header, ["class"])
    actual_col = _find_column_index(header, ["actual"])
    label_col = actual_col if actual_col is not None else class_col
    if label_col is None:
        raise ValueError(f"Could not find label columns in worksheet {worksheet.title!r}.")

    feature_names = [str(value).strip() for value in header[:feature_end] if not _is_blank(value)]
    mapping = dict(token_to_label or {})
    X: list[list[float]] = []
    y: list[int] = []

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(row)
        scan_limit = max(feature_end, label_col + 1, (class_col + 1) if class_col is not None else 0)
        scanned_values = values[:scan_limit]
        if not any(value is not None for value in scanned_values):
            if X:
                break
            continue

        features = values[:feature_end] + [None] * max(0, feature_end - len(values))
        if any(_is_blank(value) for value in features):
            continue
        if not all(_is_number(value) for value in features):
            continue

        class_token = values[class_col] if class_col is not None and class_col < len(values) else None
        token_key = str(class_token).strip() if not _is_blank(class_token) else ""

        label_value = values[label_col] if label_col < len(values) else None
        if _is_blank(label_value):
            continue

        try:
            label = _convert_label(label_value)
            mapping[token_key or str(label)] = label
        except ValueError:
            if token_key in mapping:
                label = mapping[token_key]
            elif token_key:
                mapping[token_key] = len(mapping) + 1
                label = mapping[token_key]
            else:
                raise

        X.append([float(value) for value in features])
        y.append(label)

    if not X:
        raise ValueError(f"No data rows found in worksheet {worksheet.title!r}.")

    return _SheetData(
        feature_names=feature_names,
        X=X,
        y=y,
        token_to_label=mapping,
    )


def load_lingo_split_workbook(path: str | Path) -> OrdinalDatasetSplit:
    workbook_path = Path(path).expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Train" not in workbook.sheetnames or "Test" not in workbook.sheetnames:
        raise ValueError(f"Workbook {workbook_path} must contain Train and Test sheets.")

    train_sheet = workbook["Train"]
    test_sheet = workbook["Test"]

    train = _parse_sheet(train_sheet)
    test = _parse_sheet(test_sheet, token_to_label=train.token_to_label)

    if train.feature_names != test.feature_names:
        raise ValueError(
            f"Feature mismatch between Train and Test sheets: {train.feature_names!r} vs {test.feature_names!r}."
        )

    return OrdinalDatasetSplit(
        workbook_path=workbook_path,
        feature_names=train.feature_names,
        X_train=train.X,
        y_train=train.y,
        X_test=test.X,
        y_test=test.y,
        class_token_to_label=train.token_to_label,
        reported_train_accuracy=_extract_reported_accuracy(train_sheet),
        reported_test_accuracy=_extract_reported_accuracy(test_sheet),
    )


def load_tabular_dataset_split(
    train_path: str | Path,
    *,
    test_path: str | Path | None = None,
    target_column: str | None = None,
    feature_columns: list[str] | None = None,
    train_sheet: str | None = None,
    test_sheet: str | None = None,
    delimiter: str | None = None,
) -> OrdinalDatasetSplit:
    train_file = Path(train_path).expanduser().resolve()
    inferred_test_path = Path(test_path).expanduser().resolve() if test_path is not None else None

    if inferred_test_path is None and train_file.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(train_file, read_only=True, data_only=True)
        if train_sheet is None and "Train" in workbook.sheetnames:
            train_sheet = "Train"
        if test_sheet is None and "Test" in workbook.sheetnames:
            test_sheet = "Test"

    train_headers, train_rows = _load_tabular_rows(train_file, sheet_name=train_sheet, delimiter=delimiter)
    resolved_target_column = target_column or _infer_target_column(train_headers)
    resolved_feature_columns = (
        [_resolve_column_name(train_headers, column_name) for column_name in feature_columns]
        if feature_columns is not None
        else None
    )
    feature_names, X_train, y_train = _parse_tabular_rows(
        train_rows,
        target_column=resolved_target_column,
        feature_columns=resolved_feature_columns,
    )

    if inferred_test_path is not None:
        _, test_rows = _load_tabular_rows(inferred_test_path, sheet_name=test_sheet, delimiter=delimiter)
    elif test_sheet is not None and test_sheet != train_sheet:
        _, test_rows = _load_tabular_rows(train_file, sheet_name=test_sheet, delimiter=delimiter)
    else:
        test_rows = train_rows

    _, X_test, y_test = _parse_tabular_rows(
        test_rows,
        target_column=resolved_target_column,
        feature_columns=feature_names,
    )

    return OrdinalDatasetSplit(
        workbook_path=train_file,
        feature_names=feature_names,
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        class_token_to_label={str(label): label for label in sorted(set(y_train))},
        reported_train_accuracy=None,
        reported_test_accuracy=None,
    )


def subset_samples_per_class(
    dataset: OrdinalDatasetSplit,
    *,
    train_limit: int | None = None,
    test_limit: int | None = None,
) -> OrdinalDatasetSplit:
    X_train, y_train = _take_first_per_class(dataset.X_train, dataset.y_train, train_limit)
    X_test, y_test = _take_first_per_class(dataset.X_test, dataset.y_test, test_limit)

    return OrdinalDatasetSplit(
        workbook_path=dataset.workbook_path,
        feature_names=list(dataset.feature_names),
        X_train=X_train,
        y_train=y_train,
        X_test=X_test,
        y_test=y_test,
        class_token_to_label=dict(dataset.class_token_to_label),
        reported_train_accuracy=None,
        reported_test_accuracy=None,
    )
