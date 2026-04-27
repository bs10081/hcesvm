"""Post-run helpers for teaching-data deadline experiments."""

from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook


BASELINE_HCESVM_TOTAL_BUDGET_SECONDS = 1800
REPORT_TIMESTAMP_RE = re.compile(r"_(\d{8}_\d{6})\.xlsx$")
FINAL_REPORT_STATUSES = {"completed", "success", "stopped_by_deadline", "failed"}


@dataclass(slots=True)
class WorkbookRunData:
    """Normalized teaching-data workbook payload."""

    path: Path
    metrics_row: dict[str, Any] | None
    parameter_rows: list[dict[str, Any]]
    progress_rows: list[dict[str, Any]]
    metadata: dict[str, Any]

    @property
    def dataset(self) -> str | None:
        if self.metrics_row is not None:
            value = self.metrics_row.get("dataset")
            if value is not None:
                return str(value)
        value = self.metadata.get("dataset")
        return None if value is None else str(value)

    @property
    def model(self) -> str | None:
        if self.metrics_row is not None:
            value = self.metrics_row.get("model")
            if value is not None:
                return str(value)
        return None

    @property
    def final_status(self) -> str:
        metadata_status = self.metadata.get("final_status")
        if metadata_status not in (None, ""):
            return str(metadata_status)
        if self.metrics_row is not None and self.metrics_row.get("status") not in (None, ""):
            return str(self.metrics_row["status"])
        return "unknown"

    @property
    def is_final(self) -> bool:
        return self.final_status in FINAL_REPORT_STATUSES

    @property
    def n_classes(self) -> int:
        if self.metrics_row is None:
            value = self.metadata.get("n_classes")
            if value is None:
                raise ValueError(f"n_classes missing in {self.path}")
            return int(value)
        return int(self.metrics_row["n_classes"])

    @property
    def configured_time_limit_seconds(self) -> int | None:
        params_raw = self.metadata.get("hcesvm_params")
        if params_raw in (None, ""):
            return None
        try:
            params = json.loads(str(params_raw))
        except json.JSONDecodeError:
            return None
        value = params.get("time_limit")
        return None if value is None else int(value)


def parse_report_timestamp(path: Path) -> datetime | None:
    """Extract the `YYYYMMDD_HHMMSS` timestamp from a report filename."""
    match = REPORT_TIMESTAMP_RE.search(path.name)
    if not match:
        return None
    return datetime.strptime(match.group(1), "%Y%m%d_%H%M%S")


def parse_compact_timestamp(value: str) -> datetime:
    """Parse a compact `YYYYMMDD_HHMMSS` timestamp string."""
    return datetime.strptime(value, "%Y%m%d_%H%M%S")


def round_up_seconds(seconds: float, *, step_seconds: int = 300) -> int:
    """Round seconds up to the nearest step."""
    if seconds <= 0:
        return 0
    return int(math.ceil(float(seconds) / step_seconds) * step_seconds)


def format_duration_label(seconds: int | None) -> str:
    """Render a compact human-friendly duration label."""
    if seconds is None:
        return ""
    total_minutes = int(math.ceil(seconds / 60))
    hours, minutes = divmod(total_minutes, 60)
    if hours and minutes:
        return f"{hours}h {minutes}m"
    if hours:
        return f"{hours}h"
    return f"{minutes}m"


def historical_three_model_per_classifier_limit(
    n_classes: int,
    *,
    total_budget_seconds: int = BASELINE_HCESVM_TOTAL_BUDGET_SECONDS,
) -> int:
    """Recover the pre-override HCESVM per-classifier limit from the old runner."""
    return max(1, int(total_budget_seconds) // max(int(n_classes) - 1, 1))


def sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return a worksheet as records keyed by the header row."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []

    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    try:
        headers = next(iterator)
    except StopIteration:
        return []

    header_cells = ["" if cell is None else str(cell) for cell in headers]
    rows: list[dict[str, Any]] = []
    for row in iterator:
        rows.append(dict(zip(header_cells, row)))
    return rows


def metadata_map(workbook_path: Path) -> dict[str, Any]:
    """Return the workbook metadata sheet as a dictionary."""
    rows = sheet_rows(workbook_path, "Run_Metadata")
    metadata: dict[str, Any] = {}
    for row in rows:
        key = row.get("key")
        if key in (None, ""):
            continue
        metadata[str(key)] = row.get("value")
    return metadata


def load_deadline_run_workbook(workbook_path: Path) -> WorkbookRunData:
    """Load a deadline-run workbook into a normalized structure."""
    metric_rows = sheet_rows(workbook_path, "Metrics")
    metrics_row = metric_rows[0] if metric_rows else None
    return WorkbookRunData(
        path=workbook_path,
        metrics_row=metrics_row,
        parameter_rows=sheet_rows(workbook_path, "Parameters"),
        progress_rows=sheet_rows(workbook_path, "Progress"),
        metadata=metadata_map(workbook_path),
    )


def load_three_model_baseline(
    workbook_path: Path,
    *,
    dataset: str,
    model_name: str = "HCESVM(test3)",
) -> WorkbookRunData:
    """Extract one dataset/model slice from a three-model comparison workbook."""
    metrics_rows = sheet_rows(workbook_path, "Metrics")
    metric_row = next(
        (
            row
            for row in metrics_rows
            if row.get("dataset") == dataset and row.get("model") == model_name
        ),
        None,
    )
    if metric_row is None:
        raise ValueError(f"{dataset} / {model_name} not found in {workbook_path}")

    parameter_rows = [
        row
        for row in sheet_rows(workbook_path, "Parameters")
        if row.get("dataset") == dataset and row.get("model") == model_name
    ]
    return WorkbookRunData(
        path=workbook_path,
        metrics_row=metric_row,
        parameter_rows=parameter_rows,
        progress_rows=[],
        metadata=metadata_map(workbook_path),
    )


def latest_dataset_deadline_report(
    reports_dir: Path,
    *,
    dataset: str,
    started_after: datetime | None = None,
) -> Path | None:
    """Locate the newest deadline workbook for one dataset after an optional cutoff."""
    pattern = f"{dataset.upper()}_HCESVM_TEST3_DEADLINE_*.xlsx"
    candidates = sorted(reports_dir.glob(pattern))
    selected: list[tuple[datetime, Path]] = []
    for path in candidates:
        timestamp = parse_report_timestamp(path)
        if timestamp is None:
            continue
        if started_after is not None and timestamp < started_after:
            continue
        selected.append((timestamp, path))
    if not selected:
        return None
    selected.sort(key=lambda item: item[0])
    return selected[-1][1]


def latest_three_model_baseline_workbook(reports_dir: Path) -> Path:
    """Return the newest three-model comparison workbook."""
    candidates = sorted(reports_dir.glob("TEACHING_DATA_THREE_MODEL_COMPARISON_*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No TEACHING_DATA_THREE_MODEL_COMPARISON workbook found")
    timestamps = []
    for path in candidates:
        timestamp = parse_report_timestamp(path)
        if timestamp is not None:
            timestamps.append((timestamp, path))
    if not timestamps:
        return candidates[-1]
    timestamps.sort(key=lambda item: item[0])
    return timestamps[-1][1]


def class_accuracy_values(metrics_row: dict[str, Any], *, split: str, n_classes: int) -> list[float]:
    """Return per-class accuracy values for one split."""
    values: list[float] = []
    for class_index in range(1, n_classes + 1):
        value = metrics_row.get(f"{split}_class_{class_index}_accuracy")
        if value is not None:
            values.append(float(value))
    return values


def per_class_accuracy_records(metrics_row: dict[str, Any], *, split: str, n_classes: int) -> list[dict[str, Any]]:
    """Return one record per class with count and accuracy."""
    rows: list[dict[str, Any]] = []
    for class_index in range(1, n_classes + 1):
        rows.append(
            {
                "class_index": class_index,
                "count": metrics_row.get(f"{split}_class_{class_index}_count"),
                "accuracy": metrics_row.get(f"{split}_class_{class_index}_accuracy"),
            }
        )
    return rows


def mean_or_none(values: list[float]) -> float | None:
    """Return the mean of a list or None when empty."""
    if not values:
        return None
    return float(mean(values))


def zero_accuracy_count(metrics_row: dict[str, Any], *, split: str, n_classes: int) -> int:
    """Count how many classes have zero accuracy for one split."""
    values = class_accuracy_values(metrics_row, split=split, n_classes=n_classes)
    return sum(1 for value in values if abs(value) <= 1e-12)


def evaluate_significant_improvement(
    *,
    baseline_metrics: dict[str, Any],
    candidate_metrics: dict[str, Any],
    n_classes: int,
) -> tuple[bool, str, dict[str, Any]]:
    """Score whether the new run improved enough to justify more runtime."""
    delta_test_total = float(candidate_metrics["test_total_accuracy"]) - float(baseline_metrics["test_total_accuracy"])
    baseline_test_macro = mean_or_none(class_accuracy_values(baseline_metrics, split="test", n_classes=n_classes))
    candidate_test_macro = mean_or_none(class_accuracy_values(candidate_metrics, split="test", n_classes=n_classes))
    delta_test_macro = (
        None
        if baseline_test_macro is None or candidate_test_macro is None
        else float(candidate_test_macro) - float(baseline_test_macro)
    )
    baseline_zero_test = zero_accuracy_count(baseline_metrics, split="test", n_classes=n_classes)
    candidate_zero_test = zero_accuracy_count(candidate_metrics, split="test", n_classes=n_classes)
    zero_reduction = baseline_zero_test - candidate_zero_test

    improved_classes = 0
    largest_gain = None
    largest_drop = None
    for class_index in range(1, n_classes + 1):
        baseline_value = baseline_metrics.get(f"test_class_{class_index}_accuracy")
        candidate_value = candidate_metrics.get(f"test_class_{class_index}_accuracy")
        if baseline_value is None or candidate_value is None:
            continue
        delta = float(candidate_value) - float(baseline_value)
        if delta >= 0.05:
            improved_classes += 1
        largest_gain = delta if largest_gain is None else max(largest_gain, delta)
        largest_drop = delta if largest_drop is None else min(largest_drop, delta)

    reasons: list[str] = []
    if delta_test_total >= 0.02:
        reasons.append(f"test total accuracy +{delta_test_total:.4f}")
    if delta_test_macro is not None and delta_test_macro >= 0.03:
        reasons.append(f"test macro class accuracy +{delta_test_macro:.4f}")
    if zero_reduction >= 1 and delta_test_total >= -0.005:
        reasons.append(f"zero-accuracy test classes reduced by {zero_reduction}")
    if improved_classes >= 2 and delta_test_total >= 0:
        reasons.append(f"{improved_classes} test classes improved by at least 5pp")

    details = {
        "baseline_test_macro_accuracy": baseline_test_macro,
        "candidate_test_macro_accuracy": candidate_test_macro,
        "delta_test_macro_accuracy": delta_test_macro,
        "baseline_zero_test_classes": baseline_zero_test,
        "candidate_zero_test_classes": candidate_zero_test,
        "zero_test_classes_reduced": zero_reduction,
        "classes_improved_by_5pp": improved_classes,
        "largest_test_class_gain": largest_gain,
        "largest_test_class_drop": largest_drop,
    }
    return bool(reasons), "; ".join(reasons) if reasons else "no material test-set improvement", details


def recommend_time_limit_seconds(
    *,
    baseline_limit_seconds: int,
    current_limit_seconds: int | None,
    progress_rows: list[dict[str, Any]],
    final_status: str,
    significant_improvement: bool,
    delta_test_total_accuracy: float | None,
    delta_test_macro_accuracy: float | None,
    zero_test_classes_reduced: int,
) -> tuple[int, str]:
    """Recommend a per-classifier time limit that balances runtime and quality."""
    limit_now = baseline_limit_seconds if current_limit_seconds is None else int(current_limit_seconds)
    min_balanced_limit = max(int(baseline_limit_seconds), 900)

    elapsed_values = [
        float(row["elapsed_seconds"])
        for row in progress_rows
        if row.get("elapsed_seconds") not in (None, "")
    ]
    runtime_needed = min_balanced_limit
    if elapsed_values:
        runtime_needed = max(min_balanced_limit, round_up_seconds(max(elapsed_values) * 1.15))

    if final_status not in {"completed", "success"}:
        recommendation = min(limit_now, runtime_needed)
        reason = (
            "latest run did not finish with full metrics; keep the limit near the "
            "slowest completed classifier instead of pushing it higher"
        )
        return recommendation, reason

    if significant_improvement:
        recommendation = min(limit_now, runtime_needed)
        reason = (
            "accuracy improved materially; keep enough headroom for the slowest "
            "successful classifier with a small buffer"
        )
        return recommendation, reason

    modest_gain = any(
        value is not None and value > 0
        for value in (delta_test_total_accuracy, delta_test_macro_accuracy)
    ) or zero_test_classes_reduced > 0
    if modest_gain:
        midpoint = round_up_seconds((baseline_limit_seconds + limit_now) / 2.0)
        recommendation = max(min_balanced_limit, min(runtime_needed, midpoint))
        reason = (
            "there is some upside, but not enough to justify the full 1-hour limit; "
            "use a midpoint budget"
        )
        return recommendation, reason

    reason = (
        "no meaningful accuracy gain over the shorter baseline; use the shorter "
        "balanced limit"
    )
    return min_balanced_limit, reason
